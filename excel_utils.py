from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import pytz

# Buenos Aires timezone for displaying dates
BUENOS_AIRES_TZ = pytz.timezone('America/Argentina/Buenos_Aires')

def to_buenos_aires(dt):
    """Convert a datetime to Buenos Aires timezone for display."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = pytz.utc.localize(dt)
    return dt.astimezone(BUENOS_AIRES_TZ)

def generate_task_excel(tasks, filters):
    """
    Generate an Excel report for tasks with professional formatting.
    
    Args:
        tasks: List of Task objects to include in the report
        filters: Dictionary of applied filters
        
    Returns:
        Workbook object ready to be saved
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Tareas"
    
    # Brand Colors - Caja de Abogados
    brand_blue = "0077BE"  # RGB(0, 119, 190)
    brand_red = "C1272D"   # RGB(193, 39, 45)

    # Get area name from filters
    area_name = filters.get('area_name', 'Todas las áreas')

    # --- Header Section ---
    ws.merge_cells('A1:J1')
    header_cell = ws['A1']
    header_cell.value = "Gestor de Tareas"
    header_cell.font = Font(name='Arial', size=22, bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color=brand_blue, end_color=brand_blue, fill_type="solid")
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # --- Area Name ---
    ws.merge_cells('A2:J2')
    area_cell = ws['A2']
    area_cell.value = area_name
    area_cell.font = Font(name='Arial', size=14, bold=False, color="FFFFFF")
    area_cell.fill = PatternFill(start_color=brand_blue, end_color=brand_blue, fill_type="solid")
    area_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    # --- Date and Time ---
    ws.merge_cells('A3:J3')
    date_cell = ws['A3']
    date_cell.value = f"Generado el {to_buenos_aires(datetime.utcnow()).strftime('%d/%m/%Y %H:%M')}"
    date_cell.font = Font(name='Arial', size=10, italic=True)
    date_cell.alignment = Alignment(horizontal='center')
    
    # --- Summary Section ---
    row = 5
    total_tasks = len(tasks)
    completed_tasks = sum(1 for t in tasks if t.status == 'Completed')
    pending_tasks = total_tasks - completed_tasks
    
    ws.merge_cells(f'A{row}:H{row}')
    summary_cell = ws[f'A{row}']
    summary_cell.value = "Resumen Ejecutivo"
    summary_cell.font = Font(name='Arial', size=14, bold=True)
    summary_cell.alignment = Alignment(horizontal='left')
    row += 1
    
    # Summary stats
    summary_data = [
        ('Total de Tareas:', total_tasks),
        ('Pendientes:', pending_tasks),
        ('Completadas:', completed_tasks)
    ]
    
    for label, value in summary_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(name='Arial', size=11)
        row += 1
    
    row += 1
    
    # --- Filters Info ---
    ws.merge_cells(f'A{row}:H{row}')
    filters_cell = ws[f'A{row}']
    filters_cell.value = "Filtros Aplicados:"
    filters_cell.font = Font(name='Arial', size=12, bold=True)
    filters_cell.alignment = Alignment(horizontal='left')
    row += 1
    
    filter_text = []
    if filters.get('creator'):
        filter_text.append(f"Usuario: {filters['creator_name']}")
    if filters.get('status'):
        status_trans = "Completada" if filters['status'] == 'Completed' else "Pendiente"
        filter_text.append(f"Estado: {status_trans}")
    if filters.get('date_range'):
        filter_text.append(f"Fecha: {filters['date_range']}")
    if filters.get('tag'):
        filter_text.append(f"Etiqueta: {filters['tag']}")
    
    if not filter_text:
        ws[f'A{row}'] = "Ninguno (Mostrando todas las tareas)"
        ws[f'A{row}'].font = Font(name='Arial', size=10, italic=True)
        row += 1
    else:
        for ft in filter_text:
            ws[f'A{row}'] = ft
            ws[f'A{row}'].font = Font(name='Arial', size=10)
            row += 1
    
    row += 2
    
    # --- Table Headers ---
    headers = ['Título', 'Descripción', 'Estado', 'Prioridad', 'Vencimiento', 'Tiempo', 'Creado por', 'Asignados', 'Completado por', 'Fecha Completado']
    header_row = row
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=brand_blue, end_color=brand_blue, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    ws.row_dimensions[header_row].height = 30
    row += 1
    
    # --- Table Data ---
    for task in tasks:
        # Assignees
        assignees_list = ', '.join([a.full_name for a in task.assignees])
        
        # Completed by
        completed_by_name = task.completed_by.full_name if task.completed_by else '-'
        
        # Completed at
        completed_at_str = to_buenos_aires(task.completed_at).strftime('%d/%m/%Y %H:%M') if task.completed_at else '-'
        
        # Status translation
        status_text = "Completada" if task.status == 'Completed' else "Pendiente"
        
        # Time spent formatted
        if task.time_spent and task.time_spent > 0:
            if task.time_spent >= 60:
                time_str = f"{round(task.time_spent / 60, 1)} hs"
            else:
                time_str = f"{task.time_spent} min"
        else:
            time_str = '-'
        
        row_data = [
            task.title,
            task.description or '-',
            status_text,
            task.priority,
            task.due_date.strftime('%d/%m/%Y'),
            time_str,
            task.creator.full_name,
            assignees_list,
            completed_by_name,
            completed_at_str
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Special formatting for status
            if col_num == 3:  # Status column
                if task.status == 'Completed':
                    cell.font = Font(name='Arial', size=10, bold=True, color="047857")  # Green
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                else:
                    cell.font = Font(name='Arial', size=10, bold=True, color=brand_red)
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            
            # Zebra striping
            elif row % 2 == 0:
                cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        
        ws.row_dimensions[row].height = 40
        row += 1
    
    # --- Adjust column widths ---
    column_widths = [30, 40, 15, 12, 15, 12, 25, 20, 18, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    return wb


def generate_import_template():
    """
    Generates an empty Excel template for importing tasks.
    Includes headers and comments/instructions.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Importación"
    
    # Headers - Must match the order expected by import_tasks in routes.py
    # Columns: 0:Título, 1:Descripción, 2:Prioridad, 3:Fecha Inicio, 4:Hora Inicio,
    #          5:Fecha Vencimiento, 6:Hora Vencimiento, 7:Asignados, 8:Etiquetas,
    #          9:ID Proceso, 10:Estado, 11:Completado Por
    headers = [
        'Título (Requerido)', 
        'Descripción', 
        'Prioridad (Normal, Media, Urgente)', 
        'Fecha Inicio (DD/MM/YYYY)',
        'Hora Inicio (HH:MM)',
        'Fecha Vencimiento (Requerido, DD/MM/YYYY)', 
        'Hora Vencimiento (HH:MM)',
        'Asignados (Usuarios separados por coma)', 
        'Etiquetas (Separadas por coma)',
        'ID Proceso (Opcional)',
        'Estado (Pendiente/Completado)',
        'Completado Por (Usuario, si Estado=Completado)'
    ]
    
    # Header Style
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0077BE", end_color="0077BE", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=Side(style='thin'))
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_num)].width = 22
    
    # Set date/time columns (D, E, F, G) as text format to prevent Excel auto-conversion
    date_time_cols = [4, 5, 6, 7]  # Fecha Inicio, Hora Inicio, Fecha Vencimiento, Hora Vencimiento
    for col in date_time_cols:
        # Format the entire column as text (rows 2 to 1000 to cover typical usage)
        for r in range(2, 1001):
            ws.cell(row=r, column=col).number_format = '@'

    # Add example row
    example_row = [
        'Revisar Contrato',  # Título
        'Revisar cláusulas del contrato',  # Descripción
        'Normal',  # Prioridad
        '05/02/2026',  # Fecha Inicio
        '08:00',  # Hora Inicio
        '10/02/2026',  # Fecha Vencimiento
        '18:00',  # Hora Vencimiento
        'admin',  # Asignados
        'Legales',  # Etiquetas
        '',  # ID Proceso
        'Pendiente',  # Estado
        ''  # Completado Por
    ]
    for col_num, value in enumerate(example_row, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.font = Font(name='Arial', size=10, italic=True, color="888888")
        cell.alignment = Alignment(horizontal='center')
        # Ensure example date/time cells keep text format
        if col_num in date_time_cols:
            cell.number_format = '@'
    
    return wb


def parse_date_flexible(value):
    """
    Parse a date value from Excel that could be a datetime object or a string
    in various formats. Returns a datetime or None.
    """
    from datetime import date
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    # Try multiple string formats
    fecha_str = str(value).strip()
    for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
        try:
            return datetime.strptime(fecha_str, fmt)
        except ValueError:
            continue
    return None


def parse_time_flexible(value):
    """
    Parse a time value from Excel that could be a datetime/time object or a string.
    Returns (hour, minute) tuple or None.
    """
    from datetime import time as time_type
    if value is None:
        return None
    if isinstance(value, datetime):
        return (value.hour, value.minute)
    if isinstance(value, time_type):
        return (value.hour, value.minute)
    hora_str = str(value).strip()
    for fmt in ['%H:%M', '%H:%M:%S', '%I:%M %p']:
        try:
            parsed = datetime.strptime(hora_str, fmt)
            return (parsed.hour, parsed.minute)
        except ValueError:
            continue
    return None


def process_excel_import(file_stream, current_user, area_id=None):
    """
    Parses an uploaded Excel file and creates tasks.

    Columns must match generate_import_template():
    0: Título, 1: Descripción, 2: Prioridad, 3: Fecha Inicio, 4: Hora Inicio,
    5: Fecha Vencimiento, 6: Hora Vencimiento, 7: Asignados, 8: Etiquetas,
    9: ID Proceso, 10: Estado, 11: Completado Por

    Args:
        file_stream: Excel file stream
        current_user: The user performing the import
        area_id: Area ID to assign tasks to (from form selection)

    Returns:
        tuple: (success_count, error_list)
    """
    from openpyxl import load_workbook
    from models import Task, User, Tag, Process
    from extensions import db

    try:
        wb = load_workbook(file_stream)
        ws = wb.active
    except Exception as e:
        return 0, [f"Error al leer el archivo Excel: {str(e)}"]

    success_count = 0
    errors = []

    # Cache users and tags to avoid repeats
    all_users = {u.username.lower(): u for u in User.query.all()}
    for u in list(all_users.values()):
        all_users[u.full_name.lower()] = u

    all_tags = {t.name.lower(): t for t in Tag.query.all()}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not any(row):
            continue

        try:
            def get_col(idx):
                return row[idx] if idx < len(row) else None

            # Column mapping matching generate_import_template()
            title = get_col(0)
            description = get_col(1)
            priority = get_col(2)
            fecha_inicio_raw = get_col(3)
            hora_inicio_raw = get_col(4)
            fecha_vencimiento_raw = get_col(5)
            hora_vencimiento_raw = get_col(6)
            assignees_raw = get_col(7)
            tags_raw = get_col(8)
            process_id_raw = get_col(9)
            status_raw = get_col(10)
            completed_by_raw = get_col(11)

            # Validation: Title
            if not title:
                errors.append(f"Fila {row_idx}: Falta el Título (Requerido).")
                continue

            # Validation & Parse: Due Date (required)
            if not fecha_vencimiento_raw:
                errors.append(f"Fila {row_idx}: Falta Fecha Vencimiento (Requerido).")
                continue

            due_date = parse_date_flexible(fecha_vencimiento_raw)
            if not due_date:
                errors.append(f"Fila {row_idx}: Formato de Fecha Vencimiento inválido. Use DD/MM/YYYY o YYYY-MM-DD.")
                continue

            # Apply time to due date
            hora_venc = parse_time_flexible(hora_vencimiento_raw)
            if hora_venc:
                due_date = due_date.replace(hour=hora_venc[0], minute=hora_venc[1])
            else:
                due_date = due_date.replace(hour=18, minute=0)  # Default end of workday

            # Parse Start Date (optional)
            start_date = None
            if fecha_inicio_raw:
                start_date = parse_date_flexible(fecha_inicio_raw)
                if not start_date:
                    errors.append(f"Fila {row_idx}: Formato de Fecha Inicio inválido. Se ignoró.")
                else:
                    hora_ini = parse_time_flexible(hora_inicio_raw)
                    if hora_ini:
                        start_date = start_date.replace(hour=hora_ini[0], minute=hora_ini[1])
                    else:
                        start_date = start_date.replace(hour=8, minute=0)  # Default start of workday

            # Priority
            valid_priorities = ['Normal', 'Media', 'Urgente']
            if priority and str(priority).strip().capitalize() in valid_priorities:
                priority = str(priority).strip().capitalize()
            else:
                priority = 'Normal'

            # Create Task
            new_task = Task(
                title=str(title),
                description=str(description) if description else '',
                priority=priority,
                due_date=due_date,
                planned_start_date=start_date,
                creator_id=current_user.id,
                status='Pending',
                area_id=area_id or (current_user.areas[0].id if current_user.areas else None)
            )

            # Process ID
            if process_id_raw:
                try:
                    pid = int(process_id_raw)
                    process = Process.query.get(pid)
                    if process:
                        new_task.process_id = process.id
                        new_task.area_id = process.area_id
                    else:
                        errors.append(f"Fila {row_idx}: Proceso ID {pid} no encontrado. Se creó sin proceso.")
                except (ValueError, TypeError):
                    errors.append(f"Fila {row_idx}: ID de Proceso inválido.")

            # Assignees
            if assignees_raw:
                names = [n.strip().lower() for n in str(assignees_raw).split(',')]
                for name in names:
                    if name in all_users:
                        new_task.assignees.append(all_users[name])

            # Tags
            if tags_raw:
                tag_names = [t.strip().lower() for t in str(tags_raw).split(',')]
                for t_name in tag_names:
                    if t_name in all_tags:
                        new_task.tags.append(all_tags[t_name])

            # Status & Completion
            if status_raw:
                status_map = {
                    'pendiente': 'Pending',
                    'pending': 'Pending',
                    'completado': 'Completed',
                    'completed': 'Completed',
                    'completada': 'Completed'
                }
                final_status = status_map.get(str(status_raw).strip().lower(), 'Pending')
                new_task.status = final_status

                if final_status == 'Completed':
                    new_task.completed_at = datetime.utcnow()
                    new_task.completed_by_id = current_user.id

                    if completed_by_raw:
                        c_name = str(completed_by_raw).strip().lower()
                        if c_name in all_users:
                            new_task.completed_by_id = all_users[c_name].id

            db.session.add(new_task)
            success_count += 1

        except Exception as e:
            errors.append(f"Fila {row_idx}: Error inesperado - {str(e)}")
            continue

    try:
        if success_count > 0:
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        return 0, [f"Error de base de datos al guardar: {str(e)}"]

    return success_count, errors
