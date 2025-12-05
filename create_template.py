from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# Main sheet with template
ws = wb.active
ws.title = 'Tareas'

# Header style
header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers - Added Etiquetas column
headers = ['Titulo', 'Descripcion', 'Prioridad', 'Fecha Vencimiento', 'Asignados', 'Etiquetas']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Example rows with tags
examples = [
    ['Revisar expediente 1234', 'Verificar documentacion completa del caso', 'Normal', '2024-12-20', 'admin', 'Urgente, Legal'],
    ['Preparar informe mensual', 'Elaborar informe de actividades del mes', 'Media', '2024-12-15', 'admin', 'Administrativo'],
    ['Audiencia caso Smith', 'Preparar alegatos para audiencia', 'Urgente', '2024-12-10', 'admin', 'Legal, Importante, Tribunal'],
]

for row_idx, row_data in enumerate(examples, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

# Adjust column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 30

# Instructions sheet
ws2 = wb.create_sheet('Instrucciones')
instructions = [
    'INSTRUCCIONES PARA IMPORTAR TAREAS',
    '',
    '1. Complete la hoja "Tareas" con los datos de las tareas a importar.',
    '',
    '2. Columnas requeridas:',
    '   - Titulo: Nombre de la tarea (obligatorio)',
    '   - Descripcion: Detalle de la tarea (opcional)',
    '   - Prioridad: Normal, Media o Urgente (obligatorio)',
    '   - Fecha Vencimiento: Formato AAAA-MM-DD, ej: 2024-12-20 (obligatorio)',
    '   - Asignados: Username(s) separados por coma, ej: admin, usuario1 (obligatorio)',
    '   - Etiquetas: Nombre(s) de etiquetas separadas por coma (opcional, max 3)',
    '',
    '3. Importante:',
    '   - No modifique los encabezados de las columnas',
    '   - Los usernames deben existir en el sistema',
    '   - Las etiquetas deben existir en el sistema (se ignoran las que no existan)',
    '   - La fecha debe estar en formato AAAA-MM-DD',
    '   - Las prioridades validas son: Normal, Media, Urgente',
]

for row_idx, text in enumerate(instructions, 1):
    cell = ws2.cell(row=row_idx, column=1, value=text)
    if row_idx == 1:
        cell.font = Font(bold=True, size=14)
    elif text.startswith('   -'):
        cell.font = Font(italic=True)

ws2.column_dimensions['A'].width = 75

wb.save('static/plantilla_tareas.xlsx')
print('Template created successfully!')

