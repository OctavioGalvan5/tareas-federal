from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, make_response
from flask_login import login_user, logout_user, login_required, current_user
from extensions import db
from models import User, Task, Tag, TaskTemplate, SubtaskTemplate, Expiration, RecurringTask, ActivityLog, ProcessType, Process, StatusTransition, TaskAttachment
from datetime import datetime, date, timedelta, time
from pdf_utils import generate_task_pdf
from excel_utils import generate_task_excel, generate_import_template, process_excel_import
from io import BytesIO
from utils import calculate_business_days_until
from sqlalchemy.orm import joinedload, subqueryload
import pytz
from werkzeug.utils import secure_filename
import storage
import json

# Buenos Aires timezone (for reference, conversion is done in templates)
BUENOS_AIRES_TZ = pytz.timezone('America/Argentina/Buenos_Aires')

def now_utc():
    """Get current datetime in UTC (consistent with other datetimes in the app)."""
    return datetime.utcnow()


def create_subtasks_from_template(template, parent_task, assignees, creator, area_id):
    """
    Recursively create subtasks from a template's subtask hierarchy.
    
    Args:
        template: TaskTemplate object
        parent_task: The parent Task that was just created
        assignees: List of User objects to assign to subtasks
        creator: User who is creating the tasks
        area_id: Area ID for the subtasks
    """
    from models import SubtaskTemplate
    
    def create_children(subtask_templates, parent_task_obj):
        """Recursively create child tasks from subtask templates."""
        for st in subtask_templates:
            # Calculate due date based on parent + offset
            subtask_due_date = parent_task_obj.due_date + timedelta(days=st.days_offset)
            
            # Create the subtask
            subtask = Task(
                title=st.title,
                description=st.description or f'Subtarea de: {parent_task_obj.title}',
                priority=st.priority,
                status='Pending',
                due_date=subtask_due_date,
                created_at=now_utc(),
                creator_id=creator.id,
                area_id=area_id,
                parent_id=parent_task_obj.id,
                enabled=parent_task_obj.status == 'Completed'  # Blocked if parent not done
            )
            
            # Assign users
            for user in assignees:
                subtask.assignees.append(user)
            
            db.session.add(subtask)
            db.session.flush()  # Get ID for nested children
            
            # Recursively create children of this subtask
            children = SubtaskTemplate.query.filter_by(
                template_id=template.id,
                parent_id=st.id
            ).order_by(SubtaskTemplate.order).all()
            
            if children:
                create_children(children, subtask)
    
    # Get top-level subtasks (no parent)
    top_level = SubtaskTemplate.query.filter_by(
        template_id=template.id,
        parent_id=None
    ).order_by(SubtaskTemplate.order).all()
    
    create_children(top_level, parent_task)


def log_activity(user, action, description, target_type=None, target_id=None, area_id=None, details=None):
    """
    Registra una actividad en el log de auditoría.
    
    Args:
        user: Usuario que realizó la acción
        action: Tipo de acción ('task_created', 'task_completed', etc.)
        description: Descripción legible ("creó la tarea #123")
        target_type: Tipo de objeto afectado ('task', 'expiration', etc.)
        target_id: ID del objeto afectado
        area_id: ID del área asociada (para filtrado de supervisores)
        details: JSON string con detalles adicionales (difs, comentarios)
    """
    try:
        log = ActivityLog(
            user_id=user.id,
            action=action,
            description=description,
            target_type=target_type,
            target_id=target_id,
            area_id=area_id,
            details=details,
            created_at=datetime.utcnow()
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        print(f"Error logging activity: {e}")
        db.session.rollback()


def log_process_event(process_id, event_type, description, user_id=None, task_id=None, extra_data=None):
    """
    Registra un evento en el historial del proceso.
    
    Args:
        process_id: ID del proceso
        event_type: Tipo de evento ('task_created', 'task_completed', 'transfer', 'status_change')
        description: Descripción legible del evento
        user_id: ID del usuario que realizó la acción (opcional)
        task_id: ID de la tarea relacionada (opcional)
        extra_data: Datos adicionales en formato JSON (opcional)
    """
    try:
        from models import ProcessEvent
        event = ProcessEvent(
            process_id=process_id,
            event_type=event_type,
            description=description,
            user_id=user_id,
            task_id=task_id,
            extra_data=extra_data
        )
        db.session.add(event)
        # Note: Don't commit here, let the caller handle the commit
    except Exception as e:
        print(f"Error logging process event: {e}")


main_bp = Blueprint('main', __name__)
auth_bp = Blueprint('auth', __name__)
admin_bp = Blueprint('admin', __name__)

# --- Auth Routes ---
@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.scrum_board'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        print(f"DEBUG: Login attempt for username='{username}', password='{password}'")
        
        user = User.query.filter_by(username=username).first()
        print(f"DEBUG: User found: {user}")
        
        if user:
            check = user.check_password(password)
            print(f"DEBUG: Password check result: {check}")
            if check:
                login_user(user)
                
                # Log activity
                log_activity(
                    user=user,
                    action='login',
                    description='inició sesión',
                    area_id=user.areas[0].id if user.areas else None
                )
                
                return redirect(url_for('main.scrum_board'))
        
        print("DEBUG: Login failed")
        flash('Usuario o contraseña incorrectos.', 'danger')
            
    return render_template('login.html')

@auth_bp.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('auth.login'))

# --- Main Routes ---
# --- Main Routes ---
@main_bp.route('/')
@login_required
def index():
    return redirect(url_for('main.scrum_board'))

@main_bp.route('/dashboard')
@login_required
def dashboard():
    # Import Area model
    from models import Area
    
    # Filter logic
    filter_assignee = request.args.get('assignee')
    filter_creator = request.args.get('creator')
    filter_status = request.args.get('status')
    filter_tag = request.args.get('tag_filter')
    filter_area = request.args.get('area')  # NEW: Area filter
    show_blocked = request.args.get('show_blocked', 'true')  # NEW: Show blocked tasks (default: true)
    sort_order = request.args.get('sort', 'asc')

    # Base query: Eager load assignees, tags, and area
    tasks_query = Task.query.options(
        joinedload(Task.assignees), 
        joinedload(Task.tags),
        joinedload(Task.area),  # NEW: Load area relationship
        joinedload(Task.parent)  # Load parent for blocked tasks
    )
    
    # Filter by enabled status (show blocked tasks if filter is on)
    if show_blocked == 'true':
        # Show all tasks (both enabled and disabled)
        pass  # No filter on enabled status
    else:
        # Only show enabled tasks
        tasks_query = tasks_query.filter(Task.enabled == True)
    
    # NEW: Role-based visibility filtering
    # - usuario/usuario_plus: only see tasks assigned to them OR created by them (within their areas)
    # - supervisor: see ALL tasks in their area
    # - gerente/admin: see ALL tasks
    if current_user.can_only_see_own_tasks():
        # Users can only see tasks they are assigned to OR tasks they created
        user_area_ids = [area.id for area in current_user.areas]
        if user_area_ids:
            tasks_query = tasks_query.filter(
                db.or_(
                    Task.assignees.any(id=current_user.id),
                    Task.creator_id == current_user.id
                )
            ).filter(Task.area_id.in_(user_area_ids))
        else:
            # User has no areas assigned - show no tasks
            tasks_query = tasks_query.filter(Task.area_id == -1)
    elif not current_user.can_see_all_areas():
        # Supervisors see all tasks in their area
        user_area_ids = [area.id for area in current_user.areas]
        if user_area_ids:
            tasks_query = tasks_query.filter(Task.area_id.in_(user_area_ids))
        else:
            tasks_query = tasks_query.filter(Task.area_id == -1)
    # Gerentes and admins see all tasks (no additional filter)
    
    # Additional area filter (for gerentes/supervisors filtering specific area)
    if filter_area:
        tasks_query = tasks_query.filter(Task.area_id == int(filter_area))
    
    if filter_assignee:
        tasks_query = tasks_query.filter(Task.assignees.any(id=filter_assignee))
    
    if filter_creator:
        tasks_query = tasks_query.filter(Task.creator_id == filter_creator)
        
    # Handle status filter - exclude 'Anulado' by default
    if filter_status:
        if filter_status == 'Overdue':
            today_date = date.today()
            # Include Pending, In Progress, and In Review tasks that are overdue
            tasks_query = tasks_query.filter(
                Task.status.in_(['Pending', 'In Progress', 'In Review']),
                db.func.date(Task.due_date) < today_date
            )
        elif filter_status in ['Pending', 'In Progress', 'In Review', 'Completed', 'Anulado', 'Scheduled']:
            tasks_query = tasks_query.filter(Task.status == filter_status)
    else:
        tasks_query = tasks_query.filter(Task.status != 'Anulado')

    if filter_tag:
        tasks_query = tasks_query.filter(Task.tags.any(id=int(filter_tag)))

    # Search filter
    search_query = request.args.get('q')
    if search_query:
        search_term = f"%{search_query}%"
        tasks_query = tasks_query.filter(
            (Task.title.ilike(search_term)) | 
            (Task.description.ilike(search_term))
        )
        
    # Sort by status first, then by due_date
    # Priority: In Progress > In Review > Pending > Completed (active work first)
    
    # NEW: Date/Period Filter Implementation (Matching Scrum Board Logic)
    filter_period = request.args.get('period', 'week') # Default to week as requested
    filter_date_from = request.args.get('date_from')
    filter_date_to = request.args.get('date_to')
    
    today = date.today()
    
    # Overdue condition: Pending/InProgress/Review tasks past due date
    # These should ALWAYS be shown regardless of date filter (unless completed)
    overdue_condition = db.and_(
        db.func.date(Task.due_date) < today,
        Task.status.in_(['Pending', 'In Progress', 'In Review'])
    )
    
    if filter_period == 'today':
        tasks_query = tasks_query.filter(
            db.or_(
                db.func.date(Task.due_date) == today,
                db.func.date(Task.planned_start_date) == today,
                Task.status.in_(['In Progress', 'In Review']), # Always show active work
                overdue_condition,
                # Show completed tasks if completed TODAY
                db.and_(Task.status == 'Completed', db.func.date(Task.completed_at) == today)
            )
        )
    elif filter_period == 'week':
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        tasks_query = tasks_query.filter(
            db.or_(
                db.and_(db.func.date(Task.due_date) >= week_start, db.func.date(Task.due_date) <= week_end),
                db.and_(db.func.date(Task.planned_start_date) >= week_start, db.func.date(Task.planned_start_date) <= week_end),
                Task.status.in_(['In Progress', 'In Review']),
                overdue_condition,
                # Show completed tasks if completed THIS WEEK
                db.and_(Task.status == 'Completed', db.func.date(Task.completed_at) >= week_start, db.func.date(Task.completed_at) <= week_end)
            )
        )
    elif filter_period == 'month':
        month_start = today.replace(day=1)
        if today.month == 12:
            month_end = today.replace(day=31)
        else:
            month_end = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        tasks_query = tasks_query.filter(
            db.or_(
                db.and_(db.func.date(Task.due_date) >= month_start, db.func.date(Task.due_date) <= month_end),
                db.and_(db.func.date(Task.planned_start_date) >= month_start, db.func.date(Task.planned_start_date) <= month_end),
                Task.status.in_(['In Progress', 'In Review']),
                overdue_condition,
                # Show completed tasks if completed THIS MONTH
                db.and_(Task.status == 'Completed', db.func.date(Task.completed_at) >= month_start, db.func.date(Task.completed_at) <= month_end)
            )
        )
    elif filter_period == 'custom':
        custom_conditions = [overdue_condition]
        if filter_date_from:
            try:
                date_from_obj = datetime.strptime(filter_date_from, '%Y-%m-%d').date()
                custom_conditions.append(db.func.date(Task.due_date) >= date_from_obj)
            except ValueError:
                pass
        if filter_date_to:
            try:
                date_to_obj = datetime.strptime(filter_date_to, '%Y-%m-%d').date()
                custom_conditions.append(db.func.date(Task.due_date) <= date_to_obj)
            except ValueError:
                pass
        
        # Combine conditions: (Overdue) OR (Range Match)
        if len(custom_conditions) > 1:
            # We have at least one valid range condition plus overdue
            range_condition = db.and_(*custom_conditions[1:])
            tasks_query = tasks_query.filter(db.or_(overdue_condition, range_condition))
        else:
            # Only overdue condition remains if dates were invalid
            tasks_query = tasks_query.filter(overdue_condition)
    status_order = db.case(
        (Task.status == 'In Progress', 0),
        (Task.status == 'In Review', 1),
        (Task.status == 'Pending', 2),
        (Task.status == 'Completed', 3),
        else_=4
    )
    
    if sort_order == 'desc':
        tasks = tasks_query.order_by(status_order, Task.due_date.desc()).all()
    else:
        tasks = tasks_query.order_by(status_order, Task.due_date.asc()).all()
    
    today = date.today()
    now = datetime.now()  # Current datetime for time-based overdue checking
    
    # Filter users by area for non-admins
    if current_user.is_admin or current_user.can_see_all_areas():
        users = User.query.all()
        all_areas = Area.query.order_by(Area.name).all()
        all_tags = Tag.query.order_by(Tag.name).all()
    else:
        # Non-admins see only users and tags in their areas
        users = [u for u in User.query.all() if any(area in u.areas for area in current_user.areas)]
        all_areas = current_user.areas
        user_area_ids = [a.id for a in current_user.areas]
        if user_area_ids:
            # Strict area filter - only show tags from user's areas
            all_tags = Tag.query.filter(Tag.area_id.in_(user_area_ids)).order_by(Tag.name).all()
        else:
            all_tags = []
    
    return render_template('dashboard.html', 
        tasks=tasks, 
        today=today,
        now=now,  # NEW: Pass current datetime for time-based overdue checking
        users=users, 
        all_tags=all_tags, 
        sort_order=sort_order,
        all_areas=all_areas,  # NEW: Pass areas to template
        current_user_is_gerente=current_user.can_see_all_areas(),  # NEW
        show_blocked=show_blocked,  # NEW: Pass blocked filter state
        filter_period=filter_period,
        filter_date_from=filter_date_from,
        filter_date_to=filter_date_to,
        filter_status=filter_status, # Also pass other filters back to maintain state
        filter_tag=filter_tag,
        filter_assignee=filter_assignee,
        filter_area=filter_area,
        filter_creator=filter_creator,
    )

@main_bp.route('/task/new', methods=['GET', 'POST'])
@login_required
def create_task():
    # Import Area model
    from models import Area
    
    # --- CHECK PERMISSION TO CREATE TASKS ---
    if not current_user.can_create_tasks():
        flash('No tienes permiso para crear tareas. Usa el calendario de vencimientos para crear recordatorios.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Check if creating a subtask (parent_id passed via query param)
    parent_task = None
    parent_id_param = request.args.get('parent_id')
    if parent_id_param:
        try:
            parent_task = Task.query.get(int(parent_id_param))
        except (ValueError, TypeError):
            parent_task = None
    
    # Check if linking to a process (process_id passed via query param)
    preselected_process = None
    process_id_param = request.args.get('process_id')
    if process_id_param:
        try:
            preselected_process = Process.query.get(int(process_id_param))
        except (ValueError, TypeError):
            preselected_process = None
    
    # Load available areas based on user role
    # Admins see all, others see only their area(s)
    if current_user.is_admin:
        available_areas = Area.query.order_by(Area.name).all()
        users = User.query.all()  # Admins see all users
        available_tags = Tag.query.order_by(Tag.name).all()
        templates = TaskTemplate.query.order_by(TaskTemplate.name).all()
    else:
        # Supervisor/usuario_plus: only their assigned area(s)
        available_areas = current_user.areas[:1] if current_user.areas else []
        # Only see users in their area
        if available_areas:
            users = [u for u in User.query.all() if any(area in u.areas for area in available_areas)]
            # Filter tags and templates by area
            user_area_ids = [a.id for a in available_areas]
            available_tags = Tag.query.filter(Tag.area_id.in_(user_area_ids)).order_by(Tag.name).all()
            templates = TaskTemplate.query.filter(TaskTemplate.area_id.in_(user_area_ids)).order_by(TaskTemplate.name).all()
        else:
            users = []
            available_tags = []

            templates = []

    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        priority = request.form.get('priority')
        
        # Parse dates and times
        start_date_str = request.form.get('start_date')
        start_time_str = request.form.get('start_time', '08:00')
        due_date_str = request.form.get('due_date')
        due_time_str = request.form.get('due_time', '14:00')
        
        time_spent_str = request.form.get('time_spent')
        area_id_str = request.form.get('area_id')  # NEW: Get area from form
        
        # Combine due_date + due_time into datetime
        due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
        if due_time_str:
            try:
                due_time = datetime.strptime(due_time_str, '%H:%M').time()
                due_date = datetime.combine(due_date.date(), due_time)
            except ValueError:
                due_date = datetime.combine(due_date.date(), datetime.strptime('14:00', '%H:%M').time())
        
        # Combine start_date + start_time into datetime (optional)
        planned_start_date = None
        if start_date_str:
            planned_start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            if start_time_str:
                try:
                    start_time = datetime.strptime(start_time_str, '%H:%M').time()
                    planned_start_date = datetime.combine(planned_start_date.date(), start_time)
                except ValueError:
                    planned_start_date = datetime.combine(planned_start_date.date(), datetime.strptime('08:00', '%H:%M').time())
        
        # Parse time_spent (in minutes)
        time_spent = None
        if time_spent_str and time_spent_str.strip():
            try:
                time_spent = int(time_spent_str)
            except ValueError:
                time_spent = None
        
        # Parse area_id
        area_id = None
        if area_id_str and area_id_str.strip():
            try:
                area_id = int(area_id_str)
            except ValueError:
                area_id = None
        
        # If no area selected and user has only one area, use that one
        if area_id is None and len(available_areas) == 1:
            area_id = available_areas[0].id
        
        # Validate: planned_start_date must be before due_date
        if planned_start_date and planned_start_date >= due_date:
            flash('La fecha/hora de inicio debe ser anterior a la fecha/hora de vencimiento.', 'danger')
            return redirect(url_for('main.create_task'))
        
        # Process task creation
        assignee_ids = request.form.getlist('assignees')
        
        # Determine initial status based on planned_start_date
        # If start datetime is in the future, the task is "Scheduled" (Programada)
        initial_status = 'Pending'
        if planned_start_date:
            now = datetime.now()
            if planned_start_date > now:
                initial_status = 'Scheduled'
        
        new_task = Task(
            title=title,
            description=description,
            priority=priority,
            status=initial_status,
            planned_start_date=planned_start_date,
            due_date=due_date,
            creator_id=current_user.id,
            time_spent=time_spent,
            area_id=area_id  # NEW: Assign area
        )
        
        # Handle process_id from form
        process_id_str = request.form.get('process_id')
        if process_id_str and process_id_str.strip():
            try:
                process_id = int(process_id_str)
                process = Process.query.get(process_id)
                if process:
                    new_task.process_id = process_id
            except (ValueError, TypeError):
                pass
        
        # Add assignees
        for user_id in assignee_ids:
            user = User.query.get(int(user_id))
            if user:
                new_task.assignees.append(user)

        # Add tags
        tag_ids = request.form.getlist('tags')
        for tag_id in tag_ids:
            tag = Tag.query.get(int(tag_id))
            if tag:
                new_task.tags.append(tag)
        
        # Handle parent task selection and dependency blocking
        parent_id_str = request.form.get('parent_id')
        if parent_id_str:
            try:
                parent_id = int(parent_id_str)
                parent_task = Task.query.get(parent_id)
                if parent_task:
                    new_task.parent_id = parent_id
                    # If parent is NOT completed, the child starts as blocked
                    if parent_task.status != 'Completed':
                        new_task.enabled = False
                        new_task.enabled_at = None
                    else:
                        # Parent is already completed, child starts enabled
                        new_task.enabled = True
                        new_task.enabled_at = now_utc()
                        new_task.enabled_by_task_id = parent_id
            except (ValueError, TypeError):
                pass  # Invalid parent_id, ignore it
                
        db.session.add(new_task)
        db.session.commit()
        
        # Log process event if part of a process
        if new_task.process_id:
            log_process_event(
                process_id=new_task.process_id,
                event_type='task_created',
                description=f'Tarea creada: {new_task.title}',
                user_id=current_user.id,
                task_id=new_task.id
            )
        
        # Process inline subtasks (if any) - with hierarchy support
        subtask_titles = request.form.getlist('subtask_title[]')
        subtask_descriptions = request.form.getlist('subtask_description[]')
        subtask_priorities = request.form.getlist('subtask_priority[]')
        subtask_assignees = request.form.getlist('subtask_assignee[]')
        subtask_assignees = request.form.getlist('subtask_assignee[]')
        subtask_due_dates = request.form.getlist('subtask_due_date[]')
        subtask_due_times = request.form.getlist('subtask_due_time[]')
        subtask_start_dates = request.form.getlist('subtask_start_date[]')
        subtask_start_times = request.form.getlist('subtask_start_time[]')
        subtask_parent_paths = request.form.getlist('subtask_parent_path[]')
        
        # First pass: create all subtasks and store by index
        index_to_subtask = {}
        subtasks_created = 0
        
        for i, title in enumerate(subtask_titles):
            if title.strip():  # Only create if title is not empty
                # Get description (or default)
                description = ''
                if i < len(subtask_descriptions) and subtask_descriptions[i]:
                    description = subtask_descriptions[i].strip()
                else:
                    description = f'Subtarea de: {new_task.title}'
                
                # Get priority (or inherit from parent)
                priority = new_task.priority
                if i < len(subtask_priorities) and subtask_priorities[i]:
                    priority = subtask_priorities[i]
                
                # Get due_date and due_time (or inherit from parent)
                subtask_due_date = due_date  # Default to parent's due_date
                if i < len(subtask_due_dates) and subtask_due_dates[i]:
                    try:
                        subtask_date_str = subtask_due_dates[i]
                        subtask_time_str = subtask_due_times[i] if i < len(subtask_due_times) and subtask_due_times[i] else '17:00'
                        subtask_due_date = datetime.strptime(f"{subtask_date_str} {subtask_time_str}", "%Y-%m-%d %H:%M")
                    except ValueError:
                        subtask_due_date = due_date  # Fallback to parent's
                
                # Get planned_start_date (or inherit from parent)
                subtask_planned_start = new_task.planned_start_date # Default inheretance
                if i < len(subtask_start_dates) and subtask_start_dates[i]:
                    try:
                        subtask_sdate_str = subtask_start_dates[i]
                        subtask_stime_str = subtask_start_times[i] if i < len(subtask_start_times) and subtask_start_times[i] else '08:00'
                        subtask_planned_start = datetime.strptime(f"{subtask_sdate_str} {subtask_stime_str}", "%Y-%m-%d %H:%M")
                    except ValueError:
                        pass # Keep default inheritance

                # Get parent path to determine if this is a child of another subtask
                parent_path = subtask_parent_paths[i] if i < len(subtask_parent_paths) else ''
                
                # For first pass, set parent_id to main task (will be updated in second pass)
                subtask = Task(
                    title=title.strip(),
                    description=description,
                    priority=priority,
                    due_date=subtask_due_date,
                    status='Pending',  # Subtasks start as Pending but disabled
                    creator_id=current_user.id,
                    area_id=new_task.area_id,  # Inherit area
                    parent_id=new_task.id,  # Default to main task (may be changed)
                    enabled=False,  # Start disabled (blocked by parent)
                    planned_start_date=subtask_planned_start,  # Use calculated or inherited start date
                    process_id=new_task.process_id  # Inherit process from parent
                )
                
                # Assign user if specified
                if i < len(subtask_assignees) and subtask_assignees[i]:
                    try:
                        assignee = User.query.get(int(subtask_assignees[i]))
                        if assignee:
                            subtask.assignees.append(assignee)
                    except (ValueError, TypeError):
                        pass
                
                # Inherit tags from parent
                for tag in new_task.tags:
                    subtask.tags.append(tag)
                
                db.session.add(subtask)
                db.session.flush()  # Get ID for linking
                
                # Store subtask by its 1-based index from form (matching frontend counter)
                # Extract the last number from the path to use as key
                if parent_path:
                    current_path = f"{parent_path}.{i+1}"
                else:
                    current_path = str(i+1)
                index_to_subtask[current_path] = subtask
                index_to_subtask[str(i+1)] = subtask  # Also store by simple index
                subtasks_created += 1
        
        # Second pass: link child subtasks to their parent subtasks
        for i, title in enumerate(subtask_titles):
            if title.strip():
                parent_path = subtask_parent_paths[i] if i < len(subtask_parent_paths) else ''
                if parent_path:
                    # This subtask has a parent subtask (not just the main task)
                    # Find the parent subtask by path
                    if parent_path in index_to_subtask:
                        if parent_path:
                            current_path = f"{parent_path}.{i+1}"
                        else:
                            current_path = str(i+1)
                        
                        if current_path in index_to_subtask:
                            # Update the parent_id to point to the parent subtask
                            index_to_subtask[current_path].parent_id = index_to_subtask[parent_path].id
        
        if subtasks_created > 0:
            db.session.commit()
        
        # NOTE: Subtasks from template are already handled by the frontend
        # which populates the subtask form fields when a template is selected.
        # No need to call create_subtasks_from_template here as it would create duplicates.
        
        # Log activity
        log_activity(
            user=current_user,
            action='task_created',
            description=f'creó la tarea "{new_task.title}"' + (f' con {subtasks_created} subtarea(s)' if subtasks_created > 0 else ''),
            target_type='task',
            target_id=new_task.id,
            area_id=new_task.area_id
        )
        
        if subtasks_created > 0:
            flash(f'Tarea creada con {subtasks_created} subtarea(s).', 'success')
        else:
            flash('Tarea creada exitosamente.', 'success')
        return redirect(url_for('main.dashboard'))
        
    # Get available processes for the user
    user_area_ids = [a.id for a in current_user.areas]
    if current_user.is_admin:
        available_processes = Process.query.filter_by(status='Active').order_by(Process.name).all()
    elif user_area_ids:
        available_processes = Process.query.filter(
            Process.area_id.in_(user_area_ids),
            Process.status == 'Active'
        ).order_by(Process.name).all()
    else:
        available_processes = []
        
    return render_template('create_task.html', users=users, available_tags=available_tags, templates=templates, available_areas=available_areas, parent_task=parent_task, available_processes=available_processes, preselected_process=preselected_process)

@main_bp.route('/task/<int:task_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_task(task_id):
    from models import Area
    
    task = Task.query.get_or_404(task_id)
    
    # Verify user has access to this task (either creator or assignee)
    # Ideally only creator or admin should edit, or maybe assignees too?
    # For now let's allow creator and assignees to edit
    if task.creator_id != current_user.id and current_user not in task.assignees and not current_user.is_admin and current_user.role != 'supervisor':
        flash('No tienes permiso para editar esta tarea.', 'danger')
        return redirect(url_for('main.dashboard'))

    # Load available areas based on user role
    if current_user.is_admin or current_user.role == 'gerente':
        users = User.query.all()
        available_tags = Tag.query.order_by(Tag.name).all()
        available_areas = Area.query.order_by(Area.name).all()
    else:
        # Filter users by area (assignees)
        users = [u for u in User.query.all() if any(area in u.areas for area in current_user.areas)]
        
        # Filter tags by area
        user_area_ids = [a.id for a in current_user.areas]
        if user_area_ids:
             available_tags = Tag.query.filter(Tag.area_id.in_(user_area_ids)).order_by(Tag.name).all()
        else:
             available_tags = []
        # Non-admins can only see their own areas
        available_areas = current_user.areas

    if request.method == 'POST':
        # Snapshot state before changes for diff
        old_state = {
            'title': task.title,
            'description': task.description,
            'priority': task.priority,
            'status': task.status,
            'due_date': task.due_date,
            'area_id': task.area_id,
            'planned_start_date': task.planned_start_date
        }

        task.title = request.form.get('title')
        task.description = request.form.get('description')
        
        # Update status - ONLY if user is admin or if status is actually provided in form
        new_status = request.form.get('status')
        if new_status:  # Only update if a value was provided (prevents None)
            task.status = new_status
        
        # Update priority and due_date - Admin and Supervisors
        if current_user.is_admin or current_user.role == 'supervisor':
            priority = request.form.get('priority')
            if priority:  # Only update if provided
                task.priority = priority
            
            # Parse start date and time
            start_date_str = request.form.get('start_date')
            start_time_str = request.form.get('start_time', '08:00')
            
            # Update planned_start_date with time
            if start_date_str:
                planned_start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                if start_time_str:
                    try:
                        start_time = datetime.strptime(start_time_str, '%H:%M').time()
                        planned_start_date = datetime.combine(planned_start_date.date(), start_time)
                    except ValueError:
                        planned_start_date = datetime.combine(planned_start_date.date(), datetime.strptime('08:00', '%H:%M').time())
                task.planned_start_date = planned_start_date
            else:
                task.planned_start_date = None
            
            # Parse due date and time
            due_date_str = request.form.get('due_date')
            due_time_str = request.form.get('due_time', '14:00')
            
            if due_date_str:
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
                if due_time_str:
                    try:
                        due_time = datetime.strptime(due_time_str, '%H:%M').time()
                        due_date = datetime.combine(due_date.date(), due_time)
                    except ValueError:
                        due_date = datetime.combine(due_date.date(), datetime.strptime('14:00', '%H:%M').time())
                task.due_date = due_date
            
            # Validate: planned_start_date must be before due_date
            if task.planned_start_date and task.due_date and task.planned_start_date >= task.due_date:
                flash('La fecha/hora de inicio debe ser anterior a la fecha/hora de vencimiento.', 'danger')
                return redirect(url_for('main.edit_task', task_id=task.id))
            
            # Auto-update status to Scheduled if start date is in the future
            if task.planned_start_date:
                now = datetime.now()
                if task.planned_start_date > now and task.status in ['Pending', 'Scheduled']:
                    task.status = 'Scheduled'
                elif task.planned_start_date <= now and task.status == 'Scheduled':
                    task.status = 'Pending'
        
        # Update time_spent - Admin and Supervisors
        if current_user.is_admin or current_user.role == 'supervisor':
            time_spent_str = request.form.get('time_spent')
            if time_spent_str and time_spent_str.strip():
                try:
                    task.time_spent = int(time_spent_str)
                except ValueError:
                    pass  # Keep existing value if invalid
            else:
                task.time_spent = None
        
        
        # Update assignees - Admin and Supervisors
        if current_user.is_admin or current_user.role == 'supervisor':
            assignee_ids = request.form.getlist('assignees')
            task.assignees = [] # Clear current assignees
            for user_id in assignee_ids:
                user = User.query.get(int(user_id))
                if user:
                    task.assignees.append(user)
        
        # Update area - ONLY if user is admin
        if current_user.is_admin:
            area_id_str = request.form.get('area_id')
            if area_id_str and area_id_str.strip():
                try:
                    task.area_id = int(area_id_str)
                except ValueError:
                    pass  # Keep existing value if invalid

        # Update tags
        tag_ids = request.form.getlist('tags')
        task.tags = [] # Clear current tags
        for tag_id in tag_ids:
            tag = Tag.query.get(int(tag_id))
            if tag:
                task.tags.append(tag)
        
        # Handle parent task selection (with circular reference prevention)
        parent_id_str = request.form.get('parent_id')
        if parent_id_str:
            try:
                parent_id = int(parent_id_str)
                # Prevent circular reference: task cannot be its own parent
                if parent_id != task.id:
                    # Prevent circular reference: parent cannot be a descendant
                    if not is_descendant(task.id, parent_id):
                        parent_task = Task.query.get(parent_id)
                        if parent_task:
                            task.parent_id = parent_id
            except (ValueError, TypeError):
                pass  # Invalid parent_id, ignore it
        else:
            # Clear parent if field is empty
            task.parent_id = None
        
        # Handle completion tracking if status changed to Completed
        if task.status == 'Completed' and not task.completed_at:
             task.completed_by_id = current_user.id
             task.completed_at = now_utc()
        elif task.status == 'Pending':
             task.completed_by_id = None
             task.completed_at = None
             task.completion_comment = None  # Clear comment when reverting to Pending
        
        # Update completion comment (editable for completed tasks)
        completion_comment = request.form.get('completion_comment', '').strip()
        if completion_comment:
            task.completion_comment = completion_comment
        elif task.status != 'Completed':
            # Only clear comment if task is not completed
            task.completion_comment = None

        # Track edit history
        task.last_edited_by_id = current_user.id
        task.last_edited_at = now_utc()
        
        # Handle child tasks assignment
        child_ids_str = request.form.get('child_ids', '')
        if child_ids_str:
            try:
                child_ids = [int(id.strip()) for id in child_ids_str.split(',') if id.strip()]
                # Get current children IDs
                current_child_ids = [c.id for c in task.children]
                
                # Update children - set parent_id for new children
                for child_id in child_ids:
                    if child_id != task.id and child_id not in current_child_ids:
                        child_task = Task.query.get(child_id)
                        if child_task and not is_descendant(child_id, task.id):
                            child_task.parent_id = task.id
                
                # Remove parent_id from children that were removed
                for current_child_id in current_child_ids:
                    if current_child_id not in child_ids:
                        child_task = Task.query.get(current_child_id)
                        if child_task:
                            child_task.parent_id = None
            except (ValueError, TypeError):
                pass  # Invalid child_ids, ignore
        else:
            # Clear all children if no child_ids provided
            for child in task.children:
                child.parent_id = None

        db.session.commit()
        
        # Log activity
        # Log activity with diff
        import json
        changes = []
        
        if old_state['title'] != task.title:
            changes.append(f'Título modificado')
        if old_state['description'] != task.description:
            changes.append(f'Descripción modificada')
        if old_state['priority'] != task.priority:
            changes.append(f'Prioridad: {old_state["priority"]} -> {task.priority}')
        if old_state['status'] != task.status:
            # Record status change (StatusTransition)
            try:
                new_transition = StatusTransition(
                    task_id=task.id,
                    from_status=old_state['status'],
                    to_status=task.status,
                    changed_by_id=current_user.id,
                    comment=request.form.get('edit_comment')
                )
                db.session.add(new_transition)
            except Exception as e:
                pass # Fail silently if history recording fails to avoid blocking the save

            changes.append(f'Estado: {old_state["status"]} -> {task.status}')
        
        # Helper to format date for comparison
        def fmt_date(d): return d.strftime('%Y-%m-%d %H:%M') if d else 'None'
        
        if fmt_date(old_state['due_date']) != fmt_date(task.due_date):
            changes.append(f'Vencimiento: {fmt_date(old_state["due_date"])} -> {fmt_date(task.due_date)}')
            
        edit_comment = request.form.get('edit_comment')
        
        details = {
            'changes': changes,
            'comment': edit_comment
        }
        
        log_activity(
            user=current_user,
            action='task_edited',
            description=f'editó la tarea "{task.title}"',
            target_type='task',
            target_id=task.id,
            area_id=task.area_id,
            details=json.dumps(details)
        )
        
        flash('Tarea actualizada exitosamente.', 'success')
        return redirect(url_for('main.task_details', task_id=task.id))
        
        return redirect(url_for('main.dashboard'))
        
    return render_template('edit_task.html', task=task, users=users, available_tags=available_tags, available_areas=available_areas)

@main_bp.route('/task/<int:task_id>')
@login_required
def task_details(task_id):
    task = Task.query.get_or_404(task_id)
    
    # Build unified timeline
    timeline_events = []
    
    # Add Status Transitions
    # Add Status Transitions
    for transition in task.status_history:
        # Determine icon based on status
        icon = 'fa-circle'
        if transition.to_status == 'In Progress':
            icon = 'fa-play'
        elif transition.to_status == 'In Review':
            icon = 'fa-eye'
        elif transition.to_status == 'Completed':
            icon = 'fa-check-circle'
        elif transition.to_status == 'Pending':
            icon = 'fa-pause'

        timeline_events.append({
            'type': 'status',
            'timestamp': transition.changed_at,
            'user': transition.changed_by,
            'label': f"{transition.from_status} → {transition.to_status}",
            'status_label': transition.to_status, # For color coding
            'comment': transition.comment,
            'icon': icon
        })
        
    # Add Edit Logs
    # Fetch logs where target_id is this task and action is 'task_edited'
    from models import ActivityLog
    edit_logs = ActivityLog.query.filter_by(
        target_id=task.id, 
        target_type='task', 
        action='task_edited'
    ).all()
    
    import json
    for log in edit_logs:
        details = {}
        if log.details:
            try:
                details = json.loads(log.details)
            except:
                pass
        
        timeline_events.append({
            'type': 'edit',
            'timestamp': log.created_at,
            'user': log.user,
            'label': 'Tarea Editada',
            'changes': details.get('changes', []),
            'comment': details.get('comment'),
            'icon': 'fa-pen'
        })
        
    # Add Creation Event (optional, but good for completeness)
    timeline_events.append({
        'type': 'creation',
        'timestamp': task.created_at,
        'user': task.creator,
        'label': 'Tarea Creada',
        'icon': 'fa-plus'
    })
    
    # Add Attachment Events (upload/delete)
    attachment_logs = ActivityLog.query.filter(
        ActivityLog.target_id == task.id,
        ActivityLog.target_type == 'task',
        ActivityLog.action.in_(['attachment_upload', 'attachment_delete'])
    ).all()
    
    for log in attachment_logs:
        if log.action == 'attachment_upload':
            timeline_events.append({
                'type': 'attachment_upload',
                'timestamp': log.created_at,
                'user': log.user,
                'label': log.description,
                'icon': 'fa-paperclip'
            })
        else:  # attachment_delete
            timeline_events.append({
                'type': 'attachment_delete',
                'timestamp': log.created_at,
                'user': log.user,
                'label': log.description,
                'icon': 'fa-trash'
            })

    # Add Postpone Events
    postpone_logs = ActivityLog.query.filter(
        ActivityLog.target_id == task.id,
        ActivityLog.target_type == 'task',
        ActivityLog.action == 'task_postponed'
    ).all()
    
    for log in postpone_logs:
        details = {}
        if log.details:
            try:
                details = json.loads(log.details)
            except:
                pass
        
        old_date = details.get('old_due_date', '?')
        new_date = details.get('new_due_date', '?')
        
        timeline_events.append({
            'type': 'postpone',
            'timestamp': log.created_at,
            'user': log.user,
            'label': f'Pospuesta: {old_date} → {new_date}',
            'comment': None,
            'icon': 'fa-clock'
        })

    # Add Transfer Events
    transfer_logs = ActivityLog.query.filter(
        ActivityLog.target_id == task.id,
        ActivityLog.target_type == 'task',
        ActivityLog.action == 'task_transferred'
    ).all()
    
    for log in transfer_logs:
        details = {}
        if log.details:
            try:
                details = json.loads(log.details)
            except:
                pass
        
        to_user = details.get('to_user', '?')
        from_users = details.get('from_users', [])
        transfer_comment = details.get('comment')
        
        timeline_events.append({
            'type': 'transfer',
            'timestamp': log.created_at,
            'user': log.user,
            'label': f'Pasada a {to_user}',
            'comment': transfer_comment,
            'icon': 'fa-exchange-alt',
            'from_users': from_users
        })

    # Sort by timestamp
    timeline_events.sort(key=lambda x: x['timestamp'])

    return render_template('task_details.html', task=task, timeline_events=timeline_events)

@main_bp.route('/task-tree')
@login_required
def task_tree():
    """Display all tasks in a hierarchical tree view"""
    from models import Area
    
    # Get filter parameters (same as dashboard)
    filter_assignee = request.args.get('assignee')
    filter_creator = request.args.get('creator')
    filter_status = request.args.get('status', '')
    filter_tag = request.args.get('tag_filter')
    filter_area = request.args.get('area')
    sort_order = request.args.get('sort', 'asc')
    search_query = request.args.get('q', '')
    
    # Base query: get all root tasks (tasks without parent)
    query = Task.query.options(
        joinedload(Task.children),
        joinedload(Task.assignees),
        joinedload(Task.tags)
    ).filter(Task.parent_id == None)
    
    # --- ROLE-BASED VISIBILITY FILTERING ---
    user_area_ids = [a.id for a in current_user.areas]
    
    if current_user.can_only_see_own_tasks():
        # Users can only see tasks they are assigned to OR created
        if user_area_ids:
            query = query.filter(
                db.or_(
                    Task.assignees.any(id=current_user.id),
                    Task.creator_id == current_user.id
                )
            ).filter(Task.area_id.in_(user_area_ids))
        else:
            query = query.filter(Task.area_id == -1)
        available_areas = current_user.areas
        show_area_filter = False
    elif current_user.can_see_all_areas():
        # Gerentes/admins see all tasks
        if filter_area:
            query = query.filter(Task.area_id == int(filter_area))
        available_areas = Area.query.order_by(Area.name).all()
        show_area_filter = True
    else:
        # Supervisors see all tasks in their area
        if user_area_ids:
            query = query.filter(Task.area_id.in_(user_area_ids))
        else:
            query = query.filter(Task.area_id == -1)
        available_areas = current_user.areas
        show_area_filter = False
    
    # Apply assignee filter
    if filter_assignee:
        query = query.filter(Task.assignees.any(id=filter_assignee))
    
    # Apply creator filter
    if filter_creator:
        query = query.filter(Task.creator_id == filter_creator)
    
    # Apply status filter - exclude 'Anulado' by default
    if filter_status:
        if filter_status in ['Pending', 'Completed', 'Anulado']:
            query = query.filter(Task.status == filter_status)
    else:
        query = query.filter(Task.status != 'Anulado')
    
    # Apply tag filter
    if filter_tag:
        query = query.filter(Task.tags.any(id=int(filter_tag)))
    
    # Apply search filter
    if search_query:
        search_term = f"%{search_query}%"
        query = query.filter(
            (Task.title.ilike(search_term)) | 
            (Task.description.ilike(search_term))
        )
    
    # Sort by status first (Pending before Completed), then by due_date
    status_order = db.case(
        (Task.status == 'Pending', 0),
        (Task.status == 'Completed', 1),
        else_=2
    )
    
    if sort_order == 'desc':
        root_tasks = query.order_by(status_order, Task.due_date.desc()).all()
    else:
        root_tasks = query.order_by(status_order, Task.due_date.asc()).all()
    
    # Get counts for stats - FILTERED BY AREA
    stats_query_base = Task.query.filter(Task.status != 'Anulado')
    if not current_user.is_admin:
        if user_area_ids:
            stats_query_base = stats_query_base.filter(Task.area_id.in_(user_area_ids))
        else:
            stats_query_base = stats_query_base.filter(Task.area_id == -1)
    
    total_tasks = stats_query_base.count()
    tasks_with_children = stats_query_base.filter(Task.children.any()).count()
    tasks_with_parent = stats_query_base.filter(Task.parent_id != None).count()
    
    # Get users and tags for filter dropdowns - FILTERED BY AREA
    if current_user.is_admin:
        users = User.query.all()
        all_tags = Tag.query.order_by(Tag.name).all()
    else:
        users = [u for u in User.query.all() if any(area in u.areas for area in current_user.areas)]
        # Strict filter - only tags from user's areas
        if user_area_ids:
            all_tags = Tag.query.filter(Tag.area_id.in_(user_area_ids)).order_by(Tag.name).all()
        else:
            all_tags = []
    
    return render_template('task_tree.html', 
                           root_tasks=root_tasks,
                           filter_status=filter_status,
                           filter_assignee=filter_assignee,
                           filter_creator=filter_creator,
                           filter_tag=filter_tag,
                           filter_area=filter_area,
                           sort_order=sort_order,
                           search_query=search_query,
                           total_tasks=total_tasks,
                           tasks_with_children=tasks_with_children,
                           tasks_with_parent=tasks_with_parent,
                           users=users,
                           all_tags=all_tags,
                           all_areas=available_areas,
                           show_area_filter=show_area_filter)

@main_bp.route('/calendar')
@login_required
def calendar():
    from models import Area
    
    # Get filter parameters
    period = request.args.get('period', 'all')  # today, week, month, all
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    filter_user = request.args.get('creator')  # Keep 'creator' param name for backward compatibility
    filter_area = request.args.get('area')
    
    # Base query: show ALL tasks by default (matching dashboard behavior)
    # Only show ENABLED tasks (not blocked by parent dependency)
    query = Task.query.options(joinedload(Task.assignees), joinedload(Task.tags)).filter(Task.enabled == True)
    
    # Exclude 'Anulado' tasks by default
    query = query.filter(Task.status != 'Anulado')
    
    # --- ROLE-BASED VISIBILITY FILTERING ---
    user_area_ids = [a.id for a in current_user.areas]
    
    if current_user.can_only_see_own_tasks():
        # Users can only see tasks they are assigned to OR created
        if user_area_ids:
            query = query.filter(
                db.or_(
                    Task.assignees.any(id=current_user.id),
                    Task.creator_id == current_user.id
                )
            ).filter(Task.area_id.in_(user_area_ids))
        else:
            query = query.filter(Task.area_id == -1)
        available_areas = current_user.areas
        show_area_filter = False
    elif current_user.can_see_all_areas():
        # Gerentes/admins see all tasks
        if filter_area:
            query = query.filter(Task.area_id == int(filter_area))
        available_areas = Area.query.order_by(Area.name).all()
        show_area_filter = True
    else:
        # Supervisors see all tasks in their area
        if user_area_ids:
            query = query.filter(Task.area_id.in_(user_area_ids))
        else:
            query = query.filter(Task.area_id == -1)
        available_areas = current_user.areas
        show_area_filter = False
    
    # Filter by user (assignee) if selected
    if filter_user:
        query = query.filter(Task.assignees.any(id=filter_user))
    
    # Apply date filters
    today = date.today()
    
    if period == 'today':
        query = query.filter(db.func.date(Task.due_date) == today)
    elif period == 'week':
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        query = query.filter(db.func.date(Task.due_date) >= week_start,
                           db.func.date(Task.due_date) <= week_end)
    elif period == 'month':
        month_start = today.replace(day=1)
        # Get last day of month
        if today.month == 12:
            month_end = today.replace(day=31)
        else:
            month_end = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        query = query.filter(db.func.date(Task.due_date) >= month_start,
                           db.func.date(Task.due_date) <= month_end)
    elif start_date_str and end_date_str:
        # Custom date range
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        query = query.filter(db.func.date(Task.due_date) >= start_date,
                           db.func.date(Task.due_date) <= end_date)
    # else: period == 'all' - no date filtering
    
    tasks = query.order_by(Task.due_date.asc()).all()
    
    # Filter users by area (same logic as dashboard)
    if current_user.can_see_all_areas():
        users = User.query.all()
    else:
        # Non-admins see only users in their areas
        users = [u for u in User.query.all() if any(area in u.areas for area in current_user.areas)]
    
    # Get event dates for calendar widget (dates with tasks the user can see)
    month_start = today.replace(day=1)
    if today.month == 12:
        month_end = today.replace(day=31)
    else:
        month_end = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
    
    # Build event_dates query with same visibility rules as main query
    event_dates_query = db.session.query(db.func.date(Task.due_date)).filter(
        Task.enabled == True,
        Task.status != 'Anulado',
        db.func.date(Task.due_date) >= month_start,
        db.func.date(Task.due_date) <= month_end
    )
    
    # Apply same visibility filtering to event_dates
    if current_user.can_only_see_own_tasks():
        if user_area_ids:
            event_dates_query = event_dates_query.filter(
                db.or_(
                    Task.assignees.any(id=current_user.id),
                    Task.creator_id == current_user.id
                )
            ).filter(Task.area_id.in_(user_area_ids))
        else:
            event_dates_query = event_dates_query.filter(Task.area_id == -1)
    elif not current_user.can_see_all_areas():
        # Supervisors see all tasks in their area
        if user_area_ids:
            event_dates_query = event_dates_query.filter(Task.area_id.in_(user_area_ids))
        else:
            event_dates_query = event_dates_query.filter(Task.area_id == -1)
    
    event_dates = [d[0].strftime('%Y-%m-%d') for d in event_dates_query.distinct().all() if d[0]]
    
    # --- CALENDAR GRID GENERATION ---
    import calendar as cal
    
    # Get month/year from query params or use current month
    view_year = int(request.args.get('year', today.year))
    view_month = int(request.args.get('month', today.month))
    
    # Calculate previous and next month for navigation
    if view_month == 1:
        prev_month, prev_year = 12, view_year - 1
    else:
        prev_month, prev_year = view_month - 1, view_year
    
    if view_month == 12:
        next_month, next_year = 1, view_year + 1
    else:
        next_month, next_year = view_month + 1, view_year
    
    # Generate calendar weeks (list of 7-day lists)
    cal_obj = cal.Calendar(firstweekday=0)  # Monday = 0
    month_days = cal_obj.monthdatescalendar(view_year, view_month)
    
    # Get tasks for the visible calendar range (includes prev/next month day spillover)
    if month_days:
        cal_start = month_days[0][0]
        cal_end = month_days[-1][-1]
        
        # Query tasks within calendar view range
        cal_tasks_query = Task.query.options(joinedload(Task.assignees), joinedload(Task.tags)).filter(
            Task.enabled == True,
            Task.status != 'Anulado',
            db.func.date(Task.due_date) >= cal_start,
            db.func.date(Task.due_date) <= cal_end
        )
        
        # Apply same visibility filters
        if current_user.can_only_see_own_tasks():
            if user_area_ids:
                cal_tasks_query = cal_tasks_query.filter(
                    db.or_(
                        Task.assignees.any(id=current_user.id),
                        Task.creator_id == current_user.id
                    )
                ).filter(Task.area_id.in_(user_area_ids))
            else:
                cal_tasks_query = cal_tasks_query.filter(Task.area_id == -1)
        elif not current_user.can_see_all_areas():
            if user_area_ids:
                cal_tasks_query = cal_tasks_query.filter(Task.area_id.in_(user_area_ids))
            else:
                cal_tasks_query = cal_tasks_query.filter(Task.area_id == -1)
        
        cal_tasks = cal_tasks_query.all()
    else:
        cal_tasks = []
    
    # Group tasks by date
    tasks_by_date = {}
    for task in cal_tasks:
        date_key = task.due_date.date().isoformat()
        if date_key not in tasks_by_date:
            tasks_by_date[date_key] = []
        tasks_by_date[date_key].append(task)
    
    # Month names in Spanish
    month_names = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                   'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    return render_template('calendar.html', 
                          tasks=tasks, 
                          current_period=period, 
                          today=today, 
                          users=users, 
                          event_dates=event_dates,
                          all_areas=available_areas,
                          show_area_filter=show_area_filter,
                          # New calendar grid data
                          calendar_weeks=month_days,
                          view_month=view_month,
                          view_year=view_year,
                          month_name=month_names[view_month],
                          prev_month=prev_month,
                          prev_year=prev_year,
                          next_month=next_month,
                          next_year=next_year,
                          tasks_by_date=tasks_by_date)


@main_bp.route('/task/<int:task_id>/toggle', methods=['POST'])
@login_required
def toggle_task_status(task_id):
    task = Task.query.get_or_404(task_id)
    old_status = task.status
    
    # Verify user has access to this task (admin, creator, or assignee)
    if not current_user.is_admin and task.creator_id != current_user.id and current_user not in task.assignees:
        return jsonify({'success': False, 'error': 'Acceso denegado'}), 403
    
    # Get optional completion comment from request body
    data = request.get_json(silent=True) or {}
    completion_comment = data.get('comment', '').strip() if data else ''
    
    # Toggle status and track completion
    if task.status != 'Completed':
        task.status = 'Completed'
        task.completed_by_id = current_user.id
        task.completed_at = now_utc()
        task.completion_comment = completion_comment if completion_comment else None
        
        # Enable child tasks when parent is completed
        enabled_children_count = 0
        adjusted_dates_count = 0
        for child in task.children:
            if not child.enabled:
                child.enabled = True
                child.enabled_at = now_utc()
                child.enabled_by_task_id = task.id
                enabled_children_count += 1
                
                # If the child's due_date has already passed, adjust it to today
                if child.due_date.date() < date.today():
                    child.original_due_date = child.due_date  # Save original date before adjustment
                    child.due_date = now_utc()
                    adjusted_dates_count += 1
        
            if adjusted_dates_count > 0:
                msg += f' Se ajustó la fecha de vencimiento de {adjusted_dates_count} tarea(s) que ya habían vencido.'
            flash(msg, 'info')
        
        # Log process event if part of a process
        if task.process_id:
            log_process_event(
                process_id=task.process_id,
                event_type='task_completed',
                description=f'Tarea completada: {task.title}',
                user_id=current_user.id,
                task_id=task.id
            )
    else:
        # Note: We don't disable children when uncompleting - they stay enabled
        task.status = 'Pending'
        task.completed_by_id = None
        task.completed_at = None
        task.completion_comment = None  # Clear comment when reopening
    
    # Record status change
    if old_status != task.status:
        try:
            new_transition = StatusTransition(
                task_id=task.id,
                from_status=old_status,
                to_status=task.status,
                changed_by_id=current_user.id,
                comment=completion_comment if task.status == 'Completed' else None
            )
            db.session.add(new_transition)
        except Exception:
            pass

    db.session.commit()
    
    # Log activity
    action = 'task_completed' if task.status == 'Completed' else 'task_reopened'
    description = f'completó la tarea "{task.title}"' if task.status == 'Completed' else f'reabrió la tarea "{task.title}"'
    log_activity(
        user=current_user,
        action=action,
        description=description,
        target_type='task',
        target_id=task.id,
        area_id=task.area_id
    )
    
    return jsonify({
        'success': True,
        'task_id': task.id,
        'new_status': task.status
    })

@main_bp.route('/task/<int:task_id>/anular', methods=['POST'])
@login_required
def anular_task(task_id):
    """Mark a task as 'Anulado' (soft delete)"""
    task = Task.query.get_or_404(task_id)
    
    # Only creator or admin can anular
    if task.creator_id != current_user.id and not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Solo el creador o un admin puede anular esta tarea'}), 403
    
    task.status = 'Anulado'
    task.completed_by_id = current_user.id  # Track who anulled it
    task.completed_at = now_utc()  # Track when
    
    db.session.commit()
    
    # Log activity
    log_activity(
        user=current_user,
        action='task_anulada',
        description=f'anuló la tarea "{task.title}"',
        target_type='task',
        target_id=task.id,
        area_id=task.area_id
    )
    
    return jsonify({
        'success': True,
        'task_id': task.id,
        'new_status': 'Anulado'
    })


# Valid status values and transitions
VALID_STATUSES = ['Pending', 'In Progress', 'In Review', 'Completed', 'Anulado', 'Scheduled']
STATUS_LABELS = {
    'Pending': 'Pendiente',
    'In Progress': 'En Proceso',
    'In Review': 'En Revisión',
    'Completed': 'Completado',
    'Anulado': 'Anulado',
    'Scheduled': 'Programada'
}


def can_change_task_status(user, task, new_status):
    """
    Check if user can change task status based on role and task ownership.
    
    Returns: (bool, str) - (allowed, error_message)
    """
    # Admin can do anything
    if user.is_admin:
        return True, None
    
    # Supervisor can change any task in their area
    if user.role == 'supervisor':
        user_area_ids = [a.id for a in user.areas]
        if task.area_id in user_area_ids:
            return True, None
        return False, "Solo puedes modificar tareas de tu área"
    
    # Usuario+ can change status on their own tasks
    if user.role == 'usuario_plus':
        if user in task.assignees or task.creator_id == user.id:
            # Can move to any active status
            if new_status in ['Pending', 'In Progress', 'In Review', 'Completed']:
                return True, None
            return False, "Estado inválido"
        return False, "Solo puedes modificar tareas asignadas a ti"

    # Usuario can change status on their own tasks
    if user.role == 'usuario':
        if user in task.assignees or task.creator_id == user.id:
            # Can move to any active status
            if new_status in ['Pending', 'In Progress', 'In Review', 'Completed']:
                return True, None
            return False, "Estado inválido"
        return False, "Solo puedes modificar tareas asignadas a ti"
    
    return False, "No tienes permiso para cambiar el estado"


@main_bp.route('/task/<int:task_id>/status', methods=['POST'])
@login_required
def update_task_status(task_id):
    """
    Update task status for Scrum board.
    Expects JSON: { "status": "In Progress" }
    """
    task = Task.query.get_or_404(task_id)
    data = request.get_json() or {}
    new_status = data.get('status')
    
    if not new_status or new_status not in VALID_STATUSES:
        return jsonify({
            'success': False, 
            'error': f'Estado invalido. Valores validos: {", ".join(VALID_STATUSES)}'
        }), 400
    
    # Check if task is blocked (disabled)
    if not task.enabled:
        parent_info = f" (Esperando tarea #{task.parent.id})" if task.parent else ""
        return jsonify({
            'success': False, 
            'error': f'Esta tarea está bloqueada y no puede cambiar de estado{parent_info}'
        }), 403
    
    # Check permission
    allowed, error_msg = can_change_task_status(current_user, task, new_status)
    if not allowed:
        return jsonify({'success': False, 'error': error_msg}), 403
    
    old_status = task.status
    
    # Update status and tracking fields
    task.status = new_status
    
    if new_status == 'In Progress' and not task.started_at:
        task.started_at = now_utc()
        task.started_by_id = current_user.id
    elif new_status == 'In Review' and not task.in_review_at:
        task.in_review_at = now_utc()
        task.in_review_by_id = current_user.id
    elif new_status == 'Completed' and not task.completed_at:
        task.completed_at = now_utc()
        
        # If completing from "In Review", credit goes to the person who did the work
        # and the current user (supervisor) is tracked as the approver
        if old_status == 'In Review' and task.in_review_by_id:
            # The person who sent to review did the work
            task.completed_by_id = task.in_review_by_id
            # The current user (supervisor) approved it
            task.approved_by_id = current_user.id
            task.approved_at = now_utc()
        else:
            # Direct completion (not from review) - current user gets credit
            task.completed_by_id = current_user.id
        
        if data.get('comment'):
            task.completion_comment = data.get('comment')
        
        # Log process event if part of a process
        if task.process_id:
            log_process_event(
                process_id=task.process_id,
                event_type='task_completed',
                description=f'Tarea completada: {task.title}',
                user_id=current_user.id,
                task_id=task.id
            )
            
        # Enable child tasks when parent is completed
        for child in task.children:
            if not child.enabled:
                child.enabled = True
                child.enabled_at = now_utc()
                child.enabled_by_task_id = task.id
                if child.due_date.date() < date.today():
                    child.original_due_date = child.due_date
                    child.due_date = now_utc()
    elif new_status == 'Pending':
        # Reset all tracking when moving back to Pending
        task.started_at = None
        task.started_by_id = None
        task.in_review_at = None
        task.in_review_by_id = None
        task.completed_at = None
        task.completed_by_id = None
        task.approved_at = None
        task.approved_by_id = None
    elif new_status == 'Anulado':
        # 1. Update current task first
        task.status = 'Anulado'
        task.enabled = False
        task.completed_at = now_utc()
        task.completed_by_id = current_user.id
        task.last_edited_by_id = current_user.id
        task.last_edited_at = now_utc()
        
        # 2. Robust Cascade: Propagate to all descendants iteratively
        # We use flush() to make sure the DB sees the parent's generic state
        db.session.flush()
        
        from sqlalchemy.orm import aliased
        Parent = aliased(Task)
        
        # Safety limit to prevent infinite loops (max depth 100)
        for _ in range(100):
            # Find any task that is NOT annulled, but whose parent IS annulled
            orphans = Task.query.join(Parent, Task.parent)\
                .filter(Task.status != 'Anulado')\
                .filter(Parent.status == 'Anulado')\
                .all()
            
            if not orphans:
                break
                
            for child in orphans:
                child.status = 'Anulado'
                child.enabled = False
                child.completed_at = now_utc()
                child.completed_by_id = current_user.id
                child.last_edited_by_id = current_user.id
                child.last_edited_at = now_utc()
            
            # Flush changes so next iteration sees these as annulled parents
            db.session.flush()
    
    task.last_edited_by_id = current_user.id
    task.last_edited_at = now_utc()
    
    # Record status change
    from models import StatusTransition
    new_transition = StatusTransition(
        task_id=task.id,
        from_status=old_status,
        to_status=new_status,
        changed_by_id=current_user.id,
        comment=data.get('comment')
    )
    db.session.add(new_transition)
    
    db.session.commit()
    
    # Log activity
    log_activity(
        user=current_user,
        action='task_status_changed',
        description=f'cambió estado de "{STATUS_LABELS.get(old_status, old_status)}" a "{STATUS_LABELS.get(new_status, new_status)}" en tarea "{task.title}"',
        target_type='task',
        target_id=task.id,
        area_id=task.area_id
    )
    
    # Auto-complete process if all tasks are completed
    process_completed = False
    if new_status == 'Completed' and task.process_id:
        process = Process.query.get(task.process_id)
        if process and process.check_and_complete():
            process.completed_by_id = current_user.id
            db.session.commit()
            process_completed = True
            log_activity(
                user=current_user,
                action='process_auto_completed',
                description=f'completó automáticamente el proceso "{process.name}" (todas las tareas completadas)',
                target_type='process',
                target_id=process.id,
                area_id=process.area_id
            )
    
    return jsonify({
        'success': True,
        'task_id': task.id,
        'old_status': old_status,
        'new_status': new_status,
        'process_completed': process_completed
    })


@main_bp.route('/scrum-board')
@login_required
def scrum_board():
    """Scrum/Kanban board view with 4 status columns - OPTIMIZED"""
    from models import Area
    
    # Limit for completed tasks (performance optimization)
    COMPLETED_LIMIT = 10
    
    # Get filter parameters
    filter_area = request.args.get('area')
    filter_assignee = request.args.get('assignee')
    filter_period = request.args.get('period', 'week')  # DEFAULT CHANGED TO 'week' for performance
    filter_date_from = request.args.get('date_from')
    filter_date_to = request.args.get('date_to')
    show_blocked = request.args.get('show_blocked', 'true')
    
    # Pagination for completed tasks (lazy loading)
    completed_offset = request.args.get('completed_offset', 0, type=int)
    
    # Base query options (reusable)
    query_options = [
        joinedload(Task.assignees),
        joinedload(Task.tags),
        joinedload(Task.area),
        joinedload(Task.parent)
    ]
    
    # Date filter setup
    today = date.today()
    overdue_condition = db.and_(
        db.func.date(Task.due_date) < today,
        Task.status.in_(['Pending', 'In Progress', 'In Review'])
    )
    
    def apply_date_filter(query):
        """Apply date/period filter to query"""
        if filter_period == 'today':
            return query.filter(
                db.or_(
                    db.func.date(Task.due_date) == today,
                    db.func.date(Task.planned_start_date) == today,
                    Task.status.in_(['In Progress', 'In Review']),
                    overdue_condition
                )
            )
        elif filter_period == 'week':
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
            return query.filter(
                db.or_(
                    db.and_(db.func.date(Task.due_date) >= week_start, db.func.date(Task.due_date) <= week_end),
                    db.and_(db.func.date(Task.planned_start_date) >= week_start, db.func.date(Task.planned_start_date) <= week_end),
                    Task.status.in_(['In Progress', 'In Review']),
                    overdue_condition
                )
            )
        elif filter_period == 'month':
            month_start = today.replace(day=1)
            if today.month == 12:
                month_end = today.replace(day=31)
            else:
                month_end = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
            return query.filter(
                db.or_(
                    db.and_(db.func.date(Task.due_date) >= month_start, db.func.date(Task.due_date) <= month_end),
                    db.and_(db.func.date(Task.planned_start_date) >= month_start, db.func.date(Task.planned_start_date) <= month_end),
                    Task.status.in_(['In Progress', 'In Review']),
                    overdue_condition
                )
            )
        elif filter_period == 'custom':
            custom_conditions = [overdue_condition]
            if filter_date_from:
                try:
                    date_from_obj = datetime.strptime(filter_date_from, '%Y-%m-%d').date()
                    custom_conditions.append(db.func.date(Task.due_date) >= date_from_obj)
                except ValueError:
                    pass
            if filter_date_to:
                try:
                    date_to_obj = datetime.strptime(filter_date_to, '%Y-%m-%d').date()
                    custom_conditions.append(db.func.date(Task.due_date) <= date_to_obj)
                except ValueError:
                    pass
            if custom_conditions:
                return query.filter(db.or_(*custom_conditions))
        # 'all' - no date filter
        return query
    
    def apply_role_filter(query):
        """Apply role-based visibility filter"""
        user_area_ids = [a.id for a in current_user.areas]
        
        if current_user.can_only_see_own_tasks():
            if user_area_ids:
                return query.filter(
                    db.or_(
                        Task.assignees.any(id=current_user.id),
                        Task.creator_id == current_user.id
                    )
                ).filter(Task.area_id.in_(user_area_ids))
            else:
                return query.filter(Task.area_id == -1)
        elif current_user.can_see_all_areas():
            if filter_area:
                return query.filter(Task.area_id == int(filter_area))
            return query
        else:
            # Supervisor
            if user_area_ids:
                return query.filter(Task.area_id.in_(user_area_ids))
            else:
                return query.filter(Task.area_id == -1)
        return query
    
    # Determine available areas based on role
    user_area_ids = [a.id for a in current_user.areas]
    if current_user.can_only_see_own_tasks():
        available_areas = current_user.areas
        show_area_filter = False
    elif current_user.can_see_all_areas():
        available_areas = Area.query.order_by(Area.name).all()
        show_area_filter = True
    else:
        available_areas = current_user.areas
        show_area_filter = False
    
    # Helper to build base query with filters
    def build_filtered_query(status_filter):
        q = Task.query.options(*query_options).filter(Task.status.in_(status_filter) if isinstance(status_filter, list) else Task.status == status_filter)
        if show_blocked != 'true':
            q = q.filter(Task.enabled == True)
        q = apply_date_filter(q)
        q = apply_role_filter(q)
        if filter_assignee:
            q = q.filter(Task.assignees.any(id=int(filter_assignee)))
        return q
    
    # === QUERY FOR EACH COLUMN (with load more support) ===
    COLUMN_LIMIT = 10
    
    # Get offset parameters for each column (for load more)
    pending_offset = request.args.get('pending_offset', 0, type=int)
    in_progress_offset = request.args.get('in_progress_offset', 0, type=int)
    in_review_offset = request.args.get('in_review_offset', 0, type=int)
    completed_offset = request.args.get('completed_offset', 0, type=int)
    
    # Pending (includes Scheduled)
    pending_query = build_filtered_query(['Pending', 'Scheduled'])
    pending_total = pending_query.count()
    pending_tasks = pending_query.order_by(Task.due_date.asc()).limit(COLUMN_LIMIT + pending_offset).all()
    
    # In Progress
    in_progress_query = build_filtered_query('In Progress')
    in_progress_total = in_progress_query.count()
    in_progress_tasks = in_progress_query.order_by(Task.due_date.asc()).limit(COLUMN_LIMIT + in_progress_offset).all()
    
    # In Review
    in_review_query = build_filtered_query('In Review')
    in_review_total = in_review_query.count()
    in_review_tasks = in_review_query.order_by(Task.due_date.asc()).limit(COLUMN_LIMIT + in_review_offset).all()
    
    # Completed (filter by completed_at instead of due_date to avoid tasks disappearing)
    completed_query = Task.query.options(*query_options).filter(Task.status == 'Completed')
    completed_query = apply_role_filter(completed_query)
    if filter_assignee:
        completed_query = completed_query.filter(Task.assignees.any(id=int(filter_assignee)))
    
    # Apply date filter based on completed_at (not due_date) for completed tasks
    if filter_period == 'today':
        completed_query = completed_query.filter(db.func.date(Task.completed_at) == today)
    elif filter_period == 'week':
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        completed_query = completed_query.filter(
            db.func.date(Task.completed_at) >= week_start,
            db.func.date(Task.completed_at) <= week_end
        )
    elif filter_period == 'month':
        month_start = today.replace(day=1)
        if today.month == 12:
            month_end = today.replace(day=31)
        else:
            month_end = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        completed_query = completed_query.filter(
            db.func.date(Task.completed_at) >= month_start,
            db.func.date(Task.completed_at) <= month_end
        )
    elif filter_period == 'custom':
        if filter_date_from:
            try:
                date_from_obj = datetime.strptime(filter_date_from, '%Y-%m-%d').date()
                completed_query = completed_query.filter(db.func.date(Task.completed_at) >= date_from_obj)
            except ValueError:
                pass
        if filter_date_to:
            try:
                date_to_obj = datetime.strptime(filter_date_to, '%Y-%m-%d').date()
                completed_query = completed_query.filter(db.func.date(Task.completed_at) <= date_to_obj)
            except ValueError:
                pass
    # 'all' - no date filter
    
    completed_total = completed_query.count()
    completed_tasks = completed_query.order_by(Task.completed_at.desc()).limit(COLUMN_LIMIT + completed_offset).all()
    
    # Group tasks by status
    tasks_by_status = {
        'Pending': pending_tasks,
        'In Progress': in_progress_tasks,
        'In Review': in_review_tasks,
        'Completed': completed_tasks
    }
    
    # Total counts per column
    column_totals = {
        'Pending': pending_total,
        'In Progress': in_progress_total,
        'In Review': in_review_total,
        'Completed': completed_total
    }
    
    # Get users for filter
    if current_user.can_see_all_areas():
        users = User.query.order_by(User.full_name).all()
    else:
        users = [u for u in User.query.all() if any(a in u.areas for a in current_user.areas)]
    
    return render_template('scrum_board.html',
                           tasks_by_status=tasks_by_status,
                           column_totals=column_totals,
                           column_limit=COLUMN_LIMIT,
                           all_areas=available_areas,
                           show_area_filter=show_area_filter,
                           users=users,
                           filter_area=filter_area,
                           filter_assignee=filter_assignee,
                           filter_period=filter_period,
                           filter_date_from=filter_date_from,
                           filter_date_to=filter_date_to,
                           status_labels=STATUS_LABELS,
                           today=today)

@main_bp.route('/export_pdf')
@login_required
def export_pdf():
    from models import Area

    # Re-use filter logic from dashboard
    filter_assignee = request.args.get('assignee')
    filter_creator = request.args.get('creator')
    filter_status = request.args.get('status')
    filter_area = request.args.get('area')

    # Calendar specific filters
    period = request.args.get('period')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    filters = {}

    # Start query
    query = Task.query.options(joinedload(Task.assignees), joinedload(Task.tags))

    # Apply Assignee Filter (works for both calendar and dashboard)
    if filter_assignee:
        query = query.filter(Task.assignees.any(id=filter_assignee))
        assignee = User.query.get(filter_assignee)
        filters['assignee_name'] = assignee.full_name if assignee else 'Desconocido'

    # Apply Creator Filter
    if filter_creator:
        query = query.filter(Task.creator_id == filter_creator)
        creator = User.query.get(filter_creator)
        filters['creator_name'] = creator.full_name if creator else 'Desconocido'
        filters['creator'] = filter_creator

    # Apply Area Filter
    if filter_area:
        query = query.filter(Task.area_id == int(filter_area))
        area = Area.query.get(int(filter_area))
        filters['area_name'] = area.name if area else 'Desconocida'

    # Apply Status Filter - exclude 'Anulado' by default like dashboard
    if filter_status:
        if filter_status in ['Pending', 'Completed', 'Anulado']:
            query = query.filter(Task.status == filter_status)
            filters['status'] = filter_status
    else:
        # By default, exclude 'Anulado' tasks (matching dashboard behavior)
        query = query.filter(Task.status != 'Anulado')

    # Apply Date/Period Filters
    today = date.today()
    if period == 'today':
        query = query.filter(db.func.date(Task.due_date) == today)
        filters['date_range'] = f'Hoy ({today.strftime("%d/%m/%Y")})'
    elif period == 'week':
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        query = query.filter(db.func.date(Task.due_date) >= week_start,
                           db.func.date(Task.due_date) <= week_end)
        filters['date_range'] = 'Esta Semana'
    elif period == 'month':
        month_start = today.replace(day=1)
        if today.month == 12:
            month_end = today.replace(day=31)
        else:
            month_end = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        query = query.filter(db.func.date(Task.due_date) >= month_start,
                           db.func.date(Task.due_date) <= month_end)
        filters['date_range'] = 'Este Mes'
    elif start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            query = query.filter(db.func.date(Task.due_date) >= start_date,
                               db.func.date(Task.due_date) <= end_date)
            filters['date_range'] = f"{start_date_str} a {end_date_str}"
        except ValueError:
            pass
            
    # Tag filter
    filter_tag = request.args.get('tag_filter')
    if filter_tag:
        query = query.filter(Task.tags.any(id=int(filter_tag)))
        tag = Tag.query.get(int(filter_tag))
        filters['tag'] = tag.name if tag else ''

    # Search filter
    search_query = request.args.get('q')
    if search_query:
        search_term = f"%{search_query}%"
        query = query.filter(
            (Task.title.ilike(search_term)) | 
            (Task.description.ilike(search_term))
        )
        filters['search'] = search_query
        
    tasks = query.order_by(Task.due_date.asc()).all()
    
    pdf = generate_task_pdf(tasks, filters)
    
    response = make_response(pdf.output(dest='S').encode('latin-1'))
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_tareas_{date.today()}.pdf'
    
    return response

@main_bp.route('/export_excel')
@login_required
def export_excel():
    from models import Area

    # Re-use filter logic from dashboard (same as export_pdf)
    filter_assignee = request.args.get('assignee')
    filter_creator = request.args.get('creator')
    filter_status = request.args.get('status')
    filter_area = request.args.get('area')

    # Calendar specific filters
    period = request.args.get('period')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    filters = {}

    # Start query
    query = Task.query.options(joinedload(Task.assignees), joinedload(Task.tags))

    # Apply Assignee Filter (works for both calendar and dashboard)
    if filter_assignee:
        query = query.filter(Task.assignees.any(id=filter_assignee))
        assignee = User.query.get(filter_assignee)
        filters['assignee_name'] = assignee.full_name if assignee else 'Desconocido'

    # Apply Creator Filter
    if filter_creator:
        query = query.filter(Task.creator_id == filter_creator)
        creator = User.query.get(filter_creator)
        filters['creator_name'] = creator.full_name if creator else 'Desconocido'
        filters['creator'] = filter_creator

    # Apply Area Filter
    if filter_area:
        query = query.filter(Task.area_id == int(filter_area))
        area = Area.query.get(int(filter_area))
        filters['area_name'] = area.name if area else 'Desconocida'

    # Apply Status Filter - exclude 'Anulado' by default like dashboard
    if filter_status:
        if filter_status in ['Pending', 'Completed', 'Anulado']:
            query = query.filter(Task.status == filter_status)
            filters['status'] = filter_status
    else:
        # By default, exclude 'Anulado' tasks (matching dashboard behavior)
        query = query.filter(Task.status != 'Anulado')

    # Apply Date/Period Filters
    today = date.today()
    if period == 'today':
        query = query.filter(db.func.date(Task.due_date) == today)
        filters['date_range'] = f'Hoy ({today.strftime("%d/%m/%Y")})'
    elif period == 'week':
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        query = query.filter(db.func.date(Task.due_date) >= week_start,
                           db.func.date(Task.due_date) <= week_end)
        filters['date_range'] = 'Esta Semana'
    elif period == 'month':
        month_start = today.replace(day=1)
        if today.month == 12:
            month_end = today.replace(day=31)
        else:
            month_end = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        query = query.filter(db.func.date(Task.due_date) >= month_start,
                           db.func.date(Task.due_date) <= month_end)
        filters['date_range'] = 'Este Mes'
    elif start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            query = query.filter(db.func.date(Task.due_date) >= start_date,
                               db.func.date(Task.due_date) <= end_date)
            filters['date_range'] = f"{start_date_str} a {end_date_str}"
        except ValueError:
            pass
            
    # Tag filter
    filter_tag = request.args.get('tag_filter')
    if filter_tag:
        query = query.filter(Task.tags.any(id=int(filter_tag)))
        tag = Tag.query.get(int(filter_tag))
        filters['tag'] = tag.name if tag else ''

    # Search filter
    search_query = request.args.get('q')
    if search_query:
        search_term = f"%{search_query}%"
        query = query.filter(
            (Task.title.ilike(search_term)) | 
            (Task.description.ilike(search_term))
        )
        filters['search'] = search_query
        
    tasks = query.order_by(Task.due_date.asc()).all()
    
    wb = generate_task_excel(tasks, filters)
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    response = make_response(excel_file.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_tareas_{date.today()}.xlsx'
    
    # Log activity
    log_activity(
        user=current_user,
        action='export_excel',
        description=f'exportó {len(tasks)} tareas a Excel',
        area_id=current_user.areas[0].id if current_user.areas else None
    )
    
    return response

# --- Admin Routes ---
@admin_bp.route('/users', methods=['GET', 'POST'])
@login_required
def manage_users():
    from models import Area
    
    # Allow both admins and supervisors to manage users
    is_supervisor = current_user.role == 'supervisor'
    
    if not current_user.is_admin and not is_supervisor:
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Supervisors can only see/manage their area
    if is_supervisor:
        supervisor_area = current_user.areas[0] if current_user.areas else None
        if not supervisor_area:
            flash('No tienes un área asignada. Contacta a un administrador.', 'danger')
            return redirect(url_for('main.dashboard'))
        all_areas = [supervisor_area]
    else:
        all_areas = Area.query.order_by(Area.name).all()
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        full_name = request.form.get('full_name')
        
        # Supervisors cannot create admins, supervisors, or gerentes
        if is_supervisor:
            is_admin = False
            requested_role = request.form.get('role', 'usuario')
            # Supervisors can only create 'usuario' or 'usuario_plus'
            role = requested_role if requested_role in ['usuario', 'usuario_plus'] else 'usuario'
            area_ids = [str(supervisor_area.id)]  # Force to supervisor's area
        else:
            is_admin = 'is_admin' in request.form
            role = request.form.get('role', 'usuario')
            area_ids = request.form.getlist('areas')
        
        if User.query.filter_by(username=username).first():
            flash('El nombre de usuario ya existe.', 'warning')
        else:
            new_user = User(
                username=username, 
                full_name=full_name, 
                is_admin=is_admin, 
                email=f"{username}@example.com",
                role=role
            )
            new_user.set_password(password)
            
            # Assign areas
            for area_id in area_ids:
                area = Area.query.get(int(area_id))
                if area:
                    new_user.areas.append(area)
            
            db.session.add(new_user)
            db.session.commit()
            
            # Log activity
            log_activity(
                user=current_user,
                action='user_created',
                description=f'creó el usuario "{new_user.full_name}"',
                target_type='user',
                target_id=new_user.id,
                area_id=new_user.areas[0].id if new_user.areas else None
            )
            
            flash('Usuario creado exitosamente.', 'success')
    
    # Get users to display
    if is_supervisor:
        # Supervisors only see users in their area
        users = [u for u in User.query.all() if supervisor_area in u.areas]
    else:
        users = User.query.all()
    
    return render_template('users.html', users=users, all_areas=all_areas, is_supervisor=is_supervisor)

# --- Area Management Routes ---
@admin_bp.route('/areas', methods=['GET', 'POST'])
@login_required
def manage_areas():
    """Manage areas (departments) - Admin only"""
    from models import Area
    
    if not current_user.is_admin:
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        color = request.form.get('color', '#6366f1')
        
        if not name:
            flash('El nombre del área es requerido.', 'warning')
        elif Area.query.filter_by(name=name).first():
            flash('Ya existe un área con ese nombre.', 'warning')
        else:
            new_area = Area(
                name=name,
                description=description,
                color=color
            )
            db.session.add(new_area)
            db.session.commit()
            flash(f'Área "{name}" creada exitosamente.', 'success')
    
    areas = Area.query.order_by(Area.name).all()
    return render_template('manage_areas.html', areas=areas)

@admin_bp.route('/areas/<int:area_id>/delete', methods=['POST'])
@login_required
def delete_area(area_id):
    """Delete an area - Admin only"""
    from models import Area
    
    if not current_user.is_admin:
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    area = Area.query.get_or_404(area_id)
    
    # Check if area has tasks
    task_count = area.tasks.count()
    if task_count > 0:
        flash(f'No se puede eliminar el área "{area.name}" porque tiene {task_count} tareas asignadas.', 'danger')
        return redirect(url_for('admin.manage_areas'))
    
    area_name = area.name
    db.session.delete(area)
    db.session.commit()
    flash(f'Área "{area_name}" eliminada.', 'success')
    return redirect(url_for('admin.manage_areas'))

@admin_bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    """Edit user - update role and areas - Admin only"""
    from models import Area
    
    if not current_user.is_admin:
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    user = User.query.get_or_404(user_id)
    all_areas = Area.query.order_by(Area.name).all()
    
    if request.method == 'POST':
        user.full_name = request.form.get('full_name', user.full_name)
        user.is_admin = 'is_admin' in request.form
        user.role = request.form.get('role', 'usuario')
        
        # Update password if provided
        new_password = request.form.get('password', '').strip()
        if new_password:
            user.set_password(new_password)
        
        # Update areas
        area_ids = request.form.getlist('areas')
        
        user.areas.clear()
        for area_id in area_ids:
            area = Area.query.get(int(area_id))
            if area:
                user.areas.append(area)
        
        db.session.commit()
        
        # Log activity
        log_activity(
            user=current_user,
            action='user_edited',
            description=f'editó el usuario "{user.full_name}"',
            target_type='user',
            target_id=user.id,
            area_id=user.areas[0].id if user.areas else None
        )
        
        flash(f'Usuario "{user.username}" actualizado.', 'success')
        return redirect(url_for('admin.manage_users'))
    
    return render_template('edit_user.html', user=user, all_areas=all_areas)


# --- API Routes for Notifications ---
@main_bp.route('/api/tasks/due_soon')
@login_required
def api_tasks_due_soon():
    """
    Return tasks and expirations that are due soon or already overdue.
    Used for popup notifications.
    """
    if not current_user.notifications_enabled:
        return jsonify({'tasks': [], 'expirations': [], 'overdue_tasks': [], 'overdue_expirations': []})
    
    # Get all pending/active AND enabled tasks assigned to current user (exclude blocked tasks)
    # Include 'Scheduled' to show reminders for future tasks based on due_date
    active_tasks = Task.query.filter(
        Task.assignees.any(id=current_user.id),
        Task.status.in_(['Pending', 'In Progress', 'In Review', 'Scheduled']),
        Task.enabled == True  # Only show enabled tasks (not blocked by parent)
    ).all()
    
    # Separate tasks into due soon and overdue
    due_soon_tasks = []
    overdue_tasks = []
    for task in active_tasks:
        business_days = calculate_business_days_until(task.due_date)
        task_data = {
            'id': task.id,
            'title': task.title,
            'due_date': task.due_date.strftime('%d/%m/%Y'),
            'priority': task.priority,
            'description': task.description[:100] if task.description else '',
            'days_remaining': business_days,
            'type': 'task',
            'enabled_at': task.enabled_at.strftime('%d/%m/%Y') if task.enabled_at else None
        }
        if business_days < 0:  # Overdue
            task_data['days_overdue'] = abs(business_days)
            overdue_tasks.append(task_data)
        elif business_days <= 2:  # Due in 0, 1, or 2 business days
            due_soon_tasks.append(task_data)
    
    # Get pending expirations filtered by area
    # Gerentes and admins see all expirations; others only see expirations from their own areas
    if current_user.can_see_all_areas():
        pending_expirations = Expiration.query.filter(
            Expiration.completed == False
        ).all()
    else:
        user_area_ids = [area.id for area in current_user.areas]
        if user_area_ids:
            pending_expirations = Expiration.query.filter(
                Expiration.completed == False,
                Expiration.area_id.in_(user_area_ids)
            ).all()
        else:
            pending_expirations = []
    
    # Separate expirations into due soon and overdue
    due_soon_expirations = []
    overdue_expirations = []
    for exp in pending_expirations:
        business_days = calculate_business_days_until(exp.due_date)
        exp_data = {
            'id': exp.id,
            'title': exp.title,
            'due_date': exp.due_date.strftime('%d/%m/%Y'),
            'description': exp.description[:100] if exp.description else '',
            'days_remaining': business_days,
            'type': 'expiration',
            'creator': exp.creator.full_name
        }
        if business_days < 0:  # Overdue
            exp_data['days_overdue'] = abs(business_days)
            overdue_expirations.append(exp_data)
        elif business_days <= 2:  # Due in 0, 1, or 2 business days
            due_soon_expirations.append(exp_data)
    
    return jsonify({
        'tasks': due_soon_tasks,
        'expirations': due_soon_expirations,
        'overdue_tasks': overdue_tasks,
        'overdue_expirations': overdue_expirations
    })

@main_bp.route('/api/user/toggle_notifications', methods=['POST'])
@login_required
def api_toggle_notifications():
    """
    Toggle user's notification preference.
    """
    current_user.notifications_enabled = not current_user.notifications_enabled
    db.session.commit()
    
    return jsonify({
        'success': True,
        'notifications_enabled': current_user.notifications_enabled
    })

@main_bp.route('/api/tasks/<int:task_id>/postpone', methods=['POST'])
@login_required
def api_postpone_task(task_id):
    """
    Posponer una tarea moviendo su fecha de vencimiento.
    Registra quién y cuándo la pospuso en el historial de actividad.
    
    Body JSON:
        days: int (1, 3, 7, etc.) - días a posponer
        custom_date: string (YYYY-MM-DD) - fecha personalizada (opcional, si se usa ignora days)
    """
    task = Task.query.get_or_404(task_id)
    
    # Validar que el usuario tiene permiso (es asignado, creador, admin o supervisor del área)
    is_assignee = current_user in task.assignees
    is_creator = task.creator_id == current_user.id
    is_admin = current_user.is_admin
    is_supervisor = current_user.role == 'supervisor' and task.area_id in [a.id for a in current_user.areas]
    
    if not (is_assignee or is_creator or is_admin or is_supervisor):
        return jsonify({'success': False, 'message': 'No tienes permiso para posponer esta tarea'}), 403
    
    data = request.get_json()
    days = data.get('days')
    custom_date_str = data.get('custom_date')
    
    # Guardar fecha original para el log
    old_due_date = task.due_date
    old_due_date_str = old_due_date.strftime('%d/%m/%Y')
    
    # Calcular nueva fecha
    if custom_date_str:
        try:
            # Parsear fecha personalizada
            new_date = datetime.strptime(custom_date_str, '%Y-%m-%d')
            # Mantener la hora original del due_date
            new_due_date = new_date.replace(
                hour=old_due_date.hour,
                minute=old_due_date.minute,
                second=old_due_date.second
            )
            postpone_description = f'fecha personalizada ({new_due_date.strftime("%d/%m/%Y")})'
        except ValueError:
            return jsonify({'success': False, 'message': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
    elif days:
        try:
            days = int(days)
            if days <= 0:
                return jsonify({'success': False, 'message': 'Los días deben ser positivos'}), 400
            new_due_date = old_due_date + timedelta(days=days)
            if days == 1:
                postpone_description = '1 día'
            elif days == 7:
                postpone_description = '1 semana'
            else:
                postpone_description = f'{days} días'
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Valor de días inválido'}), 400
    else:
        return jsonify({'success': False, 'message': 'Debe especificar days o custom_date'}), 400
    
    # Actualizar la tarea
    task.due_date = new_due_date
    task.last_edited_by_id = current_user.id
    task.last_edited_at = datetime.utcnow()
    
    new_due_date_str = new_due_date.strftime('%d/%m/%Y')
    
    # Registrar en el log de actividad
    log_activity(
        user=current_user,
        action='task_postponed',
        description=f'pospuso la tarea "{task.title}" ({postpone_description}): {old_due_date_str} → {new_due_date_str}',
        target_type='task',
        target_id=task.id,
        area_id=task.area_id,
        details=json.dumps({
            'old_due_date': old_due_date_str,
            'new_due_date': new_due_date_str,
            'postpone_type': 'custom' if custom_date_str else 'days',
            'postpone_value': custom_date_str if custom_date_str else days
        })
    )
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'Tarea pospuesta exitosamente',
        'new_due_date': new_due_date_str,
        'old_due_date': old_due_date_str
    })

@main_bp.route('/api/tasks/<int:task_id>/transfer', methods=['POST'])
@login_required
def api_transfer_task(task_id):
    """
    Pasar una tarea a otro usuario.
    - Reemplaza los asignados actuales con el nuevo usuario
    - Cambia el estado a 'Pending'
    - Limpia los campos de tracking (started_at, in_review_at, etc.)
    - Registra en el historial quién la pasó y a quién
    
    Body JSON:
        to_user_id: int (ID del usuario destino)
        comment: string (opcional, motivo del pase)
    """
    task = Task.query.get_or_404(task_id)
    
    # Validar que el usuario tiene permiso
    is_assignee = current_user in task.assignees
    is_creator = task.creator_id == current_user.id
    is_admin = current_user.is_admin
    is_supervisor = current_user.role == 'supervisor' and task.area_id in [a.id for a in current_user.areas]
    
    if not (is_assignee or is_creator or is_admin or is_supervisor):
        return jsonify({'success': False, 'message': 'No tienes permiso para pasar esta tarea'}), 403
    
    # No permitir pasar tareas completadas o anuladas
    if task.status in ['Completed', 'Anulado']:
        return jsonify({'success': False, 'message': 'No se puede pasar una tarea completada o anulada'}), 400
    
    data = request.get_json()
    to_user_id = data.get('to_user_id')
    comment = data.get('comment', '').strip()
    
    if not to_user_id:
        return jsonify({'success': False, 'message': 'Debe especificar el usuario destino'}), 400
    
    # Obtener usuario destino
    to_user = User.query.get(to_user_id)
    if not to_user:
        return jsonify({'success': False, 'message': 'Usuario destino no encontrado'}), 404
    
    # Guardar información anterior para el log
    old_assignees = [u.full_name for u in task.assignees]
    old_status = task.status
    
    # Limpiar asignados y agregar el nuevo
    task.assignees.clear()
    task.assignees.append(to_user)
    
    # Cambiar estado a Pending
    task.status = 'Pending'
    
    # Limpiar campos de tracking
    task.started_at = None
    task.started_by_id = None
    task.in_review_at = None
    task.in_review_by_id = None
    task.completed_at = None
    task.completed_by_id = None
    task.approved_at = None
    task.approved_by_id = None
    
    # Actualizar último editor
    task.last_edited_by_id = current_user.id
    task.last_edited_at = datetime.utcnow()
    
    # Registrar transición de estado si cambió
    if old_status != 'Pending':
        from models import StatusTransition
        transition = StatusTransition(
            task_id=task.id,
            from_status=old_status,
            to_status='Pending',
            changed_by_id=current_user.id,
            comment=f'Pase de tarea a {to_user.full_name}'
        )
        db.session.add(transition)
    
    # Registrar en el log de actividad
    log_activity(
        user=current_user,
        action='task_transferred',
        description=f'pasó la tarea "{task.title}" a {to_user.full_name}',
        target_type='task',
        target_id=task.id,
        area_id=task.area_id,
        details=json.dumps({
            'from_users': old_assignees,
            'to_user': to_user.full_name,
            'to_user_id': to_user.id,
            'old_status': old_status,
            'comment': comment if comment else None
        })
    )
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'Tarea pasada a {to_user.full_name}',
        'new_assignee': to_user.full_name
    })

@main_bp.route('/api/users', methods=['GET'])
@login_required
def api_get_users():
    """
    Obtener lista de usuarios para el selector de pase de tareas.
    Devuelve todos los usuarios activos ordenados por nombre con sus áreas.
    """
    users = User.query.order_by(User.full_name).all()
    
    users_data = [{
        'id': u.id,
        'username': u.username,
        'full_name': u.full_name,
        'role': u.role,
        'area_ids': [a.id for a in u.areas]
    } for u in users]
    
    return jsonify({
        'success': True,
        'users': users_data
    })

@main_bp.route('/api/areas', methods=['GET'])
@login_required
def api_get_areas():
    """
    Obtener lista de áreas para el selector de pase de tareas.
    """
    from models import Area
    areas = Area.query.order_by(Area.name).all()
    
    areas_data = [{
        'id': a.id,
        'name': a.name,
        'color': a.color
    } for a in areas]
    
    return jsonify({
        'success': True,
        'areas': areas_data
    })

# Tags Routes - Append to routes.py

# --- Tags Management Routes ---
@main_bp.route('/tags')
@login_required
def tags():
    """Tags management page - filtered by area"""
    if current_user.is_admin:
        all_tags = Tag.query.order_by(Tag.name).all()
    else:
        user_area_ids = [a.id for a in current_user.areas]
        if user_area_ids:
            all_tags = Tag.query.filter(Tag.area_id.in_(user_area_ids)).order_by(Tag.name).all()
        else:
            all_tags = []
    return render_template('tags.html', tags=all_tags)

@main_bp.route('/api/tags', methods=['GET'])
@login_required
def api_get_tags():
    """Get all tags - filtered by area"""
    if current_user.is_admin:
        tags = Tag.query.order_by(Tag.name).all()
    else:
        user_area_ids = [a.id for a in current_user.areas]
        if user_area_ids:
            tags = Tag.query.filter(Tag.area_id.in_(user_area_ids)).order_by(Tag.name).all()
        else:
            tags = []
    return jsonify({
        'tags': [{
            'id': tag.id,
            'name': tag.name,
            'color': tag.color
        } for tag in tags]
    })

@main_bp.route('/api/tags', methods=['POST'])
@login_required
def api_create_tag():
    """Create a new tag - accessible to all users, assigns area automatically"""
    
    data = request.get_json()
    name = data.get('name', '').strip()
    color = data.get('color', '#2563eb')
    
    if not name:
        return jsonify({'success': False, 'message': 'El nombre es requerido'}), 400
    
    # Check if tag already exists
    existing = Tag.query.filter_by(name=name).first()
    if existing:
        return jsonify({'success': False, 'message': 'Este tag ya existe'}), 400
    
    # Determine area_id based on user
    # Non-admins get their first area automatically
    area_id = None
    if not current_user.is_admin and current_user.areas:
        area_id = current_user.areas[0].id
    
    new_tag = Tag(
        name=name,
        color=color,
        created_by_id=current_user.id,
        area_id=area_id
    )
    
    db.session.add(new_tag)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'tag': {
            'id': new_tag.id,
            'name': new_tag.name,
            'color': new_tag.color
        }
    })

@main_bp.route('/api/tags/<int:tag_id>', methods=['PUT'])
@login_required
def api_update_tag(tag_id):
    """Update a tag - accessible to all users"""
    
    tag = Tag.query.get_or_404(tag_id)
    data = request.get_json()
    
    name = data.get('name', '').strip()
    color = data.get('color')
    
    if name:
        # Check if new name conflicts with another tag
        existing = Tag.query.filter(Tag.name == name, Tag.id != tag_id).first()
        if existing:
            return jsonify({'success': False, 'message': 'Este nombre ya está en uso'}), 400
        tag.name = name
    
    if color:
        tag.color = color
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'tag': {
            'id': tag.id,
            'name': tag.name,
            'color': tag.color
        }
    })

@main_bp.route('/api/tags/<int:tag_id>', methods=['DELETE'])
@login_required
def api_delete_tag(tag_id):
    """Delete a tag - accessible to all users"""
    
    tag = Tag.query.get_or_404(tag_id)
    
    db.session.delete(tag)
    db.session.commit()
    
    return jsonify({'success': True})

# --- Reports Routes ---
@main_bp.route('/reports')
@login_required
def reports():
    from models import Area
    
    # --- CHECK PERMISSION TO VIEW REPORTS ---
    if not current_user.can_see_reports():
        flash('No tienes acceso a los reportes.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Filter users by area for non-admins
    if current_user.is_admin:

        users = User.query.all()
        available_areas = Area.query.order_by(Area.name).all()
        show_area_filter = True
    else:
        # Non-admins see only users in their areas
        user_area_ids = [a.id for a in current_user.areas]
        if user_area_ids:
            users = [u for u in User.query.all() if any(area in u.areas for area in current_user.areas)]
        else:
            users = []
        available_areas = current_user.areas
        show_area_filter = False
    
    # Filter tags by area
    if current_user.is_admin:
        tags = Tag.query.order_by(Tag.name).all()
    else:
        user_area_ids = [a.id for a in current_user.areas]
        if user_area_ids:
            tags = Tag.query.filter(Tag.area_id.in_(user_area_ids)).order_by(Tag.name).all()
        else:
            tags = []
            
    return render_template('reports.html', 
                          users=users, 
                          tags=tags,
                          all_areas=available_areas,
                          show_area_filter=show_area_filter)

@main_bp.route('/api/reports/data', methods=['POST'])
@login_required
def reports_data():
    # --- CHECK PERMISSION TO VIEW REPORTS ---
    if not current_user.can_see_reports():
        return jsonify({'error': 'No tienes acceso a los reportes'}), 403
    
    data = request.get_json()

    user_ids = data.get('user_ids', [])
    tag_ids = data.get('tag_ids', []) # New filter
    status_filter = data.get('status') # New filter
    area_filter = data.get('area')  # Area filter (admin only)
    start_date_str = data.get('start_date')
    end_date_str = data.get('end_date')
    
    # Base query - ALWAYS exclude 'Anulado' and blocked tasks from reports
    query = Task.query.options(joinedload(Task.assignees), joinedload(Task.tags)).filter(
        Task.status != 'Anulado',
        Task.enabled == True  # Only include enabled tasks in reports
    )
    
    # --- FILTER BY AREA ---
    if current_user.is_admin:
        # Admin can filter by specific area
        if area_filter and area_filter != 'all':
            query = query.filter(Task.area_id == int(area_filter))
    else:
        # Non-admins see only tasks from their specific areas (NOT including NULL areas)
        user_area_ids = [a.id for a in current_user.areas]
        if user_area_ids:
            query = query.filter(Task.area_id.in_(user_area_ids))
        else:
            # User has no areas - show nothing
            query = query.filter(Task.area_id == -1)
    
    # Filter by users if provided
    if user_ids:
        query = query.filter(Task.assignees.any(User.id.in_(user_ids)))
        
    # Filter by tags if provided
    if tag_ids:
        query = query.filter(Task.tags.any(Tag.id.in_(tag_ids)))
        
    # Filter by status if provided
    if status_filter and status_filter != 'All':
        if status_filter == 'Overdue':
            # Overdue = Pending tasks with due_date < today
            today_date = date.today()
            query = query.filter(
                Task.status == 'Pending',
                db.func.date(Task.due_date) < today_date
            )
        else:
            query = query.filter(Task.status == status_filter)
    
    # Filter by date range
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        end_date = end_date.replace(hour=23, minute=59, second=59)
        query = query.filter(Task.due_date >= start_date, Task.due_date <= end_date)
        
    tasks = query.all()
    
    # --- Aggregation ---
    
    # 1. Stats per User
    user_stats = []
    target_users = User.query.filter(User.id.in_(user_ids)).all() if user_ids else User.query.all()
    
    for user in target_users:
        user_tasks = [t for t in tasks if user in t.assignees]
        # Only skip if we are filtering by specific users and this user has no tasks
        # But if we are showing all users, we might want to show 0s? 
        # Let's keep existing logic: show all target_users
        
        completed = sum(1 for t in user_tasks if t.status == 'Completed')
        pending = len(user_tasks) - completed
        user_stats.append({
            'name': user.full_name,
            'completed': completed,
            'pending': pending
        })
        
    # 2. Global Status
    global_completed = sum(1 for t in tasks if t.status == 'Completed')
    global_pending = len(tasks) - global_completed
    
    # --- Trends (Time-based) ---
    # We need a date range for the X-axis
    t_start = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else datetime.now() - timedelta(days=30)
    
    if end_date_str:
        t_end = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    else:
        t_end = datetime.now()
    
    # Generate list of dates
    date_labels = []
    current = t_start
    while current <= t_end:
        date_labels.append(current.strftime('%Y-%m-%d'))
        current += timedelta(days=1)
        
    # 3. Global Trend (Completed tasks over time)
    # Base trend query needs to respect the same filters as above? 
    # Usually "Trend" implies "History", so we look at completed_at.
    # But we should respect the User/Tag filters.
    
    trend_query = Task.query.filter(Task.status == 'Completed', Task.completed_at.isnot(None))
    
    if user_ids:
        trend_query = trend_query.filter(Task.assignees.any(User.id.in_(user_ids)))
    if tag_ids:
        trend_query = trend_query.filter(Task.tags.any(Tag.id.in_(tag_ids)))
        
    trend_query = trend_query.filter(Task.completed_at >= t_start, Task.completed_at <= t_end)
    completed_tasks_trend = trend_query.all()
    
    # Group global completed by date
    global_date_counts = {d: 0 for d in date_labels}
    for t in completed_tasks_trend:
        d_str = t.completed_at.strftime('%Y-%m-%d')
        if d_str in global_date_counts:
            global_date_counts[d_str] += 1
            
    global_trend_data = {
        'dates': date_labels,
        'completed_counts': [global_date_counts[d] for d in date_labels]
    }

    # 4. Employee Trend (Line chart per employee)
    # We reuse 'completed_tasks_trend' but group by user
    employee_trend_datasets = []
    # If too many users, chart gets messy. Maybe limit to top 5 or selected?
    # For now, do all selected users (or all if none selected)
    
    for user in target_users:
        # Get tasks for this user from the trend set
        user_trend_tasks = [t for t in completed_tasks_trend if user in t.assignees]
        
        # If user has 0 completed tasks in this period, maybe skip? Or show flat line?
        # Let's show flat line if they are in the filter list.
        
        u_date_counts = {d: 0 for d in date_labels}
        for t in user_trend_tasks:
            d_str = t.completed_at.strftime('%Y-%m-%d')
            if d_str in u_date_counts:
                u_date_counts[d_str] += 1
        
        # Only add dataset if there's at least one task or if explicitly filtered?
        # Let's add all to be safe, frontend can hide if needed.
        # Optimization: if sum is 0, maybe skip to keep chart clean?
        if sum(u_date_counts.values()) > 0 or user_ids:
             employee_trend_datasets.append({
                'label': user.full_name,
                'data': [u_date_counts[d] for d in date_labels],
                'fill': False
             })

    # 5. Tag Trend (Line chart per tag)
    # We need to query tags. If tag_ids filter is on, use those. Else all tags?
    # If all tags, that might be too many lines. Let's use top 5 active tags or just the filtered ones.
    
    tag_trend_datasets = []
    target_tags = Tag.query.filter(Tag.id.in_(tag_ids)).all() if tag_ids else Tag.query.all()
    
    # If no tag filter, maybe we only show tags that actually have data in this period to avoid clutter?
    # Or just all tags. Let's try all tags but skip empty ones.
    
    for tag in target_tags:
        # Filter trend tasks that have this tag
        tag_trend_tasks = [t for t in completed_tasks_trend if tag in t.tags]
        
        t_date_counts = {d: 0 for d in date_labels}
        for t in tag_trend_tasks:
            d_str = t.completed_at.strftime('%Y-%m-%d')
            if d_str in t_date_counts:
                t_date_counts[d_str] += 1
                
        if sum(t_date_counts.values()) > 0 or tag_ids:
            tag_trend_datasets.append({
                'label': tag.name,
                'borderColor': tag.color, # Use tag color!
                'data': [t_date_counts[d] for d in date_labels],
                'fill': False
            })

    # 6. Detailed KPIs
    kpis = calculate_kpis(tasks, global_completed)

    return jsonify({
        'user_stats': user_stats,
        'global_stats': {'completed': global_completed, 'pending': global_pending},
        'trend': global_trend_data,
        'employee_trend': employee_trend_datasets,
        'tag_trend': tag_trend_datasets,
        'kpis': kpis
    })

@main_bp.route('/api/reports/calculate_difference', methods=['POST'])
@login_required
def api_calculate_difference():
    data = request.get_json()
    tag_a_ids = data.get('tag_a_ids', [])
    tag_b_ids = data.get('tag_b_ids', [])
    
    # Fallback for old single ID calls
    if not tag_a_ids and 'tag_a_id' in data:
        tag_a_ids = [data['tag_a_id']]
    if not tag_b_ids and 'tag_b_id' in data:
        tag_b_ids = [data['tag_b_id']]
        
    start_date_str = data.get('start_date')
    end_date_str = data.get('end_date')
    
    def get_group_time(t_ids):
        if not t_ids: return 0
        q = Task.query.filter(Task.status != 'Anulado')
        # Filter tasks that have ANY of the tags in t_ids
        q = q.filter(Task.tags.any(Tag.id.in_(t_ids)))
        
        if start_date_str and end_date_str:
            try:
                s_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                e_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                q = q.filter(Task.due_date >= s_date, Task.due_date <= e_date)
            except ValueError:
                pass
                
        tasks = q.all()
        # Sum time_spent (Task objects are unique due to SQLAlchemys identity map in this query context)
        total = sum(t.time_spent for t in tasks if t.time_spent)
        return total

    time_a = get_group_time(tag_a_ids)
    time_b = get_group_time(tag_b_ids)
    
    diff = time_a - time_b
    
    abs_diff = abs(diff)
    hours = int(abs_diff / 60)
    minutes = int(abs_diff % 60)
    formatted = f"{hours}h {minutes}m"
    if diff < 0:
        formatted = f"- {formatted}"
    
    return jsonify({
        'time_a': time_a,
        'time_b': time_b,
        'diff': diff,
        'formatted': formatted
    })

@main_bp.route('/reports/export', methods=['POST'])
@login_required
def export_report():
    # Get filters from form data
    user_ids_str = request.form.get('user_ids')
    tag_ids_str = request.form.get('tag_ids') # New
    status_filter = request.form.get('status') # New
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    include_kpis_str = request.form.get('include_kpis')
    include_kpis = include_kpis_str == 'true'
    
    import json
    user_ids = json.loads(user_ids_str) if user_ids_str else []
    tag_ids = json.loads(tag_ids_str) if tag_ids_str else [] # New
    
    # Fetch data - ALWAYS exclude 'Anulado' and blocked tasks from reports
    query = Task.query.filter(Task.status != 'Anulado', Task.enabled == True)
    if user_ids:
        query = query.filter(Task.assignees.any(User.id.in_(user_ids)))
        
    if tag_ids: # New
        query = query.filter(Task.tags.any(Tag.id.in_(tag_ids)))
        
    if status_filter and status_filter != 'All': # New
        query = query.filter(Task.status == status_filter)
        
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        query = query.filter(Task.due_date >= start_date, Task.due_date <= end_date)
        
    tasks = query.order_by(Task.due_date).all()
    
    # Prepare data for charts
    
    # Generate PDF
    from pdf_utils import generate_report_pdf
    
    # Stats calculation:
    target_users = User.query.filter(User.id.in_(user_ids)).all() if user_ids else User.query.all()
    user_stats = []
    for user in target_users:
        u_tasks = [t for t in tasks if user in t.assignees]
        # If filtering by users, only show those users. If not, show all.
        # Logic: if user_ids is set, we only iterate those. If not, we iterate all.
        # But if we filter by tag, a user might have 0 tasks with that tag.
        # Should we show them? Yes, with 0.
        
        completed = sum(1 for t in u_tasks if t.status == 'Completed')
        user_stats.append({'name': user.full_name, 'completed': completed, 'pending': len(u_tasks)-completed})
        
    # Global Stats
    global_completed = sum(1 for t in tasks if t.status == 'Completed')
    global_pending = len(tasks) - global_completed
    
    # Trend Data
    t_start = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else datetime.now() - timedelta(days=30)
    
    if end_date_str:
        t_end = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    else:
        t_end = datetime.now()
    
    # Filter completed tasks for trend (respecting all filters)
    # We can just filter 'tasks' list since it already has all filters applied
    completed_tasks_trend = [t for t in tasks if t.status == 'Completed' and t.completed_at and t_start <= t.completed_at <= t_end]
    
    date_counts = {}
    current = t_start
    while current <= t_end:
        date_str = current.strftime('%Y-%m-%d')
        date_counts[date_str] = 0
        current += timedelta(days=1)
        
    for t in completed_tasks_trend:
        d_str = t.completed_at.strftime('%Y-%m-%d')
        if d_str in date_counts:
            date_counts[d_str] += 1
            
    trend_data = {
        'dates': list(date_counts.keys()),
        'completed_counts': list(date_counts.values())
    }
    
    # --- Employee Trend (for PDF) ---
    employee_trend_datasets = []
    for user in target_users:
        user_trend_tasks = [t for t in completed_tasks_trend if user in t.assignees]
        u_date_counts = {d: 0 for d in date_counts.keys()}
        for t in user_trend_tasks:
            d_str = t.completed_at.strftime('%Y-%m-%d')
            if d_str in u_date_counts:
                u_date_counts[d_str] += 1
        
        if sum(u_date_counts.values()) > 0 or user_ids:
             employee_trend_datasets.append({
                'label': user.full_name,
                'data': [u_date_counts[d] for d in date_counts.keys()]
             })

    # --- Tag Trend (for PDF) ---
    tag_trend_datasets = []
    target_tags = Tag.query.filter(Tag.id.in_(tag_ids)).all() if tag_ids else Tag.query.all()
    
    for tag in target_tags:
        tag_trend_tasks = [t for t in completed_tasks_trend if tag in t.tags]
        t_date_counts = {d: 0 for d in date_counts.keys()}
        for t in tag_trend_tasks:
            d_str = t.completed_at.strftime('%Y-%m-%d')
            if d_str in t_date_counts:
                t_date_counts[d_str] += 1
                
        if sum(t_date_counts.values()) > 0 or tag_ids:
            tag_trend_datasets.append({
                'label': tag.name,
                'color': tag.color,
                'data': [t_date_counts[d] for d in date_counts.keys()]
            })
    
    # Filter Info for PDF
    filter_info = {
        'users': [u.full_name for u in target_users] if user_ids else ['Todos'],
        'tags': [t.name for t in Tag.query.filter(Tag.id.in_(tag_ids)).all()] if tag_ids else ['Todas'],
        'status': status_filter if status_filter and status_filter != 'All' else 'Todos'
    }
        
    report_data = {
        'tasks': tasks,
        'user_stats': user_stats,
        'global_stats': {'completed': global_completed, 'pending': global_pending},
        'trend': trend_data,
        'employee_trend': employee_trend_datasets, # New
        'tag_trend': tag_trend_datasets, # New
        'start_date': start_date_str,
        'end_date': end_date_str,
        'filters': filter_info
    }
    
    if include_kpis:
        report_data['kpis'] = calculate_kpis(tasks, global_completed)
        
    # Calculate difference if tags provided
    # Input names are 'diff_tag_a' and 'diff_tag_b' which now contain JSON lists
    import json
    diff_tag_a_json = request.form.get('diff_tag_a')
    diff_tag_b_json = request.form.get('diff_tag_b')
    
    if diff_tag_a_json and diff_tag_b_json:
        try:
            tag_a_ids = json.loads(diff_tag_a_json)
            tag_b_ids = json.loads(diff_tag_b_json)
            
            # Helper to calculate group time (duplicated logic, should be refactored)
            def get_group_time_export(t_ids):
                if not t_ids: return 0, "0h 0m"
                q = Task.query.filter(Task.status != 'Anulado')
                q = q.filter(Task.tags.any(Tag.id.in_(t_ids)))
                if start_date_str and end_date_str:
                    s_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                    e_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                    q = q.filter(Task.due_date >= s_date, Task.due_date <= e_date)
                ts = q.all()
                total_min = sum(t.time_spent for t in ts if t.time_spent)
                
                h = int(total_min / 60)
                m = int(total_min % 60)
                return total_min, f"{h}h {m}m"

            # Get objects for names
            tags_a = Tag.query.filter(Tag.id.in_(tag_a_ids)).all()
            tags_b = Tag.query.filter(Tag.id.in_(tag_b_ids)).all()
            
            # Format names (e.g. "TagA, TagB")
            name_a = ", ".join([t.name for t in tags_a])
            name_b = ", ".join([t.name for t in tags_b])
            
            if tags_a and tags_b:
                time_a, str_a = get_group_time_export(tag_a_ids)
                time_b, str_b = get_group_time_export(tag_b_ids)
                
                diff = time_a - time_b
                abs_diff = abs(diff)
                d_h = int(abs_diff / 60)
                d_m = int(abs_diff % 60)
                diff_str = f"{d_h}h {d_m}m"
                if diff < 0: diff_str = "- " + diff_str
                
                report_data['diff_calc'] = {
                    'tag_a': {'name': name_a, 'time': str_a},
                    'tag_b': {'name': name_b, 'time': str_b},
                    'result': diff_str
                }
        except Exception as e:
            print(f"Error calculating diff for export: {e}")
    
    pdf = generate_report_pdf(report_data)
    
    response = make_response(pdf.output(dest='S').encode('latin-1'))
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_avanzado_{date.today()}.pdf'
    
    return response

def calculate_kpis(tasks, global_completed):
    # Total Tasks
    kpi_total = len(tasks)
    
    # Completion Rate
    kpi_completion_rate = round((global_completed / kpi_total * 100), 1) if kpi_total > 0 else 0
    
    # Overdue Tasks (Pending and due_date < now)
    now = datetime.now()
    kpi_overdue = sum(1 for t in tasks if t.status == 'Pending' and t.due_date < now)
    
    # Time calculations (using time_spent field in minutes)
    tasks_with_time = [t for t in tasks if t.time_spent and t.time_spent > 0]
    
    if tasks_with_time:
        total_minutes = sum(t.time_spent for t in tasks_with_time)
        avg_minutes = total_minutes / len(tasks_with_time)
        
        # Format avg_time: show hours if >= 60 min, otherwise minutes
        if avg_minutes >= 60:
            hours = avg_minutes / 60
            kpi_avg_time = f"{round(hours, 1)} horas"
        else:
            kpi_avg_time = f"{round(avg_minutes)} min"
        
        # Format total_time: show both minutes and hours
        total_hours = total_minutes / 60
        kpi_total_time = f"{int(total_minutes)} min ({round(total_hours, 1)} hs)"
    else:
        kpi_avg_time = "N/A"
        kpi_total_time = "0 min"

    return {
        'total': kpi_total,
        'completion_rate': kpi_completion_rate,
        'overdue': kpi_overdue,
        'avg_time': kpi_avg_time,
        'total_time': kpi_total_time
    }

# --- Excel Import Routes ---

def download_import_template():
    """Download the Excel template for importing tasks"""
    if not current_user.can_create_tasks():
        flash('No tienes permiso para importar tareas.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    import os
    template_path = os.path.join(os.path.dirname(__file__), 'static', 'plantilla_tareas.xlsx')
    
    if not os.path.exists(template_path):
        flash('Plantilla no encontrada. Contacte al administrador.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    with open(template_path, 'rb') as f:
        data = f.read()
    
    response = make_response(data)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=plantilla_tareas.xlsx'
    return response


def import_tasks():
    """Import tasks from an uploaded Excel file"""
    # Check permission first
    if not current_user.can_create_tasks():
        flash('No tienes permiso para importar tareas.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Determine user's area for automatic assignment (for non-admins)
    is_limited_creator = current_user.role in ['supervisor', 'usuario_plus']
    user_area_id = None
    
    if is_limited_creator:
        if current_user.areas:
            user_area_id = current_user.areas[0].id
        else:
            flash('No tienes un área asignada. Contacta a un administrador.', 'danger')
            return redirect(url_for('main.dashboard'))
    
    from openpyxl import load_workbook
    
    if 'file' not in request.files:
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    file = request.files['file']
    
    if file.filename == '':
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('El archivo debe ser un Excel (.xlsx o .xls).', 'danger')
        return redirect(url_for('main.dashboard'))
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        # Skip header row
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        created_count = 0
        errors = []
        
        for row_num, row in enumerate(rows, start=2):
            if not row or not row[0]:  # Skip empty rows
                continue
            
            # Columnas según la plantilla Excel:
            # 0: Título, 1: Descripción, 2: Prioridad, 3: Fecha Inicio, 4: Hora Inicio,
            # 5: Fecha Vencimiento, 6: Hora Vencimiento, 7: Asignados, 8: Etiquetas,
            # 9: ID Proceso, 10: Estado, 11: Completado Por
            titulo = row[0] if len(row) > 0 else None
            descripcion = row[1] if len(row) > 1 else ''
            prioridad = row[2] if len(row) > 2 else 'Normal'
            fecha_inicio_str = row[3] if len(row) > 3 else None
            hora_inicio_str = row[4] if len(row) > 4 else None
            fecha_vencimiento_str = row[5] if len(row) > 5 else None
            hora_vencimiento_str = row[6] if len(row) > 6 else None
            asignados_str = row[7] if len(row) > 7 else ''
            etiquetas_str = row[8] if len(row) > 8 else ''
            proceso_id_str = row[9] if len(row) > 9 else ''
            estado_str = row[10] if len(row) > 10 else 'Pendiente'
            completado_por_str = row[11] if len(row) > 11 else ''
            
            # Validate required fields
            if not titulo:
                errors.append(f'Fila {row_num}: Título requerido')
                continue
            
            if not fecha_vencimiento_str:
                errors.append(f'Fila {row_num}: Fecha de vencimiento requerida')
                continue
            
            # Helper function to parse date+time
            def parse_datetime(fecha_val, hora_val, default_hour=0, default_minute=0):
                """Parse date and optional time, returning a datetime object"""
                if not fecha_val:
                    return None
                
                # Parse date
                if isinstance(fecha_val, datetime):
                    base_date = fecha_val
                elif isinstance(fecha_val, date):
                    base_date = datetime.combine(fecha_val, datetime.min.time())
                else:
                    fecha_str_clean = str(fecha_val).strip()
                    # Try different date formats
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                        try:
                            base_date = datetime.strptime(fecha_str_clean, fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        return None  # Could not parse date
                
                # Parse time if provided
                if hora_val:
                    if isinstance(hora_val, datetime):
                        hour, minute = hora_val.hour, hora_val.minute
                    elif isinstance(hora_val, time):
                        hour, minute = hora_val.hour, hora_val.minute
                    else:
                        hora_str_clean = str(hora_val).strip()
                        # Try different time formats
                        for fmt in ['%H:%M', '%H:%M:%S', '%I:%M %p']:
                            try:
                                parsed_time = datetime.strptime(hora_str_clean, fmt)
                                hour, minute = parsed_time.hour, parsed_time.minute
                                break
                            except ValueError:
                                continue
                        else:
                            hour, minute = default_hour, default_minute
                else:
                    hour, minute = default_hour, default_minute
                
                return base_date.replace(hour=hour, minute=minute, second=0, microsecond=0)
            
            # Parse due date (required) - Default to 18:00 (end of workday)
            try:
                due_date = parse_datetime(fecha_vencimiento_str, hora_vencimiento_str, default_hour=18, default_minute=0)
                if not due_date:
                    errors.append(f'Fila {row_num}: Formato de fecha de vencimiento inválido')
                    continue
            except Exception:
                errors.append(f'Fila {row_num}: Error al procesar fecha de vencimiento')
                continue
            
            # Parse start date (optional) - Default to 08:00 (start of workday)
            planned_start_date = None
            if fecha_inicio_str:
                try:
                    planned_start_date = parse_datetime(fecha_inicio_str, hora_inicio_str, default_hour=8, default_minute=0)
                except Exception:
                    pass  # Ignore invalid start dates
            
            # Validate priority
            if prioridad not in ['Normal', 'Media', 'Urgente']:
                prioridad = 'Normal'
            
            # Parse status
            status = 'Pending'  # Default
            if estado_str:
                estado_lower = str(estado_str).strip().lower()
                if estado_lower in ['completada', 'completed', 'completado', 'c']:
                    status = 'Completed'
                elif estado_lower in ['pendiente', 'pending', 'p']:
                    status = 'Pending'
            
            # Parse completed_by user (optional)
            completed_by_user = None
            if completado_por_str:
                completed_by_username = str(completado_por_str).strip()
                if completed_by_username:
                    completed_by_user = User.query.filter_by(username=completed_by_username).first()
                    if not completed_by_user:
                        errors.append(f'Fila {row_num}: Usuario "{completed_by_username}" (Completado Por) no encontrado')
            
            # Parse assignees
            assignee_usernames = [u.strip() for u in str(asignados_str).split(',') if u.strip()]
            assignees = []
            for username in assignee_usernames:
                user = User.query.filter_by(username=username).first()
                if user:
                    assignees.append(user)
                else:
                    errors.append(f'Fila {row_num}: Usuario "{username}" no encontrado')
            
            if not assignees:
                errors.append(f'Fila {row_num}: Ningún usuario válido asignado')
                continue
            
            # Parse tags (optional)
            tag_names = [t.strip() for t in str(etiquetas_str).split(',') if t.strip()]
            tags = []
            for tag_name in tag_names:
                tag = Tag.query.filter_by(name=tag_name).first()
                if tag:
                    tags.append(tag)
                # Silently ignore non-existing tags (they're optional)
            
            # Parse process ID (optional)
            process_id = None
            if proceso_id_str:
                try:
                    process_id = int(proceso_id_str)
                    # Verify process exists
                    process = Process.query.get(process_id)
                    if not process:
                        errors.append(f'Fila {row_num}: Proceso ID {process_id} no encontrado')
                        process_id = None
                except (ValueError, TypeError):
                    errors.append(f'Fila {row_num}: ID de Proceso inválido')
            
            # Create task
            new_task = Task(
                title=str(titulo),
                description=str(descripcion) if descripcion else '',
                priority=prioridad,
                due_date=due_date,
                planned_start_date=planned_start_date,
                creator_id=current_user.id,
                status=status,
                area_id=user_area_id,  # For non-admins, assigns to their area; for admins, None (inherits default)
                process_id=process_id  # Link to process if provided
            )
            
            # Set completed info if status is Completed
            if status == 'Completed':
                # Use specified user, or fallback to current user
                new_task.completed_by_id = completed_by_user.id if completed_by_user else current_user.id
                new_task.completed_at = now_utc()
            
            for user in assignees:
                new_task.assignees.append(user)
            
            for tag in tags:
                new_task.tags.append(tag)
            
            db.session.add(new_task)
            created_count += 1
        
        db.session.commit()
        
        # Log activity
        if created_count > 0:
            log_activity(
                user=current_user,
                action='tasks_imported',
                description=f'importó {created_count} tareas desde Excel',
                area_id=user_area_id or (current_user.areas[0].id if current_user.areas else None)
            )
        
        if created_count > 0:
            flash(f'Se importaron {created_count} tareas exitosamente.', 'success')
        
        if errors:
            error_msg = 'Errores encontrados: ' + '; '.join(errors[:5])
            if len(errors) > 5:
                error_msg += f' ... y {len(errors) - 5} más.'
            flash(error_msg, 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al procesar el archivo: {str(e)}', 'danger')
    
    return redirect(url_for('main.dashboard'))

# --- Task Template Routes ---
@main_bp.route('/templates', methods=['GET', 'POST'])
@login_required
def manage_templates():
    """View and create task templates - assigns area automatically"""
    # Check permission to create tasks (which includes templates)
    if not current_user.can_create_tasks():
        flash('No tienes permiso para gestionar plantillas de tareas.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    if request.method == 'POST':
        name = request.form.get('name')
        title = request.form.get('title')
        description = request.form.get('description', '')
        priority = request.form.get('priority', 'Normal')
        default_days = int(request.form.get('default_days', 0))
        tag_ids = request.form.getlist('tags')
        
        # Parse time_spent
        time_spent_str = request.form.get('time_spent', '0')
        try:
            time_spent = int(time_spent_str) if time_spent_str else 0
        except ValueError:
            time_spent = 0
            
        # Parse start config
        start_days_offset = int(request.form.get('start_days_offset', 0))
        start_time_str = request.form.get('start_time')
        start_time = None
        if start_time_str:
            try:
                start_time = datetime.strptime(start_time_str, '%H:%M').time()
            except ValueError:
                pass
        
        # Check if template name already exists
        if TaskTemplate.query.filter_by(name=name).first():
            flash('Ya existe una plantilla con ese nombre.', 'danger')
            return redirect(url_for('main.manage_templates'))
        
        # Determine area_id based on user
        area_id = None
        if not current_user.is_admin and current_user.areas:
            area_id = current_user.areas[0].id
        
        template = TaskTemplate(
            name=name,
            title=title,
            description=description,
            priority=priority,
            default_days=default_days,
            created_by_id=current_user.id,
            time_spent=time_spent,
            area_id=area_id,
            start_days_offset=start_days_offset,
            start_time=start_time
        )
        
        # Add tags
        for tag_id in tag_ids:
            tag = Tag.query.get(int(tag_id))
            if tag:
                template.tags.append(tag)
        
        db.session.add(template)
        db.session.commit()
        
        # Process subtask templates (if any) - with hierarchy support
        subtask_titles = request.form.getlist('subtask_title[]')
        subtask_descriptions = request.form.getlist('subtask_description[]')
        subtask_priorities = request.form.getlist('subtask_priority[]')
        subtask_priorities = request.form.getlist('subtask_priority[]')
        subtask_days_offsets = request.form.getlist('subtask_days_offset[]')
        subtask_start_days_offsets = request.form.getlist('subtask_start_days_offset[]')
        subtask_start_times = request.form.getlist('subtask_start_time[]')
        subtask_parent_paths = request.form.getlist('subtask_parent_path[]')
        
        # First pass: create all subtasks and store by path
        path_to_subtask = {}
        subtasks_created = 0
        
        for i, st_title in enumerate(subtask_titles):
            if st_title.strip():
                st_description = subtask_descriptions[i] if i < len(subtask_descriptions) else ''
                st_priority = subtask_priorities[i] if i < len(subtask_priorities) else 'Normal'
                st_days_offset = 0
                if i < len(subtask_days_offsets) and subtask_days_offsets[i]:
                    try:
                        st_days_offset = int(subtask_days_offsets[i])
                    except ValueError:
                        st_days_offset = 0
                
                st_start_days_offset = 0
                if i < len(subtask_start_days_offsets) and subtask_start_days_offsets[i]:
                    try:
                        st_start_days_offset = int(subtask_start_days_offsets[i])
                    except ValueError:
                        st_start_days_offset = 0
                
                st_start_time = None
                if i < len(subtask_start_times) and subtask_start_times[i]:
                    try:
                        st_start_time = datetime.strptime(subtask_start_times[i], '%H:%M').time()
                    except ValueError:
                        st_start_time = None
                
                # Get parent path (empty string = top-level, "0" = child of path 0, etc.)
                parent_path = subtask_parent_paths[i] if i < len(subtask_parent_paths) else ''
                
                subtask_template = SubtaskTemplate(
                    template_id=template.id,
                    title=st_title.strip(),
                    description=st_description,
                    priority=st_priority,
                    days_offset=st_days_offset,
                    start_days_offset=st_start_days_offset,
                    start_time=st_start_time,
                    order=i,
                    parent_id=None  # Will be set in second pass
                )
                db.session.add(subtask_template)
                db.session.flush()  # Get ID
                
                # Calculate this subtask's path for children to reference
                # The path is based on the index in the form, matching frontend logic
                if parent_path:
                    current_path = f"{parent_path}.{i}"
                else:
                    current_path = str(i)
                
                path_to_subtask[current_path] = subtask_template
                path_to_subtask[str(i)] = subtask_template  # Also store by simple index
                subtasks_created += 1
        
        # Second pass: link children to parents
        for i, st_title in enumerate(subtask_titles):
            if st_title.strip():
                parent_path = subtask_parent_paths[i] if i < len(subtask_parent_paths) else ''
                if parent_path and parent_path in path_to_subtask:
                    if parent_path:
                        current_path = f"{parent_path}.{i}"
                    else:
                        current_path = str(i)
                    
                    if current_path in path_to_subtask:
                        path_to_subtask[current_path].parent_id = path_to_subtask[parent_path].id
        
        if subtasks_created > 0:
            db.session.commit()
        
        # Log activity
        log_activity(
            user=current_user,
            action='template_created',
            description=f'creó la plantilla "{template.name}"' + (f' con {subtasks_created} subtarea(s)' if subtasks_created > 0 else ''),
            target_type='template',
            target_id=template.id,
            area_id=template.area_id
        )
        
        flash(f'Plantilla "{name}" creada exitosamente.', 'success')
        return redirect(url_for('main.manage_templates'))
    
    # Filter templates and tags by area for non-admins
    if current_user.is_admin:
        templates = TaskTemplate.query.order_by(TaskTemplate.name).all()
        available_tags = Tag.query.order_by(Tag.name).all()
    else:
        user_area_ids = [a.id for a in current_user.areas]
        if user_area_ids:
            # Strict filter - only from user's areas
            templates = TaskTemplate.query.filter(TaskTemplate.area_id.in_(user_area_ids)).order_by(TaskTemplate.name).all()
            available_tags = Tag.query.filter(Tag.area_id.in_(user_area_ids)).order_by(Tag.name).all()
        else:
            templates = []
            available_tags = []
    
    return render_template('manage_templates.html', templates=templates, available_tags=available_tags)

@main_bp.route('/templates/<int:template_id>/delete', methods=['POST'])
@login_required
def delete_template(template_id):
    """Delete a task template"""
    template = TaskTemplate.query.get_or_404(template_id)
    
    # Only creator or admin can delete
    if template.created_by_id != current_user.id and not current_user.is_admin:
        flash('No tienes permiso para eliminar esta plantilla.', 'danger')
        return redirect(url_for('main.manage_templates'))
    
    name = template.name
    db.session.delete(template)
    db.session.commit()
    
    # Log activity
    log_activity(
        user=current_user,
        action='template_deleted',
        description=f'eliminó la plantilla "{name}"',
        target_type='template',
        target_id=template_id,
        area_id=template.area_id
    )
    
    flash(f'Plantilla "{name}" eliminada.', 'success')
    return redirect(url_for('main.manage_templates'))

@main_bp.route('/templates/<int:template_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_template(template_id):
    """Edit an existing task template"""
    template = TaskTemplate.query.get_or_404(template_id)
    
    # Only creator or admin can edit
    if template.created_by_id != current_user.id and not current_user.is_admin:
        flash('No tienes permiso para editar esta plantilla.', 'danger')
        return redirect(url_for('main.manage_templates'))
    
    # Get available tags
    if current_user.is_admin:
        available_tags = Tag.query.order_by(Tag.name).all()
    else:
        user_area_ids = [a.id for a in current_user.areas]
        available_tags = Tag.query.filter(Tag.area_id.in_(user_area_ids)).order_by(Tag.name).all() if user_area_ids else []
    
    if request.method == 'POST':
        template.name = request.form.get('name', template.name)
        template.title = request.form.get('title', template.title)
        template.description = request.form.get('description', '')
        template.priority = request.form.get('priority', 'Normal')
        template.default_days = int(request.form.get('default_days', 1))
        template.start_days_offset = int(request.form.get('start_days_offset', 0))
        
        start_time_str = request.form.get('start_time')
        if start_time_str:
            try:
                template.start_time = datetime.strptime(start_time_str, '%H:%M').time()
            except ValueError:
                template.start_time = None
        else:
            template.start_time = None
            
        template.time_spent = int(request.form.get('time_spent', 0)) if request.form.get('time_spent') else None
        
        # Update tags
        template.tags.clear()
        tag_ids = request.form.getlist('tags')
        for tag_id in tag_ids:
            tag = Tag.query.get(int(tag_id))
            if tag:
                template.tags.append(tag)
        
        # Clear existing subtask templates
        SubtaskTemplate.query.filter_by(template_id=template.id).delete()
        db.session.flush()
        
        # Process new subtask templates (same logic as create)
        subtask_titles = request.form.getlist('subtask_title[]')
        subtask_descriptions = request.form.getlist('subtask_description[]')
        subtask_priorities = request.form.getlist('subtask_priority[]')
        subtask_days_offsets = request.form.getlist('subtask_days_offset[]')
        subtask_start_days_offsets = request.form.getlist('subtask_start_days_offset[]')
        subtask_start_times = request.form.getlist('subtask_start_time[]')
        subtask_parent_paths = request.form.getlist('subtask_parent_path[]')
        
        path_to_subtask = {}
        subtasks_created = 0
        
        for i, st_title in enumerate(subtask_titles):
            if st_title.strip():
                st_description = subtask_descriptions[i] if i < len(subtask_descriptions) else ''
                st_priority = subtask_priorities[i] if i < len(subtask_priorities) else 'Normal'
                st_days_offset = 0
                if i < len(subtask_days_offsets) and subtask_days_offsets[i]:
                    try:
                        st_days_offset = int(subtask_days_offsets[i])
                    except ValueError:
                        st_days_offset = 0
                
                st_start_days_offset = 0
                if i < len(subtask_start_days_offsets) and subtask_start_days_offsets[i]:
                    try:
                        st_start_days_offset = int(subtask_start_days_offsets[i])
                    except ValueError:
                        st_start_days_offset = 0
                
                st_start_time = None
                if i < len(subtask_start_times) and subtask_start_times[i]:
                    try:
                        st_start_time = datetime.strptime(subtask_start_times[i], '%H:%M').time()
                    except ValueError:
                        st_start_time = None
                
                parent_path = subtask_parent_paths[i] if i < len(subtask_parent_paths) else ''
                
                subtask_template = SubtaskTemplate(
                    template_id=template.id,
                    title=st_title.strip(),
                    description=st_description,
                    priority=st_priority,
                    days_offset=st_days_offset,
                    start_days_offset=st_start_days_offset,
                    start_time=st_start_time,
                    order=i,
                    parent_id=None
                )
                db.session.add(subtask_template)
                db.session.flush()
                
                if parent_path:
                    current_path = f"{parent_path}.{i}"
                else:
                    current_path = str(i)
                path_to_subtask[current_path] = subtask_template
                path_to_subtask[str(i)] = subtask_template
                subtasks_created += 1
        
        # Second pass: link children to parents
        for i, st_title in enumerate(subtask_titles):
            if st_title.strip():
                parent_path = subtask_parent_paths[i] if i < len(subtask_parent_paths) else ''
                if parent_path and parent_path in path_to_subtask:
                    if parent_path:
                        current_path = f"{parent_path}.{i}"
                    else:
                        current_path = str(i)
                    
                    if current_path in path_to_subtask:
                        path_to_subtask[current_path].parent_id = path_to_subtask[parent_path].id
        
        db.session.commit()
        
        log_activity(
            user=current_user,
            action='template_edited',
            description=f'editó la plantilla "{template.name}"',
            target_type='template',
            target_id=template.id,
            area_id=template.area_id
        )
        
        flash(f'Plantilla "{template.name}" actualizada.', 'success')
        return redirect(url_for('main.manage_templates'))
    
    # Get existing subtask templates for display
    subtask_templates_obj = SubtaskTemplate.query.filter_by(template_id=template.id).order_by(SubtaskTemplate.order).all()
    
    # Convert to list of dicts for JSON serialization (handling start_time objects)
    subtask_templates = []
    for st in subtask_templates_obj:
        subtask_templates.append({
            'id': st.id,
            'title': st.title,
            'description': st.description or '',
            'priority': st.priority,
            'days_offset': st.days_offset,
            'start_days_offset': st.start_days_offset,
            'start_time': st.start_time.strftime('%H:%M') if st.start_time else '',
            'parent_id': st.parent_id,
            'order': st.order
        })
    
    return render_template('edit_template.html', 
                          template=template, 
                          available_tags=available_tags,
                          subtask_templates=subtask_templates)

@main_bp.route('/api/templates/<int:template_id>')
@login_required
def get_template_data(template_id):
    """API to get template data for form autofill"""
    template = TaskTemplate.query.get_or_404(template_id)
    
    # Calculate due date based on default_days
    due_date = date.today() + timedelta(days=template.default_days)
    
    # Get all subtask templates for this template
    subtask_templates = SubtaskTemplate.query.filter_by(template_id=template_id).order_by(SubtaskTemplate.order).all()
    
    # Build subtask data with hierarchy info
    subtasks_data = []
    for st in subtask_templates:
        subtasks_data.append({
            'id': st.id,
            'title': st.title,
            'description': st.description or '',
            'priority': st.priority,
            'days_offset': st.days_offset,
            'start_days_offset': st.start_days_offset,
            'start_time': st.start_time.strftime('%H:%M') if st.start_time else None,
            'parent_id': st.parent_id
        })
    
    return jsonify({
        'title': template.title,
        'description': template.description or '',
        'priority': template.priority,
        'start_days_offset': template.start_days_offset,
        'start_time': template.start_time.strftime('%H:%M') if template.start_time else None,
        'due_date': due_date.strftime('%Y-%m-%d'),
        'tag_ids': [tag.id for tag in template.tags],
        'time_spent': template.time_spent or 0,
        'subtask_templates': subtasks_data
    })


# --- Task Hierarchy Helper Functions ---
def is_descendant(parent_id, potential_child_id):
    """
    Check if potential_child_id is a descendant of parent_id.
    This prevents circular references in the task hierarchy.
    """
    if parent_id == potential_child_id:
        return True
    
    task = Task.query.get(potential_child_id)
    if not task or not task.parent_id:
        return False
    
    # Recursively check if parent_id is an ancestor
    return is_descendant(parent_id, task.parent_id)


# --- Task Hierarchy API Routes ---
@main_bp.route('/api/tasks/search')
@login_required
def api_search_tasks():
    """Search tasks for parent selection modal"""
    query = request.args.get('q', '').strip()
    exclude_id = request.args.get('exclude_id')
    
    if not query:
        return jsonify({'tasks': []})
    
    search_query = Task.query.filter(Task.status != 'Anulado')
    
    if query.isdigit():
        search_query = search_query.filter(
            (Task.id == int(query)) | (Task.title.ilike(f'%{query}%'))
        )
    else:
        search_query = search_query.filter(Task.title.ilike(f'%{query}%'))
    
    if exclude_id:
        try:
            search_query = search_query.filter(Task.id != int(exclude_id))
        except (ValueError, TypeError):
            pass
    
    tasks = search_query.order_by(Task.due_date.desc()).limit(20).all()
    
    result = []
    for task in tasks:
        result.append({
            'id': task.id,
            'title': task.title,
            'priority': task.priority,
            'status': task.status,
            'due_date': task.due_date.strftime('%d/%m/%Y'),
            'assignees': ', '.join([u.full_name for u in task.assignees])
        })
    
    return jsonify({'tasks': result})


@main_bp.route('/api/tasks/<int:task_id>/validate_parent/<int:parent_id>')
@login_required
def api_validate_parent(task_id, parent_id):
    """Validate that parent_id is not a descendant of task_id (prevent circular references)"""
    if task_id == parent_id:
        return jsonify({'valid': False, 'error': 'Una tarea no puede ser su propia tarea padre'})
    
    if is_descendant(task_id, parent_id):
        return jsonify({'valid': False, 'error': 'Referencia circular detectada'})
    
    return jsonify({'valid': True})


# --- Expiration Calendar Routes ---

@main_bp.route('/expirations')
@login_required
def expiration_calendar():
    """Calendario de vencimientos - filtrado por área"""
    from models import Area
    
    period = request.args.get('period', 'all')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    filter_tag = request.args.get('tag_filter')
    filter_area = request.args.get('area')
    
    # Query expirations with eager loading
    query = Expiration.query.options(joinedload(Expiration.tags), joinedload(Expiration.creator))
    
    # --- FILTER BY AREA ---
    # Admins can see all areas and use filter
    if current_user.is_admin:
        if filter_area:
            query = query.filter(Expiration.area_id == int(filter_area))
        available_areas = Area.query.order_by(Area.name).all()
        show_area_filter = True
    else:
        # Non-admins see only expirations from their specific areas (NOT including NULL areas)
        user_area_ids = [a.id for a in current_user.areas]
        if user_area_ids:
            query = query.filter(Expiration.area_id.in_(user_area_ids))
        else:
            # User has no areas - show nothing
            query = query.filter(Expiration.area_id == -1)
        available_areas = current_user.areas
        show_area_filter = False
    
    # Apply date filters
    today = date.today()
    
    if period == 'today':
        query = query.filter(db.func.date(Expiration.due_date) == today)
    elif period == 'week':
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        query = query.filter(db.func.date(Expiration.due_date) >= week_start,
                           db.func.date(Expiration.due_date) <= week_end)
    elif period == 'month':
        month_start = today.replace(day=1)
        if today.month == 12:
            month_end = today.replace(day=31)
        else:
            month_end = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        query = query.filter(db.func.date(Expiration.due_date) >= month_start,
                           db.func.date(Expiration.due_date) <= month_end)
    elif start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            query = query.filter(db.func.date(Expiration.due_date) >= start_date,
                               db.func.date(Expiration.due_date) <= end_date)
        except ValueError:
            pass
    
    # Apply tag filter
    if filter_tag:
        query = query.filter(Expiration.tags.any(id=int(filter_tag)))
    
    expirations = query.order_by(Expiration.due_date.asc()).all()
    available_tags = Tag.query.order_by(Tag.name).all()
    
    # Get event dates for calendar widget (all dates with expirations in current month)
    month_start = today.replace(day=1)
    if today.month == 12:
        month_end = today.replace(day=31)
    else:
        month_end = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
    
    event_dates_query = db.session.query(db.func.date(Expiration.due_date)).filter(
        db.func.date(Expiration.due_date) >= month_start,
        db.func.date(Expiration.due_date) <= month_end
    ).distinct().all()
    event_dates = [d[0].strftime('%Y-%m-%d') for d in event_dates_query if d[0]]
    
    # --- CALENDAR GRID GENERATION ---
    import calendar as cal
    
    # Get month/year from query params or use current month
    view_year = int(request.args.get('year', today.year))
    view_month = int(request.args.get('month', today.month))
    
    # Calculate previous and next month for navigation
    if view_month == 1:
        prev_month, prev_year = 12, view_year - 1
    else:
        prev_month, prev_year = view_month - 1, view_year
    
    if view_month == 12:
        next_month, next_year = 1, view_year + 1
    else:
        next_month, next_year = view_month + 1, view_year
    
    # Generate calendar weeks (list of 7-day lists)
    cal_obj = cal.Calendar(firstweekday=0)  # Monday = 0
    month_days = cal_obj.monthdatescalendar(view_year, view_month)
    
    # Get expirations for the visible calendar range
    if month_days:
        cal_start = month_days[0][0]
        cal_end = month_days[-1][-1]
        
        # Query expirations within calendar view range
        cal_exp_query = Expiration.query.options(joinedload(Expiration.tags)).filter(
            db.func.date(Expiration.due_date) >= cal_start,
            db.func.date(Expiration.due_date) <= cal_end
        )
        
        # Apply same visibility filters
        if not current_user.is_admin:
            user_area_ids = [a.id for a in current_user.areas]
            if user_area_ids:
                cal_exp_query = cal_exp_query.filter(Expiration.area_id.in_(user_area_ids))
            else:
                cal_exp_query = cal_exp_query.filter(Expiration.area_id == -1)
        
        cal_expirations = cal_exp_query.all()
    else:
        cal_expirations = []
    
    # Group expirations by date
    expirations_by_date = {}
    for exp in cal_expirations:
        date_key = exp.due_date.date().isoformat()
        if date_key not in expirations_by_date:
            expirations_by_date[date_key] = []
        expirations_by_date[date_key].append(exp)
    
    # Month names in Spanish
    month_names = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                   'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    return render_template('expiration_calendar.html', 
                          expirations=expirations, 
                          current_period=period, 
                          today=today,
                          available_tags=available_tags,
                          event_dates=event_dates,
                          all_areas=available_areas,
                          show_area_filter=show_area_filter,
                          # New calendar grid data
                          calendar_weeks=month_days,
                          view_month=view_month,
                          view_year=view_year,
                          month_name=month_names[view_month],
                          prev_month=prev_month,
                          prev_year=prev_year,
                          next_month=next_month,
                          next_year=next_year,
                          expirations_by_date=expirations_by_date)


@main_bp.route('/expirations/create', methods=['POST'])
@login_required
def create_expiration():
    """Crear un nuevo vencimiento - asigna área automáticamente"""
    title = request.form.get('title')
    description = request.form.get('description', '')
    due_date_str = request.form.get('due_date')
    
    if not title or not due_date_str:
        flash('Título y fecha de vencimiento son requeridos.', 'danger')
        return redirect(url_for('main.expiration_calendar'))
    
    try:
        due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
    except ValueError:
        flash('Formato de fecha inválido.', 'danger')
        return redirect(url_for('main.expiration_calendar'))
    
    # Determine area_id based on user role
    # Non-admins get their first area automatically
    area_id = None
    if not current_user.is_admin and current_user.areas:
        area_id = current_user.areas[0].id
    
    new_expiration = Expiration(
        title=title,
        description=description,
        due_date=due_date,
        creator_id=current_user.id,
        area_id=area_id
    )
    
    # Add tags
    tag_ids = request.form.getlist('tags')
    for tag_id in tag_ids:
        tag = Tag.query.get(int(tag_id))
        if tag:
            new_expiration.tags.append(tag)
    
    db.session.add(new_expiration)
    db.session.commit()
    
    # Log activity
    log_activity(
        user=current_user,
        action='expiration_created',
        description=f'creó el vencimiento "{new_expiration.title}"',
        target_type='expiration',
        target_id=new_expiration.id,
        area_id=new_expiration.area_id
    )
    
    flash('Vencimiento creado exitosamente.', 'success')
    return redirect(url_for('main.expiration_calendar'))


@main_bp.route('/expirations/<int:expiration_id>/edit', methods=['POST'])
@login_required
def edit_expiration(expiration_id):
    """Editar un vencimiento existente"""
    expiration = Expiration.query.get_or_404(expiration_id)
    
    # Only creator or admin can edit
    if expiration.creator_id != current_user.id and not current_user.is_admin:
        flash('No tienes permiso para editar este vencimiento.', 'danger')
        return redirect(url_for('main.expiration_calendar'))
    
    title = request.form.get('title')
    description = request.form.get('description', '')
    due_date_str = request.form.get('due_date')
    
    if not title or not due_date_str:
        flash('Título y fecha de vencimiento son requeridos.', 'danger')
        return redirect(url_for('main.expiration_calendar'))
    
    try:
        due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
    except ValueError:
        flash('Formato de fecha inválido.', 'danger')
        return redirect(url_for('main.expiration_calendar'))
    
    expiration.title = title
    expiration.description = description
    expiration.due_date = due_date
    
    # Update tags
    tag_ids = request.form.getlist('tags')
    expiration.tags = []
    for tag_id in tag_ids:
        tag = Tag.query.get(int(tag_id))
        if tag:
            expiration.tags.append(tag)
    
    db.session.commit()
    
    # Log activity
    log_activity(
        user=current_user,
        action='expiration_edited',
        description=f'editó el vencimiento "{expiration.title}"',
        target_type='expiration',
        target_id=expiration.id,
        area_id=expiration.area_id
    )
    
    flash('Vencimiento actualizado exitosamente.', 'success')
    return redirect(url_for('main.expiration_calendar'))


@main_bp.route('/expirations/<int:expiration_id>/toggle', methods=['POST'])
@login_required
def toggle_expiration(expiration_id):
    """Marcar vencimiento como completado/pendiente"""
    expiration = Expiration.query.get_or_404(expiration_id)
    
    if expiration.completed:
        expiration.completed = False
        expiration.completed_at = None
    else:
        expiration.completed = True
        expiration.completed_at = now_utc()
    
    db.session.commit()
    
    # Log activity
    action = 'expiration_completed' if expiration.completed else 'expiration_reopened'
    description = f'completó el vencimiento "{expiration.title}"' if expiration.completed else f'reabrió el vencimiento "{expiration.title}"'
    log_activity(
        user=current_user,
        action=action,
        description=description,
        target_type='expiration',
        target_id=expiration.id,
        area_id=expiration.area_id
    )
    
    return jsonify({
        'success': True,
        'expiration_id': expiration.id,
        'completed': expiration.completed
    })


@main_bp.route('/expirations/<int:expiration_id>/delete', methods=['POST'])
@login_required
def delete_expiration(expiration_id):
    """Eliminar un vencimiento"""
    expiration = Expiration.query.get_or_404(expiration_id)
    
    # Only creator or admin can delete
    if expiration.creator_id != current_user.id and not current_user.is_admin:
        return jsonify({'success': False, 'error': 'No tienes permiso para eliminar este vencimiento.'}), 403
    
    db.session.delete(expiration)
    db.session.commit()
    
    return jsonify({'success': True})


@main_bp.route('/api/expirations/<int:expiration_id>')
@login_required
def api_get_expiration(expiration_id):
    """API para obtener datos de un vencimiento para edición"""
    expiration = Expiration.query.get_or_404(expiration_id)
    
    return jsonify({
        'id': expiration.id,
        'title': expiration.title,
        'description': expiration.description or '',
        'due_date': expiration.due_date.strftime('%Y-%m-%d'),
        'tag_ids': [tag.id for tag in expiration.tags],
        'completed': expiration.completed
    })


# --- Recurring Tasks Routes (Admin Only) ---

@main_bp.route('/recurring-tasks')
@login_required
def manage_recurring_tasks():
    """Listar y gestionar tareas recurrentes"""
    # Check permission to create tasks
    if not current_user.can_create_tasks():
        flash('No tienes permiso para gestionar tareas recurrentes.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Filter by area - admins/gerentes see all, others see only their area
    if current_user.can_see_all_areas():
        recurring_tasks = RecurringTask.query.order_by(RecurringTask.created_at.desc()).all()
        users = User.query.order_by(User.full_name).all()
        available_tags = Tag.query.order_by(Tag.name).all()
        templates = TaskTemplate.query.order_by(TaskTemplate.name).all()
    else:
        # Supervisor/usuario_plus: only see recurring tasks from their area
        user_area_ids = [a.id for a in current_user.areas]
        if user_area_ids:
            # Strict filter - only from user's areas
            recurring_tasks = RecurringTask.query.filter(RecurringTask.area_id.in_(user_area_ids)).order_by(RecurringTask.created_at.desc()).all()
            # Only show users from their areas
            users = [u for u in User.query.order_by(User.full_name).all() if any(area in u.areas for area in current_user.areas)]
            available_tags = Tag.query.filter(Tag.area_id.in_(user_area_ids)).order_by(Tag.name).all()
            templates = TaskTemplate.query.filter(TaskTemplate.area_id.in_(user_area_ids)).order_by(TaskTemplate.name).all()
        else:
            recurring_tasks = []
            users = []
            available_tags = []
            templates = []
    
    return render_template('manage_recurring_tasks.html',
                           recurring_tasks=recurring_tasks,
                           users=users,
                           available_tags=available_tags,
                           templates=templates)


@main_bp.route('/recurring-tasks/create', methods=['POST'])
@login_required
def create_recurring_task():
    """Crear una nueva tarea recurrente"""
    # Check permission to create tasks
    if not current_user.can_create_tasks():
        flash('No tienes permiso para crear tareas recurrentes.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    title = request.form.get('title')
    description = request.form.get('description', '')
    priority = request.form.get('priority', 'Normal')
    recurrence_type = request.form.get('recurrence_type')
    days_of_week = request.form.get('days_of_week', '')
    day_of_month_str = request.form.get('day_of_month', '')
    due_time_str = request.form.get('due_time')
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date', '')
    time_spent_str = request.form.get('time_spent', '0')
    assignee_ids = request.form.getlist('assignees')
    tag_ids = request.form.getlist('tags')
    
    # NEW: Optional template_id
    template_id_str = request.form.get('template_id', '')
    template_id = None
    selected_template = None
    if template_id_str:
        try:
            template_id = int(template_id_str)
            selected_template = TaskTemplate.query.get(template_id)
        except ValueError:
            pass
    
    # Validations - title is optional if template is selected
    if not selected_template and not title:
        flash('Título es requerido si no se selecciona una plantilla.', 'danger')
        return redirect(url_for('main.manage_recurring_tasks'))
    
    if not recurrence_type or not due_time_str or not start_date_str:
        flash('Tipo de recurrencia, hora y fecha de inicio son requeridos.', 'danger')
        return redirect(url_for('main.manage_recurring_tasks'))
    
    if not assignee_ids:
        flash('Debes asignar al menos un usuario.', 'danger')
        return redirect(url_for('main.manage_recurring_tasks'))
    
    # Parse time
    try:
        from datetime import time as dt_time
        hour, minute = map(int, due_time_str.split(':'))
        due_time = dt_time(hour, minute)
    except ValueError:
        flash('Formato de hora inválido.', 'danger')
        return redirect(url_for('main.manage_recurring_tasks'))
    
    # Parse dates
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None
    except ValueError:
        flash('Formato de fecha inválido.', 'danger')
        return redirect(url_for('main.manage_recurring_tasks'))
    
    # Parse day_of_month for monthly recurrence
    day_of_month = None
    if recurrence_type == 'monthly' and day_of_month_str:
        try:
            day_of_month = int(day_of_month_str)
            if day_of_month < 1 or day_of_month > 31:
                raise ValueError()
        except ValueError:
            flash('Día del mes inválido (1-31).', 'danger')
            return redirect(url_for('main.manage_recurring_tasks'))
    
    # Parse time_spent
    try:
        time_spent = int(time_spent_str) if time_spent_str else 0
    except ValueError:
        time_spent = 0
    
    # Determine area_id based on user
    area_id = None
    if not current_user.is_admin and current_user.areas:
        area_id = current_user.areas[0].id
    
    # Parse custom_dates for custom recurrence
    custom_dates = None
    if recurrence_type == 'custom':
        custom_dates_str = request.form.get('custom_dates', '')
        if custom_dates_str:
            custom_dates = custom_dates_str  # Already JSON string from frontend
    
    # Create recurring task
    # Use template values if template selected, otherwise use form values
    final_title = title if title else (selected_template.title if selected_template else 'Sin título')
    final_description = description if description else (selected_template.description if selected_template else '')
    final_priority = priority if priority != 'Normal' else (selected_template.priority if selected_template else 'Normal')
    
    new_recurring = RecurringTask(
        title=final_title,
        description=final_description,
        priority=final_priority,
        recurrence_type=recurrence_type,
        days_of_week=days_of_week if recurrence_type == 'weekly' else None,
        day_of_month=day_of_month,
        custom_dates=custom_dates,
        due_time=due_time,
        start_date=start_date,
        end_date=end_date,
        time_spent=time_spent,
        creator_id=current_user.id,
        is_active=True,
        area_id=area_id,
        template_id=template_id  # NEW: Link to template
    )
    
    # Add assignees
    for uid in assignee_ids:
        user = User.query.get(int(uid))
        if user:
            new_recurring.assignees.append(user)
    
    # Add tags
    for tid in tag_ids:
        tag = Tag.query.get(int(tid))
        if tag:
            new_recurring.tags.append(tag)
    
    db.session.add(new_recurring)
    db.session.commit()
    
    # Log activity
    log_activity(
        user=current_user,
        action='recurring_created',
        description=f'creó la tarea recurrente "{new_recurring.title}"',
        target_type='recurring_task',
        target_id=new_recurring.id,
        area_id=new_recurring.area_id
    )
    
    flash(f'Tarea recurrente "{title}" creada exitosamente.', 'success')
    return redirect(url_for('main.manage_recurring_tasks'))


@main_bp.route('/recurring-tasks/<int:rt_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_recurring_task(rt_id):
    """Editar una tarea recurrente"""
    # Check permission to create tasks
    if not current_user.can_create_tasks():
        flash('No tienes permiso para editar tareas recurrentes.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    rt = RecurringTask.query.get_or_404(rt_id)
    
    # Non-admins can only edit recurring tasks from their area
    if not current_user.can_see_all_areas():
        user_area_ids = [a.id for a in current_user.areas]
        if rt.area_id not in user_area_ids and rt.area_id is not None:
            flash('No tienes permiso para editar esta tarea recurrente.', 'danger')
            return redirect(url_for('main.manage_recurring_tasks'))
    
    if request.method == 'POST':
        rt.title = request.form.get('title', rt.title)
        rt.description = request.form.get('description', '')
        rt.priority = request.form.get('priority', 'Normal')
        rt.recurrence_type = request.form.get('recurrence_type', rt.recurrence_type)
        
        # Update days_of_week for weekly
        if rt.recurrence_type == 'weekly':
            rt.days_of_week = request.form.get('days_of_week', '')
        else:
            rt.days_of_week = None
        
        # Update day_of_month for monthly
        if rt.recurrence_type == 'monthly':
            day_str = request.form.get('day_of_month', '')
            rt.day_of_month = int(day_str) if day_str else None
        else:
            rt.day_of_month = None
        
        # Update custom_dates for custom recurrence
        if rt.recurrence_type == 'custom':
            rt.custom_dates = request.form.get('custom_dates', '')
        else:
            rt.custom_dates = None
        
        # Update time
        due_time_str = request.form.get('due_time')
        if due_time_str:
            from datetime import time as dt_time
            hour, minute = map(int, due_time_str.split(':'))
            rt.due_time = dt_time(hour, minute)
        
        # Update dates
        start_date_str = request.form.get('start_date')
        if start_date_str:
            rt.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        
        end_date_str = request.form.get('end_date', '')
        rt.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None
        
        # Update time_spent
        time_spent_str = request.form.get('time_spent', '0')
        rt.time_spent = int(time_spent_str) if time_spent_str else 0
        
        # Update assignees
        assignee_ids = request.form.getlist('assignees')
        rt.assignees = []
        for uid in assignee_ids:
            user = User.query.get(int(uid))
            if user:
                rt.assignees.append(user)
        
        # Update tags
        tag_ids = request.form.getlist('tags')
        rt.tags = []
        for tid in tag_ids:
            tag = Tag.query.get(int(tid))
            if tag:
                rt.tags.append(tag)
        
        db.session.commit()
        
        # Log activity
        log_activity(
            user=current_user,
            action='recurring_edited',
            description=f'editó la tarea recurrente "{rt.title}"',
            target_type='recurring_task',
            target_id=rt.id,
            area_id=rt.area_id
        )
        
        flash('Tarea recurrente actualizada.', 'success')
        return redirect(url_for('main.manage_recurring_tasks'))
    
    # GET - show edit form
    users = User.query.order_by(User.full_name).all()
    available_tags = Tag.query.order_by(Tag.name).all()
    return render_template('edit_recurring_task.html', rt=rt, users=users, available_tags=available_tags)


@main_bp.route('/recurring-tasks/<int:rt_id>/toggle', methods=['POST'])
@login_required
def toggle_recurring_task(rt_id):
    """Pausar/Reanudar una tarea recurrente"""
    # Check permission to create tasks
    if not current_user.can_create_tasks():
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    rt = RecurringTask.query.get_or_404(rt_id)
    
    # Non-admins can only toggle recurring tasks from their area
    if not current_user.can_see_all_areas():
        user_area_ids = [a.id for a in current_user.areas]
        if rt.area_id not in user_area_ids and rt.area_id is not None:
            return jsonify({'success': False, 'error': 'No autorizado para esta tarea'}), 403
    
    rt.is_active = not rt.is_active
    db.session.commit()
    
    # Log activity
    action = 'recurring_activated' if rt.is_active else 'recurring_paused'
    description = f'activó la tarea recurrente "{rt.title}"' if rt.is_active else f'pausó la tarea recurrente "{rt.title}"'
    log_activity(
        user=current_user,
        action=action,
        description=description,
        target_type='recurring_task',
        target_id=rt.id,
        area_id=rt.area_id
    )
    
    status = 'activada' if rt.is_active else 'pausada'
    return jsonify({
        'success': True,
        'is_active': rt.is_active,
        'message': f'Tarea recurrente {status}'
    })


@main_bp.route('/recurring-tasks/<int:rt_id>/delete', methods=['POST'])
@login_required
def delete_recurring_task(rt_id):
    """Eliminar una tarea recurrente - Solo admin"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    rt = RecurringTask.query.get_or_404(rt_id)
    title = rt.title
    db.session.delete(rt)
    db.session.commit()
    
    # Log activity
    log_activity(
        user=current_user,
        action='recurring_deleted',
        description=f'eliminó la tarea recurrente "{title}"',
        target_type='recurring_task',
        target_id=rt_id,
        area_id=rt.area_id
    )
    
    return jsonify({'success': True, 'message': f'Tarea recurrente "{title}" eliminada'})


@main_bp.route('/api/recurring-tasks/generate-now', methods=['POST'])
@login_required
def generate_recurring_tasks_now():
    """Forzar generación de tareas recurrentes ahora - Solo admin (para testing)"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    from flask import current_app
    from scheduler import generate_daily_tasks
    
    try:
        generate_daily_tasks(current_app._get_current_object())
        return jsonify({'success': True, 'message': 'Tareas generadas exitosamente'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/api/recurring-tasks/<int:rt_id>')
@login_required
def api_get_recurring_task(rt_id):
    """API para obtener datos de una tarea recurrente"""
    if not current_user.is_admin:
        return jsonify({'error': 'No autorizado'}), 403
    
    rt = RecurringTask.query.get_or_404(rt_id)
    
    return jsonify({
        'id': rt.id,
        'title': rt.title,
        'description': rt.description or '',
        'priority': rt.priority,
        'recurrence_type': rt.recurrence_type,
        'days_of_week': rt.days_of_week or '',
        'day_of_month': rt.day_of_month,
        'custom_dates': rt.custom_dates or '',
        'due_time': rt.due_time.strftime('%H:%M') if rt.due_time else '',
        'start_date': rt.start_date.strftime('%Y-%m-%d') if rt.start_date else '',
        'end_date': rt.end_date.strftime('%Y-%m-%d') if rt.end_date else '',
        'time_spent': rt.time_spent or 0,
        'is_active': rt.is_active,
        'assignee_ids': [u.id for u in rt.assignees],
        'tag_ids': [t.id for t in rt.tags]
    })


# --- Activity Log Routes ---

@main_bp.route('/activity-log')
@login_required
def activity_log():
    """Registro de actividades - solo admin y supervisores"""
    from models import Area
    
    # Only admins and supervisors can access
    if current_user.role not in ['admin', 'supervisor', 'gerente']:
        flash('No tienes permiso para acceder al registro de actividades.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Filters from query params
    filter_user = request.args.get('user')
    filter_action = request.args.get('action')
    filter_area = request.args.get('area')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Base query
    query = ActivityLog.query.options(joinedload(ActivityLog.user), joinedload(ActivityLog.area))
    
    # Area filtering based on role
    if current_user.is_admin or current_user.role == 'gerente':
        # Admin/Gerente can see all, and can filter by area
        if filter_area:
            query = query.filter(ActivityLog.area_id == int(filter_area))
        available_areas = Area.query.order_by(Area.name).all()
        show_area_filter = True
    else:
        # Supervisor only sees their area(s)
        user_area_ids = [a.id for a in current_user.areas]
        if user_area_ids:
            query = query.filter(ActivityLog.area_id.in_(user_area_ids))
        else:
            query = query.filter(ActivityLog.area_id == -1)  # No results
        available_areas = current_user.areas
        show_area_filter = False
    
    # User filter
    if filter_user:
        query = query.filter(ActivityLog.user_id == int(filter_user))
    
    # Action filter
    if filter_action:
        query = query.filter(ActivityLog.action == filter_action)
    
    # Date range filter
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            query = query.filter(db.func.date(ActivityLog.created_at) >= start_date)
        except ValueError:
            pass
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            query = query.filter(db.func.date(ActivityLog.created_at) <= end_date)
        except ValueError:
            pass
    
    # Order by most recent first, paginate
    logs = query.order_by(ActivityLog.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    # Get all users for filter dropdown
    all_users = User.query.filter(User.is_active == True).order_by(User.full_name).all()
    
    # Get unique action types for filter dropdown
    action_types = [
        ('task_created', 'Tarea Creada'),
        ('task_completed', 'Tarea Completada'),
        ('task_reopened', 'Tarea Reabierta'),
        ('task_edited', 'Tarea Editada'),
        ('task_anulada', 'Tarea Anulada'),
        ('expiration_created', 'Vencimiento Creado'),
        ('expiration_completed', 'Vencimiento Completado'),
        ('expiration_edited', 'Vencimiento Editado'),
        ('login', 'Inicio de Sesión'),
    ]
    
    today = date.today()
    
    return render_template('activity_log.html',
                          logs=logs,
                          all_users=all_users,
                          action_types=action_types,
                          all_areas=available_areas,
                          show_area_filter=show_area_filter,
                          current_filter_user=filter_user,
                          current_filter_action=filter_action,
                          current_filter_area=filter_area,
                          today=today)


# --- Process Type Management Routes ---

@main_bp.route('/process-types', methods=['GET', 'POST'])
@login_required
def manage_process_types():
    """View and create process types - Admin/Supervisor only"""
    from models import Area
    
    # Only admin and supervisors can manage process types
    if not current_user.is_admin and current_user.role != 'supervisor':
        flash('No tienes permiso para gestionar tipos de proceso.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get user's areas for filtering
    user_area_ids = [a.id for a in current_user.areas]
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        color = request.form.get('color', '#6366f1')
        icon = request.form.get('icon', 'fa-folder')
        template_id = request.form.get('template_id')
        area_id = request.form.get('area_id')
        
        if not name:
            flash('El nombre del tipo de proceso es requerido.', 'danger')
            return redirect(url_for('main.manage_process_types'))
        
        # Determine area_id
        if current_user.is_admin and area_id:
            area_id = int(area_id)
        elif current_user.areas:
            area_id = current_user.areas[0].id
        else:
            flash('No tienes un área asignada.', 'danger')
            return redirect(url_for('main.manage_process_types'))
        
        # Check for duplicate name in same area
        existing = ProcessType.query.filter_by(name=name, area_id=area_id).first()
        if existing:
            flash('Ya existe un tipo de proceso con ese nombre en esta área.', 'danger')
            return redirect(url_for('main.manage_process_types'))
        
        # Create process type
        process_type = ProcessType(
            name=name,
            description=description,
            color=color,
            icon=icon,
            area_id=area_id,
            created_by_id=current_user.id,
            template_id=int(template_id) if template_id else None
        )
        
        db.session.add(process_type)
        db.session.commit()
        
        # Log activity
        log_activity(
            user=current_user,
            action='process_type_created',
            description=f'creó el tipo de proceso "{name}"',
            target_type='process_type',
            target_id=process_type.id,
            area_id=area_id
        )
        
        flash(f'Tipo de proceso "{name}" creado exitosamente.', 'success')
        return redirect(url_for('main.manage_process_types'))
    
    # GET - List process types
    if current_user.is_admin:
        process_types = ProcessType.query.filter_by(is_active=True).order_by(ProcessType.name).all()
        templates = TaskTemplate.query.order_by(TaskTemplate.name).all()
        areas = Area.query.order_by(Area.name).all()
    else:
        if user_area_ids:
            process_types = ProcessType.query.filter(
                ProcessType.area_id.in_(user_area_ids),
                ProcessType.is_active == True
            ).order_by(ProcessType.name).all()
            templates = TaskTemplate.query.filter(TaskTemplate.area_id.in_(user_area_ids)).order_by(TaskTemplate.name).all()
        else:
            process_types = []
            templates = []
        areas = current_user.areas
    
    return render_template('manage_process_types.html', 
                           process_types=process_types,
                           templates=templates,
                           areas=areas,
                           show_area_selector=current_user.is_admin)


@main_bp.route('/process-types/<int:pt_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_process_type(pt_id):
    """Edit a process type"""
    from models import Area
    
    process_type = ProcessType.query.get_or_404(pt_id)
    
    # Check permissions
    user_area_ids = [a.id for a in current_user.areas]
    if not current_user.is_admin and process_type.area_id not in user_area_ids:
        flash('No tienes permiso para editar este tipo de proceso.', 'danger')
        return redirect(url_for('main.manage_process_types'))
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        color = request.form.get('color', '#6366f1')
        icon = request.form.get('icon', 'fa-folder')
        template_id = request.form.get('template_id')
        
        if not name:
            flash('El nombre del tipo de proceso es requerido.', 'danger')
            return redirect(url_for('main.edit_process_type', pt_id=pt_id))
        
        # Check for duplicate name in same area (excluding self)
        existing = ProcessType.query.filter(
            ProcessType.name == name,
            ProcessType.area_id == process_type.area_id,
            ProcessType.id != pt_id
        ).first()
        if existing:
            flash('Ya existe un tipo de proceso con ese nombre en esta área.', 'danger')
            return redirect(url_for('main.edit_process_type', pt_id=pt_id))
        
        process_type.name = name
        process_type.description = description
        process_type.color = color
        process_type.icon = icon
        process_type.template_id = int(template_id) if template_id else None
        
        db.session.commit()
        
        # Log activity
        log_activity(
            user=current_user,
            action='process_type_edited',
            description=f'editó el tipo de proceso "{name}"',
            target_type='process_type',
            target_id=process_type.id,
            area_id=process_type.area_id
        )
        
        flash(f'Tipo de proceso "{name}" actualizado exitosamente.', 'success')
        return redirect(url_for('main.manage_process_types'))
    
    # GET - Show edit form
    if current_user.is_admin:
        templates = TaskTemplate.query.order_by(TaskTemplate.name).all()
    else:
        templates = TaskTemplate.query.filter(TaskTemplate.area_id.in_(user_area_ids)).order_by(TaskTemplate.name).all()
    
    return render_template('edit_process_type.html', 
                           process_type=process_type,
                           templates=templates)


@main_bp.route('/process-types/<int:pt_id>/toggle', methods=['POST'])
@login_required
def toggle_process_type(pt_id):
    """Activate/Deactivate a process type"""
    process_type = ProcessType.query.get_or_404(pt_id)
    
    # Check permissions
    user_area_ids = [a.id for a in current_user.areas]
    if not current_user.is_admin and process_type.area_id not in user_area_ids:
        return jsonify({'success': False, 'error': 'No tienes permiso'}), 403
    
    # Check if there are active processes
    if process_type.is_active:
        active_processes = Process.query.filter_by(
            process_type_id=pt_id,
            status='Active'
        ).count()
        if active_processes > 0:
            return jsonify({
                'success': False, 
                'error': f'No se puede desactivar: hay {active_processes} proceso(s) activo(s)'
            }), 400
    
    process_type.is_active = not process_type.is_active
    db.session.commit()
    
    action = 'activó' if process_type.is_active else 'desactivó'
    log_activity(
        user=current_user,
        action='process_type_toggled',
        description=f'{action} el tipo de proceso "{process_type.name}"',
        target_type='process_type',
        target_id=process_type.id,
        area_id=process_type.area_id
    )
    
    return jsonify({
        'success': True,
        'is_active': process_type.is_active
    })


@main_bp.route('/process-types/<int:pt_id>/delete', methods=['POST'])
@login_required
def delete_process_type(pt_id):
    """Delete a process type - only if no processes exist"""
    process_type = ProcessType.query.get_or_404(pt_id)
    
    # Check permissions
    user_area_ids = [a.id for a in current_user.areas]
    if not current_user.is_admin and process_type.area_id not in user_area_ids:
        return jsonify({'success': False, 'error': 'No tienes permiso'}), 403
    
    # Check if there are any processes
    process_count = Process.query.filter_by(process_type_id=pt_id).count()
    if process_count > 0:
        return jsonify({
            'success': False,
            'error': f'No se puede eliminar: hay {process_count} proceso(s) asociados. Desactívalo en su lugar.'
        }), 400
    
    name = process_type.name
    area_id = process_type.area_id
    
    db.session.delete(process_type)
    db.session.commit()
    
    log_activity(
        user=current_user,
        action='process_type_deleted',
        description=f'eliminó el tipo de proceso "{name}"',
        target_type='process_type',
        target_id=pt_id,
        area_id=area_id
    )
    
    return jsonify({'success': True})


# --- Process Management Routes ---

@main_bp.route('/processes')
@login_required
def list_processes():
    """List all processes"""
    from models import Area
    
    # Get filter parameters
    filter_type = request.args.get('type')
    filter_status = request.args.get('status', 'Active')
    filter_area = request.args.get('area')
    
    user_area_ids = [a.id for a in current_user.areas]
    
    # Base query
    query = Process.query.options(
        joinedload(Process.process_type),
        joinedload(Process.area),
        joinedload(Process.created_by)
    )
    
    # Role-based filtering
    if current_user.can_see_all_areas():
        if filter_area:
            query = query.filter(Process.area_id == int(filter_area))
        available_areas = Area.query.order_by(Area.name).all()
        show_area_filter = True
    else:
        if user_area_ids:
            query = query.filter(Process.area_id.in_(user_area_ids))
        else:
            query = query.filter(Process.area_id == -1)
        available_areas = current_user.areas
        show_area_filter = False
    
    # Status filter
    if filter_status and filter_status != 'all':
        query = query.filter(Process.status == filter_status)
    
    # Type filter
    if filter_type:
        query = query.filter(Process.process_type_id == int(filter_type))
    
    # Order by due date
    processes = query.order_by(Process.due_date.asc()).all()
    
    # Get process types for filter dropdown
    if current_user.is_admin:
        process_types = ProcessType.query.filter_by(is_active=True).order_by(ProcessType.name).all()
    else:
        process_types = ProcessType.query.filter(
            ProcessType.area_id.in_(user_area_ids),
            ProcessType.is_active == True
        ).order_by(ProcessType.name).all()
    
    # Calculate stats
    active_count = sum(1 for p in processes if p.status == 'Active')
    completed_count = sum(1 for p in processes if p.status == 'Completed')
    
    return render_template('processes.html',
                           processes=processes,
                           process_types=process_types,
                           filter_type=filter_type,
                           filter_status=filter_status,
                           filter_area=filter_area,
                           areas=available_areas,
                           show_area_filter=show_area_filter,
                           active_count=active_count,
                           completed_count=completed_count)


@main_bp.route('/processes/create', methods=['GET', 'POST'])
@login_required
def create_process():
    """Create a new process"""
    from models import Area
    
    # Only supervisor+ can create processes
    if not current_user.can_create_tasks():
        flash('No tienes permiso para crear procesos.', 'danger')
        return redirect(url_for('main.list_processes'))
    
    user_area_ids = [a.id for a in current_user.areas]
    
    if request.method == 'POST':
        process_type_id = request.form.get('process_type_id')
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        due_date_str = request.form.get('due_date')
        due_time_str = request.form.get('due_time', '18:00')
        
        if not process_type_id or not name or not due_date_str:
            flash('Tipo de proceso, nombre y fecha límite son requeridos.', 'danger')
            return redirect(url_for('main.create_process'))
        
        # Get process type
        process_type = ProcessType.query.get_or_404(int(process_type_id))
        
        # Check permission for this area
        if not current_user.is_admin and process_type.area_id not in user_area_ids:
            flash('No tienes permiso para crear procesos de este tipo.', 'danger')
            return redirect(url_for('main.create_process'))
        
        # Parse due date
        try:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
            if due_time_str:
                due_time = datetime.strptime(due_time_str, '%H:%M').time()
                due_date = datetime.combine(due_date.date(), due_time)
        except ValueError:
            flash('Fecha inválida.', 'danger')
            return redirect(url_for('main.create_process'))
        
        # Create process
        process = Process(
            process_type_id=process_type.id,
            name=name,
            description=description,
            area_id=process_type.area_id,
            due_date=due_date,
            created_by_id=current_user.id,
            status='Active'
        )
        
        db.session.add(process)
        db.session.flush()  # Get process ID
        
        # If process type has template, create tasks automatically
        tasks_created = 0
        if process_type.template:
            template = process_type.template
            
            # Get assignees from form
            assignee_ids = request.form.getlist('assignees')
            assignees = [User.query.get(int(uid)) for uid in assignee_ids if uid]
            assignees = [a for a in assignees if a]
            
            # Create main task from template
            main_task = Task(
                title=template.title,
                description=template.description or description,
                priority=template.priority,
                status='Pending',
                creator_id=current_user.id,
                area_id=process_type.area_id,
                due_date=due_date,
                process_id=process.id,
                planned_start_date=datetime.now()  # Start now
            )
            
            for tag in template.tags:
                main_task.tags.append(tag)
            for assignee in assignees:
                main_task.assignees.append(assignee)
            
            db.session.add(main_task)
            db.session.flush()
            tasks_created = 1
            
            # Create subtasks from template
            if template.subtask_templates.count() > 0:
                create_subtasks_from_template(template, main_task, assignees, current_user, process_type.area_id)
                tasks_created += template.subtask_templates.count()
                
                # Associate all created subtasks with the process
                for child in main_task.children:
                    child.process_id = process.id
        
        db.session.commit()
        
        # Log activity
        log_activity(
            user=current_user,
            action='process_created',
            description=f'creó el proceso "{name}"' + (f' con {tasks_created} tarea(s)' if tasks_created > 0 else ''),
            target_type='process',
            target_id=process.id,
            area_id=process.area_id
        )
        
        flash(f'Proceso "{name}" creado exitosamente' + (f' con {tasks_created} tarea(s).' if tasks_created > 0 else '.'), 'success')
        return redirect(url_for('main.process_details', process_id=process.id))
    
    # GET - Show create form
    if current_user.is_admin:
        process_types = ProcessType.query.filter_by(is_active=True).order_by(ProcessType.name).all()
        users = User.query.all()
    else:
        process_types = ProcessType.query.filter(
            ProcessType.area_id.in_(user_area_ids),
            ProcessType.is_active == True
        ).order_by(ProcessType.name).all()
        users = [u for u in User.query.all() if any(area in u.areas for area in current_user.areas)]
    
    # Default due date: 7 days from now
    default_due_date = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    
    return render_template('create_process.html',
                           process_types=process_types,
                           users=users,
                           default_due_date=default_due_date)


@main_bp.route('/processes/<int:process_id>')
@login_required
def process_details(process_id):
    """View process details"""
    from models import Area
    
    process = Process.query.options(
        joinedload(Process.process_type),
        joinedload(Process.area),
        joinedload(Process.created_by),
        joinedload(Process.completed_by)
    ).get_or_404(process_id)
    
    # Check permission - allow current area OR involved areas (read-only)
    user_area_ids = [a.id for a in current_user.areas]
    involved_area_ids = [a.id for a in process.involved_areas]
    can_view = (current_user.is_admin or 
                process.area_id in user_area_ids or 
                any(aid in involved_area_ids for aid in user_area_ids))
    
    if not can_view:
        flash('No tienes permiso para ver este proceso.', 'danger')
        return redirect(url_for('main.list_processes'))
    
    # Get tasks in this process with necessary relationships loaded
    tasks = Task.query.filter_by(process_id=process_id)\
        .options(
            joinedload(Task.assignees),
            subqueryload(Task.children),
            subqueryload(Task.status_history).joinedload(StatusTransition.changed_by)
        )\
        .order_by(Task.created_at.asc())\
        .all()
    
    # Get all areas for transfer modal
    all_areas = Area.query.order_by(Area.name).all()
    
    # Get process events (history)
    events = process.events.all()
    
    return render_template('process_details.html',
                           process=process,
                           tasks=tasks,
                           all_areas=all_areas,
                           events=events,
                           now_utc=now_utc)


@main_bp.route('/processes/<int:process_id>/cancel', methods=['POST'])
@login_required
def cancel_process(process_id):
    """Cancel a process and annul all its tasks"""
    process = Process.query.get_or_404(process_id)
    
    # Check permission
    user_area_ids = [a.id for a in current_user.areas]
    if not current_user.is_admin and process.area_id not in user_area_ids:
        return jsonify({'success': False, 'error': 'No tienes permiso'}), 403
    
    if process.status != 'Active':
        return jsonify({'success': False, 'error': 'Solo se pueden cancelar procesos activos'}), 400
    
    # Cancel process and annul tasks
    process.cancel_with_tasks(current_user)
    db.session.commit()
    
    # Log activity
    log_activity(
        user=current_user,
        action='process_cancelled',
        description=f'canceló el proceso "{process.name}" y anuló sus tareas',
        target_type='process',
        target_id=process.id,
        area_id=process.area_id
    )
    
    return jsonify({'success': True})


@main_bp.route('/processes/<int:process_id>/complete', methods=['POST'])
@login_required
def complete_process(process_id):
    """Manually complete a process"""
    process = Process.query.get_or_404(process_id)
    
    # Check permission - only current area can complete
    user_area_ids = [a.id for a in current_user.areas]
    if not current_user.is_admin and process.area_id not in user_area_ids:
        return jsonify({'success': False, 'error': 'No tienes permiso'}), 403
    
    if process.status != 'Active':
        return jsonify({'success': False, 'error': 'Solo se pueden completar procesos activos'}), 400
    
    # Check for pending tasks
    pending_tasks = process.tasks.filter(Task.status.in_(['Pending', 'In Progress', 'In Review'])).count()
    
    data = request.get_json() or {}
    force = data.get('force', False)
    
    if pending_tasks > 0 and not force:
        return jsonify({
            'success': False,
            'needs_confirmation': True,
            'pending_count': pending_tasks,
            'error': f'Hay {pending_tasks} tarea(s) sin completar en este proceso. Si tu área ya terminó su trabajo, deberías usar "Pasar a otra Área" en lugar de completar.'
        }), 400
    
    # Complete the process
    process.status = 'Completed'
    process.completed_at = now_utc()
    process.completed_by_id = current_user.id
    
    # Log process event for unified history
    log_process_event(
        process_id=process.id,
        event_type='process_completed',
        description=f'Proceso completado por {current_user.full_name}',
        user_id=current_user.id
    )
    
    db.session.commit()
    
    # Log activity
    log_activity(
        user=current_user,
        action='process_completed',
        description=f'completó el proceso "{process.name}"',
        target_type='process',
        target_id=process.id,
        area_id=process.area_id
    )
    
    return jsonify({'success': True})


@main_bp.route('/processes/<int:process_id>/transfer', methods=['POST'])
@login_required
def transfer_process(process_id):
    """Transfer a process to another area, keeping visibility for involved areas."""
    from models import Process, Area, ProcessTransfer
    
    # Only admin and supervisor can transfer
    if not current_user.is_admin and current_user.role != 'supervisor':
        return jsonify({'success': False, 'error': 'No tienes permiso para transferir procesos'}), 403
    
    process = Process.query.get_or_404(process_id)
    
    # Check if user has access to current area
    if not current_user.is_admin:
        user_area_ids = [a.id for a in current_user.areas]
        if process.area_id not in user_area_ids:
            return jsonify({'success': False, 'error': 'Solo puedes transferir procesos de tu área'}), 403
    
    data = request.get_json() or {}
    to_area_id = data.get('to_area_id')
    comment = data.get('comment', '')
    
    if not to_area_id:
        return jsonify({'success': False, 'error': 'Debe seleccionar un área destino'}), 400
    
    to_area = Area.query.get(to_area_id)
    if not to_area:
        return jsonify({'success': False, 'error': 'Área destino no encontrada'}), 404
    
    if process.area_id == to_area_id:
        return jsonify({'success': False, 'error': 'El proceso ya está en esa área'}), 400
    
    # Store current area in involved_areas (for read-only access)
    current_area = process.area
    if current_area not in process.involved_areas:
        process.involved_areas.append(current_area)
    
    # Create transfer record
    transfer = ProcessTransfer(
        process_id=process.id,
        from_area_id=process.area_id,
        to_area_id=to_area_id,
        transferred_by_id=current_user.id,
        comment=comment
    )
    db.session.add(transfer)
    
    # Update process area
    old_area_name = process.area.name
    process.area_id = to_area_id
    
    # If process was completed, reopen it - a transfer means more work is needed
    if process.status == 'Completed':
        process.status = 'Active'
        process.completed_at = None
        process.completed_by_id = None
    
    # Log process event for unified history
    comment_suffix = f' - "{comment}"' if comment else ''
    log_process_event(
        process_id=process.id,
        event_type='transfer',
        description=f'Proceso transferido de {old_area_name} a {to_area.name}{comment_suffix}',
        user_id=current_user.id
    )
    
    db.session.commit()
    
    # Log activity
    log_activity(
        user=current_user,
        action='process_transferred',
        description=f'transfirió el proceso "{process.name}" de {old_area_name} a {to_area.name}',
        target_type='process',
        target_id=process.id,
        area_id=to_area_id
    )
    
    return jsonify({
        'success': True,
        'from_area': old_area_name,
        'to_area': to_area.name
    })


# --- Helper for Template Subtasks (Appended) ---
def create_subtasks_from_template(template, parent_task, assignees=None, creator=None, area_id=None):
    """
    Creates subtasks based on a template, handling hierarchy and start times.
    Appended to ensure latest version is used.
    """
    from models import Task, SubtaskTemplate, User
    from extensions import db
    from datetime import datetime, timedelta
    
    # Get subtask templates
    st_templates = SubtaskTemplate.query.filter_by(template_id=template.id).order_by(SubtaskTemplate.order).all()
    
    # Map for hierarchy
    id_to_task = {}
    
    # 1. Create all subtasks
    for st in st_templates:
        # Determine start datetime
        subtask_start_date = None
        if parent_task.planned_start_date:
            # Base date: Parent Start Date + Offset
            offset = st.start_days_offset or 0
            base_date = parent_task.planned_start_date + timedelta(days=offset)
            
            # Determine time
            if st.start_time:
                # Use specific time
                subtask_start_date = datetime.combine(base_date.date(), st.start_time)
            else:
                # Inherit time from parent
                subtask_start_date = datetime.combine(base_date.date(), parent_task.planned_start_date.time())
        else:
            # Fallback if parent has no start date
            pass
            
        # Determine due date (based on existing logic for days_offset)
        subtask_due_date = parent_task.due_date
        if st.days_offset is not None and parent_task.due_date:
             try:
                subtask_due_date = parent_task.due_date + timedelta(days=st.days_offset)
             except:
                pass
        
        # Create Task
        subtask = Task(
             title=st.title,
             description=st.description,
             priority=st.priority,
             status='Pending',
             creator_id=creator.id if creator else parent_task.creator_id,
             area_id=area_id or parent_task.area_id,
             parent_id=parent_task.id, # Temp, will be updated in pass 2
             planned_start_date=subtask_start_date,
             due_date=subtask_due_date,
             enabled=False, # Initially blocked
             process_id=parent_task.process_id  # Inherit process from parent
        )
        
        # Tags
        for tag in parent_task.tags:
             subtask.tags.append(tag)
        
        # Assignees (from arg)
        if assignees:
             for assignee in assignees:
                 subtask.assignees.append(assignee)
                 
        db.session.add(subtask)
        db.session.flush()
        id_to_task[st.id] = subtask
        
    # 2. Link parents and Enabled status
    for st in st_templates:
        child = id_to_task[st.id]
        if st.parent_id and st.parent_id in id_to_task:
            parent = id_to_task[st.parent_id]
            child.parent_id = parent.id
            # Child blocked by subtask parent
            child.enabled = False 
        else:
            child.parent_id = parent_task.id
            # Top level subtask - enable if parent is enabled/pending
            child.enabled = parent_task.enabled if parent_task.enabled else True
            
    db.session.commit()


@main_bp.route('/download_import_template')
@login_required
def download_import_template():
    """Download the Excel template for importing tasks."""
    wb = generate_import_template()
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_importacion_tareas.xlsx'
    )


@main_bp.route('/import_tasks', methods=['POST'])
@login_required
def import_tasks():
    """Import tasks from an Excel file."""
    # Check if a file was uploaded
    if 'file' not in request.files:
        flash('No se seleccionó ningún archivo', 'danger')
        return redirect(url_for('main.dashboard'))
    
    file = request.files['file']
    
    # Check if user selected a file
    if file.filename == '':
        flash('No se seleccionó ningún archivo', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get selected area_id from form
    area_id = request.form.get('area_id', type=int)

    # Process the file
    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        success_count, errors = process_excel_import(file, current_user, area_id=area_id)
        
        if success_count > 0:
            flash(f'Se importaron exitosamente {success_count} tarea(s).', 'success')
            
        if errors:
            # If many errors, maybe just show first 5?
            error_msg = '<br>'.join(errors[:5])
            if len(errors) > 5:
                error_msg += f'<br>... y {len(errors) - 5} errores más.'
            flash(f'Errores durante la importación:<br>{error_msg}', 'warning')
            
        if success_count == 0 and not errors:
             flash('No se encontraron tareas válidas para importar.', 'warning')

    else:
        flash('Formato de archivo inválido. Por favor sube un Excel (.xlsx)', 'danger')
        
    return redirect(url_for('main.dashboard'))


# =====================================================
# FILE ATTACHMENTS (MinIO)
# =====================================================

@main_bp.route('/tasks/<int:task_id>/attachments', methods=['POST'])
@login_required
def upload_attachment(task_id):
    """Upload a file attachment to a task."""
    print(f"=== INICIO upload_attachment para task_id={task_id} ===")
    task = Task.query.get_or_404(task_id)
    print(f"DEBUG: Tarea encontrada: {task.title}")
    
    # Check if user has access to this task
    if not current_user.is_admin and current_user not in task.assignees and task.creator_id != current_user.id:
        print(f"DEBUG: Usuario sin permiso: {current_user.username}")
        flash('No tienes permiso para adjuntar archivos a esta tarea.', 'danger')
        return redirect(url_for('main.task_details', task_id=task_id))
    
    print(f"DEBUG: request.files keys: {list(request.files.keys())}")
    if 'file' not in request.files:
        print("DEBUG: 'file' no está en request.files")
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('main.task_details', task_id=task_id))
    
    file = request.files['file']
    print(f"DEBUG: Archivo recibido: filename={file.filename}, content_type={file.content_type}")
    
    if file.filename == '':
        print("DEBUG: filename está vacío")
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('main.task_details', task_id=task_id))
    
    # Check file extension
    if not storage.allowed_file(file.filename):
        print(f"DEBUG: Extensión no permitida para: {file.filename}")
        flash(f'Tipo de archivo no permitido. Extensiones válidas: {", ".join(storage.ALLOWED_EXTENSIONS)}', 'danger')
        return redirect(url_for('main.task_details', task_id=task_id))
    
    # Check file size
    file.seek(0, 2)  # Seek to end
    file_size = file.tell()
    file.seek(0)  # Reset to start
    print(f"DEBUG: Tamaño del archivo: {file_size} bytes ({file_size / (1024*1024):.2f} MB)")
    
    if file_size > storage.MAX_FILE_SIZE:
        print(f"DEBUG: Archivo muy grande. Máximo: {storage.MAX_FILE_SIZE}")
        flash(f'El archivo excede el tamaño máximo de {storage.MAX_FILE_SIZE / (1024*1024):.0f} MB.', 'danger')
        return redirect(url_for('main.task_details', task_id=task_id))
    
    # Secure the filename
    filename = secure_filename(file.filename)
    print(f"DEBUG: Filename seguro: {filename}")
    
    # Upload to MinIO
    print(f"DEBUG: Llamando a storage.upload_file...")
    success, result = storage.upload_file(file, task_id, filename)
    print(f"DEBUG: Resultado de upload_file: success={success}, result={result}")
    
    if not success:
        print(f"DEBUG: Error al subir a MinIO: {result}")
        flash(f'Error al subir archivo: {result}', 'danger')
        return redirect(url_for('main.task_details', task_id=task_id))
    
    # Save metadata to database
    print("DEBUG: Guardando metadata en base de datos...")
    attachment = TaskAttachment(
        task_id=task_id,
        filename=filename,
        file_key=result,  # This is the file_key returned from upload_file
        file_size=file_size,
        content_type=file.content_type,
        uploaded_by_id=current_user.id
    )
    db.session.add(attachment)
    
    # Log activity
    activity = ActivityLog(
        user_id=current_user.id,
        action='attachment_upload',
        description=f'Subió el archivo "{filename}" a la tarea #{task_id}',
        target_type='task',
        target_id=task_id,
        area_id=task.area_id
    )
    db.session.add(activity)
    db.session.commit()
    print(f"DEBUG: Archivo guardado exitosamente. Attachment ID: {attachment.id}")
    
    flash(f'Archivo "{filename}" subido y guardado exitosamente.', 'success')
    print("=== FIN upload_attachment (éxito) ===")
    return redirect(url_for('main.edit_task', task_id=task_id))


@main_bp.route('/attachments/<int:attachment_id>/download')
@login_required
def download_attachment(attachment_id):
    """Download a file attachment."""
    attachment = TaskAttachment.query.get_or_404(attachment_id)
    task = attachment.task
    
    # Check if user has access to this task
    if not current_user.is_admin and current_user not in task.assignees and task.creator_id != current_user.id:
        flash('No tienes permiso para descargar este archivo.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Download from MinIO
    success, result = storage.download_file(attachment.file_key)
    
    if not success:
        flash(f'Error al descargar archivo: {result}', 'danger')
        return redirect(url_for('main.task_details', task_id=task.id))
    
    # Create response with file data
    response = make_response(result['data'])
    response.headers['Content-Type'] = result['content_type']
    response.headers['Content-Disposition'] = f'attachment; filename="{attachment.filename}"'
    
    return response


@main_bp.route('/attachments/<int:attachment_id>/delete', methods=['POST'])
@login_required
def delete_attachment(attachment_id):
    """Delete a file attachment."""
    attachment = TaskAttachment.query.get_or_404(attachment_id)
    task = attachment.task
    
    # Check permission: only admin, the uploader, or task creator can delete
    if not current_user.is_admin and attachment.uploaded_by_id != current_user.id and task.creator_id != current_user.id:
        flash('No tienes permiso para eliminar este archivo.', 'danger')
        return redirect(url_for('main.task_details', task_id=task.id))
    
    # Delete from MinIO
    success, message = storage.delete_file(attachment.file_key)
    
    if not success:
        flash(f'Error al eliminar archivo: {message}', 'warning')
        # Continue to delete from DB anyway
    
    # Log activity
    activity = ActivityLog(
        user_id=current_user.id,
        action='attachment_delete',
        description=f'Eliminó el archivo "{attachment.filename}" de la tarea #{task.id}',
        target_type='task',
        target_id=task.id,
        area_id=task.area_id
    )
    db.session.add(activity)
    
    # Delete from database
    db.session.delete(attachment)
    db.session.commit()
    
    flash(f'Archivo "{attachment.filename}" eliminado.', 'success')
    return redirect(url_for('main.edit_task', task_id=task.id))


@main_bp.route('/attachments/<int:attachment_id>/view')
@login_required
def view_attachment(attachment_id):
    """View a file attachment in browser (for PDFs and images)."""
    attachment = TaskAttachment.query.get_or_404(attachment_id)
    task = attachment.task
    
    # Check if user has access to this task
    if not current_user.is_admin and current_user not in task.assignees and task.creator_id != current_user.id:
        flash('No tienes permiso para ver este archivo.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Download from MinIO
    success, result = storage.download_file(attachment.file_key)
    
    if not success:
        flash(f'Error al cargar archivo: {result}', 'danger')
        return redirect(url_for('main.task_details', task_id=task.id))
    
    # Create response - display inline instead of download
    response = make_response(result['data'])
    response.headers['Content-Type'] = result['content_type']
    response.headers['Content-Disposition'] = f'inline; filename="{attachment.filename}"'
    
    return response
